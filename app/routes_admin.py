import base64
import csv
import json
from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO, StringIO
import mimetypes
from pathlib import Path
import re
import secrets

from sqlalchemy import or_
from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from app.auth import (
    attach_current_user,
    can_admin_page,
    can_edit,
    can_manage_users,
    can_read_page,
    can_review,
    can_view_analytics,
    can_view_audit_logs,
    can_view_security,
    can_write_page,
    client_ip,
    get_page_permission,
    get_csrf_token,
    login_required,
    validate_csrf_token,
)
from app.models import (
    CONTENT_STATUS_DRAFT,
    CONTENT_STATUS_OFFLINE,
    CONTENT_STATUS_PUBLISHED,
    FEEDBACK_APPROVED,
    FEEDBACK_PENDING,
    FEEDBACK_REJECTED,
    MERCH_BATCH_ACTIVE,
    MERCH_BATCH_ENDED,
    MERCH_BATCH_UPCOMING,
    MERCH_ORDER_CANCELLED,
    MERCH_ORDER_CONFIRMED,
    MERCH_ORDER_PENDING,
    MERCH_ORDER_PICKED_UP,
    MERCH_ORDER_READY,
    MERCH_ORDER_STATUSES,
    PAGE_ACCOUNTS,
    PAGE_ACTIVITIES,
    PAGE_ANALYTICS,
    PAGE_ANNOUNCEMENTS,
    PAGE_AUDIT_LOGS,
    PAGE_FEEDBACK,
    PAGE_KEYS,
    PAGE_KIT_PREORDERS,
    PAGE_PERMISSION_LEVELS,
    PAGE_ROUTES,
    PAGE_SECURITY,
    PERMISSION_ADMIN,
    PERMISSION_NONE,
    PERMISSION_READ,
    PERMISSION_WRITE,
    REGISTRATION_CONFIRMED,
    REGISTRATION_PENDING,
    ROLE_CONTENT_ADMIN,
    ROLE_OPS_ADMIN,
    ROLE_PAGE_PERMISSION_PRESETS,
    ROLE_SUPER_ADMIN,
    ROLE_VIEWER,
    ROLES,
    ROUTE_STATUSES,
    SITE_FEEDBACK_DONE,
    SITE_FEEDBACK_PENDING,
    STATUS_DRAFT,
    STATUS_OFFLINE,
    STATUS_PENDING_REVIEW,
    STATUS_PUBLISHED,
    Activity,
    ActivityRouteOption,
    AccessLog,
    AuditLog,
    Announcement,
    EventRegistration,
    ImportReport,
    MediaAsset,
    MerchPreorderBatch,
    MerchPreorderImage,
    MerchPreorderRegistration,
    Route,
    RouteFeedback,
    RouteVersion,
    SiteFeedback,
    User,
    UserPagePermission,
    db,
    merch_batch_status_for_window,
    utcnow,
)
from app.querying import query_routes_from_request
from app.gpx_utils import parse_gpx_points_and_stats, parse_gpx_waypoints
from app.route_ops import allowed_file, file_size_ok, parse_distance, save_gpx_file
from app.security_monitor import WATCHLIST_PROBE_PATHS, build_non_probe_filters
from app.security_limits import check_lock, clear_state, register_failure
from app.services import (
    approved_rating_summary,
    create_route_version,
    ensure_user_page_permissions,
    rollback_route_to_version,
    route_snapshot,
    save_import_report,
    write_audit_log,
    write_field_audit_log,
)

bp = Blueprint("admin", __name__, url_prefix="/manage")
SH_TZ = timezone(timedelta(hours=8))
LOGIN_WINDOW_SECONDS = 15 * 60
LOGIN_MAX_FAILURES = 5
LOGIN_LOCK_SECONDS = 15 * 60
ROLE_LABELS = {
    ROLE_SUPER_ADMIN: "超级管理员",
    ROLE_OPS_ADMIN: "安全运维",
    ROLE_CONTENT_ADMIN: "内容管理员",
    ROLE_VIEWER: "只读观察员",
}
PAGE_PERMISSION_LABELS = {
    PAGE_ROUTES: "路线",
    PAGE_ACTIVITIES: "活动",
    PAGE_KIT_PREORDERS: "预定",
    PAGE_ANNOUNCEMENTS: "公告",
    PAGE_FEEDBACK: "反馈",
    PAGE_ANALYTICS: "流量",
    PAGE_SECURITY: "安全",
    PAGE_ACCOUNTS: "账号",
    PAGE_AUDIT_LOGS: "审计日志",
}
PERMISSION_LEVEL_LABELS = {
    PERMISSION_NONE: "无权限",
    PERMISSION_READ: "可读",
    PERMISSION_WRITE: "可修改",
    PERMISSION_ADMIN: "完全管理",
}
ACTIVITY_ROUTE_LEVELS = (
    ("beginner", "初级"),
    ("intermediate", "中级"),
    ("advanced", "高级"),
)
MERCH_BATCH_STATUS_LABELS = {
    MERCH_BATCH_UPCOMING: "即将开始",
    MERCH_BATCH_ACTIVE: "正在进行",
    MERCH_BATCH_ENDED: "已经结束",
}
MERCH_ORDER_STATUS_LABELS = {
    MERCH_ORDER_PENDING: "待确认",
    MERCH_ORDER_CONFIRMED: "已确认",
    MERCH_ORDER_CANCELLED: "已取消",
    MERCH_ORDER_READY: "待领取",
    MERCH_ORDER_PICKED_UP: "已领取",
}
MERCH_ACTIVE_ORDER_STATUSES = (
    MERCH_ORDER_PENDING,
    MERCH_ORDER_CONFIRMED,
    MERCH_ORDER_READY,
    MERCH_ORDER_PICKED_UP,
)
AUDIT_ACTION_LABELS = {
    "auth.login": "登录后台",
    "auth.logout": "退出后台",
    "route.create": "创建路线",
    "route.update": "更新路线",
    "route.soft_delete": "移入路线回收站",
    "route.restore": "恢复路线",
    "route.status": "更新路线状态",
    "route.rollback": "回滚路线版本",
    "route.bulk_import": "批量导入路线",
    "feedback.create": "提交路线反馈",
    "feedback.review": "审核路线反馈",
    "feedback.reopen": "重开路线反馈",
    "feedback.delete": "删除路线反馈",
    "activity.create": "创建活动",
    "activity.update": "更新活动",
    "activity.delete": "删除活动",
    "activity.media.assign": "绑定活动媒体",
    "activity.media.delete": "删除活动媒体",
    "merch_preorder.create": "创建预定批次",
    "merch_preorder.update": "更新预定批次",
    "merch_preorder.visibility": "更新预定批次展示状态",
    "merch_preorder.delete": "删除预定批次",
    "merch_preorder.image.delete": "删除预定图片",
    "merch_preorder.registration.status": "更新预报名状态",
    "announcement.create": "创建公告",
    "announcement.update": "更新公告",
    "announcement.status": "更新公告状态",
    "announcement.delete": "删除公告",
    "site_feedback.status_update": "更新网站反馈",
    "user.create": "创建账号",
    "user.update": "更新账号",
    "user.deactivate": "停用账号",
}


def _parse_registration_notes(raw_notes: str) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in (raw_notes or "").split(";"):
        chunk = item.strip()
        if not chunk or "=" not in chunk:
            continue
        key, value = chunk.split("=", 1)
        result[key.strip()] = value.strip()
    return result


def _excel_safe_sheet_name(raw: str, used: set[str]) -> str:
    base = (raw or "未指定路线").strip()
    base = re.sub(r'[:\\/?*\[\]]', "_", base)
    if not base:
        base = "未指定路线"
    base = base[:31]
    candidate = base
    index = 2
    while candidate in used:
        suffix = f"_{index}"
        keep = max(1, 31 - len(suffix))
        candidate = f"{base[:keep]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def _display_app_version() -> str:
    raw = str(current_app.config.get("APP_VERSION", "") or "").strip()
    if not raw:
        return "unknown"
    if re.fullmatch(r"\d+\.\d+\.\d+", raw):
        return f"v{raw}"
    return raw


def _audit_action_label(action: str | None) -> str:
    raw = (action or "").strip()
    if not raw:
        return "记录操作"
    return AUDIT_ACTION_LABELS.get(raw, raw)


def _to_local_time(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(SH_TZ)


def _deployed_at_utc() -> datetime | None:
    raw = (current_app.config.get("APP_DEPLOYED_AT") or "").strip()
    if not raw:
        return None
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=SH_TZ)
    return parsed.astimezone(timezone.utc)


def _resolve_period_start(days: int, scope: str, now: datetime) -> tuple[str, datetime, str]:
    deployed = _deployed_at_utc()
    if scope == "post_deploy" and deployed is not None:
        local = _to_local_time(deployed)
        return "post_deploy", deployed, f"上线后（{local.strftime('%Y-%m-%d %H:%M')}）"
    start = now - timedelta(days=days)
    return "recent", start, f"最近 {days} 天"


def _watchlist_probe_expression():
    conditions = [AccessLog.path.in_(WATCHLIST_PROBE_PATHS), AccessLog.path.like("/wordpress/wp-admin/%")]
    return or_(*conditions)


def _probe_expression():
    suffixes = (".php", ".asp", ".aspx", ".jsp", ".bak", ".sql", ".zip", ".tar.gz")
    prefixes = ("/wp-", "/wordpress", "/xmlrpc.php", "/phpmyadmin", "/pma", "/.git", "/.env", "/vendor", "/cgi-bin")
    conditions = [AccessLog.path.like(f"{prefix}%") for prefix in prefixes]
    conditions.extend([AccessLog.path.like(f"%{suffix}") for suffix in suffixes])
    conditions.append(AccessLog.path.like("%../%"))
    conditions.append(AccessLog.path.ilike("%2e%2e%"))
    conditions.append(_watchlist_probe_expression())
    return or_(*conditions)


def _default_page_permissions_for_role(role: str) -> dict[str, str]:
    role = (role or "").strip()
    preset = ROLE_PAGE_PERMISSION_PRESETS.get(role, {})
    return {page_key: preset.get(page_key, PERMISSION_NONE) for page_key in PAGE_KEYS}


def _page_permissions_from_form(role: str) -> dict[str, str]:
    defaults = _default_page_permissions_for_role(role)
    result: dict[str, str] = {}
    for page_key in PAGE_KEYS:
        raw = (request.form.get(f"page_perm_{page_key}") or defaults.get(page_key) or PERMISSION_NONE).strip()
        result[page_key] = raw if raw in PAGE_PERMISSION_LEVELS else PERMISSION_NONE
    return result


def _user_page_permissions(user: User) -> dict[str, str]:
    existing = {item.page_key: item.permission_level for item in user.page_permissions}
    return {page_key: existing.get(page_key, PERMISSION_NONE) for page_key in PAGE_KEYS}


def _apply_user_page_permissions(user: User, permissions: dict[str, str]) -> None:
    existing = {item.page_key: item for item in user.page_permissions}
    for page_key in PAGE_KEYS:
        level = permissions.get(page_key, PERMISSION_NONE)
        if level not in PAGE_PERMISSION_LEVELS:
            level = PERMISSION_NONE
        record = existing.get(page_key)
        if record is None:
            db.session.add(UserPagePermission(user=user, page_key=page_key, permission_level=level))
        else:
            record.permission_level = level


def _save_activity_media(activity_id: int, uploads, activity_route_option_id: int | None = None) -> int:
    media_dir = Path(current_app.config["MEDIA_UPLOAD_FOLDER"])
    media_dir.mkdir(parents=True, exist_ok=True)
    allowed_exts = {str(item).lower() for item in current_app.config.get("ALLOWED_MEDIA_EXTENSIONS", set())}
    max_media_bytes = int(current_app.config.get("MAX_MEDIA_BYTES", 10 * 1024 * 1024))

    saved_count = 0
    for upload in uploads:
        if not upload or not (upload.filename or "").strip():
            continue
        original_name = secure_filename(Path(upload.filename).name)
        ext = Path(original_name).suffix.lower()
        if not original_name or ext not in allowed_exts:
            continue
        if not file_size_ok(upload, max_media_bytes):
            continue

        token = f"activity_{activity_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{secrets.token_hex(4)}{ext}"
        target_path = media_dir / token
        upload.save(target_path)
        mime_type = (upload.mimetype or "").strip() or (mimetypes.guess_type(original_name)[0] or "application/octet-stream")
        db.session.add(
            MediaAsset(
                activity_id=activity_id,
                activity_route_option_id=activity_route_option_id,
                original_filename=original_name[:255],
                storage_path=token,
                mime_type=mime_type[:128],
                size_bytes=int(target_path.stat().st_size),
                created_at=utcnow(),
            )
        )
        saved_count += 1
    return saved_count


def _save_merch_preorder_images(batch_id: int, uploads, image_kind: str) -> int:
    media_dir = Path(current_app.config["MEDIA_UPLOAD_FOLDER"])
    media_dir.mkdir(parents=True, exist_ok=True)
    allowed_exts = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
    max_media_bytes = int(current_app.config.get("MAX_MEDIA_BYTES", 10 * 1024 * 1024))

    saved_count = 0
    existing_max = (
        db.session.query(db.func.coalesce(db.func.max(MerchPreorderImage.sort_order), 0))
        .filter_by(batch_id=batch_id, image_kind=image_kind)
        .scalar()
        or 0
    )
    for upload in uploads:
        if not upload or not (upload.filename or "").strip():
            continue
        original_name = secure_filename(Path(upload.filename).name)
        ext = Path(original_name).suffix.lower()
        if not original_name or ext not in allowed_exts:
            continue
        if not file_size_ok(upload, max_media_bytes):
            continue

        token = f"merch_{batch_id}_{image_kind}_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{secrets.token_hex(4)}{ext}"
        target_path = media_dir / token
        upload.save(target_path)
        mime_type = (upload.mimetype or "").strip() or (mimetypes.guess_type(original_name)[0] or "application/octet-stream")
        existing_max += 1
        db.session.add(
            MerchPreorderImage(
                batch_id=batch_id,
                image_kind=image_kind,
                original_filename=original_name[:255],
                storage_path=token,
                mime_type=mime_type[:128],
                size_bytes=int(target_path.stat().st_size),
                sort_order=int(existing_max),
                created_at=utcnow(),
            )
        )
        saved_count += 1
    return saved_count


def _merch_registration_count(batch_id: int) -> int:
    return int(
        db.session.query(db.func.coalesce(db.func.sum(MerchPreorderRegistration.quantity), 0))
        .filter(
            MerchPreorderRegistration.batch_id == batch_id,
            MerchPreorderRegistration.status.in_(MERCH_ACTIVE_ORDER_STATUSES),
        )
        .scalar()
        or 0
    )


def _read_wechat_qr_upload(upload):
    try:
        import cv2
        import numpy as np
    except Exception:
        return None, None, "missing_dependency"

    current_pos = upload.stream.tell()
    upload.stream.seek(0)
    raw_bytes = upload.stream.read()
    upload.stream.seek(current_pos)
    if not raw_bytes:
        return None, None, "empty_file"

    image_data = np.frombuffer(raw_bytes, dtype=np.uint8)
    image = cv2.imdecode(image_data, cv2.IMREAD_COLOR)
    if image is None:
        return None, None, "decode_failed"
    return cv2, image, ""


def _extract_wechat_qr_preview_pair(upload) -> tuple[dict | None, str]:
    cv2, image, error_code = _read_wechat_qr_upload(upload)
    if error_code:
        return None, error_code

    def decode_qr_text(candidate_image) -> tuple[str, object | None]:
        detector = cv2.QRCodeDetector()
        try:
            decoded_text, points, _straight_qr = detector.detectAndDecode(candidate_image)
        except Exception:
            return "", None
        return (decoded_text or "").strip(), points

    def detect_qr_points(candidate_image):
        decoded_text, points = decode_qr_text(candidate_image)
        if points is None:
            detector = cv2.QRCodeDetector()
            try:
                detected, points = detector.detect(candidate_image)
            except Exception:
                detected = False
                points = None
            if not detected:
                points = None
        return decoded_text, points

    def crop_qr(candidate_image, points):
        import numpy as np

        points = np.asarray(points, dtype=np.float32).reshape(-1, 2)
        if points.shape[0] < 4:
            return None, "no_qr"
        points = points[:4]
        sums = points.sum(axis=1)
        diffs = np.diff(points, axis=1).reshape(-1)
        ordered = np.array(
            [
                points[np.argmin(sums)],
                points[np.argmin(diffs)],
                points[np.argmax(sums)],
                points[np.argmax(diffs)],
            ],
            dtype=np.float32,
        )
        width_top = np.linalg.norm(ordered[1] - ordered[0])
        width_bottom = np.linalg.norm(ordered[2] - ordered[3])
        height_right = np.linalg.norm(ordered[2] - ordered[1])
        height_left = np.linalg.norm(ordered[3] - ordered[0])
        side = int(max(width_top, width_bottom, height_right, height_left))
        if side < 40:
            return None, "no_qr"

        target = np.array(
            [[0, 0], [side - 1, 0], [side - 1, side - 1], [0, side - 1]],
            dtype=np.float32,
        )
        matrix = cv2.getPerspectiveTransform(ordered, target)
        qr_image = cv2.warpPerspective(candidate_image, matrix, (side, side))
        border = max(12, int(side * 0.06))
        qr_image = cv2.copyMakeBorder(
            qr_image,
            border,
            border,
            border,
            border,
            cv2.BORDER_CONSTANT,
            value=[255, 255, 255],
        )
        encoded, output = cv2.imencode(".png", qr_image)
        if not encoded:
            return None, "encode_failed"
        return output.tobytes(), qr_image, ""

    def decode_from_options(*images) -> str:
        for item in images:
            if item is None:
                continue
            decoded_text, _points = decode_qr_text(item)
            if decoded_text:
                return decoded_text
            decoded_text, _points = decode_qr_text(cv2.bitwise_not(item))
            if decoded_text:
                return decoded_text
        return ""

    def normalize_detected_qr(candidate_image):
        import numpy as np

        if candidate_image is None:
            return None, "no_qr"
        if len(candidate_image.shape) == 3:
            gray = cv2.cvtColor(candidate_image, cv2.COLOR_BGR2GRAY)
        else:
            gray = candidate_image
        _threshold, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        edge = max(1, int(binary.shape[0] * 0.12))
        edge_samples = np.concatenate(
            [
                binary[:edge, :].reshape(-1),
                binary[-edge:, :].reshape(-1),
                binary[:, :edge].reshape(-1),
                binary[:, -edge:].reshape(-1),
            ]
        )
        if float(np.mean(edge_samples)) < 127:
            binary = cv2.bitwise_not(binary)

        scale = max(1, 360 // max(1, int(binary.shape[0])))
        if scale > 1:
            binary = cv2.resize(binary, None, fx=scale, fy=scale, interpolation=cv2.INTER_NEAREST)
        border = max(24, int(binary.shape[0] * 0.10))
        binary = cv2.copyMakeBorder(
            binary,
            border,
            border,
            border,
            border,
            cv2.BORDER_CONSTANT,
            value=255,
        )
        encoded, output = cv2.imencode(".png", binary)
        if not encoded:
            return None, "encode_failed"
        return output.tobytes(), ""

    selected_image = image
    decoded_text, points = detect_qr_points(image)
    if points is None:
        selected_image = cv2.bitwise_not(image)
        decoded_text, points = detect_qr_points(selected_image)
    if points is None:
        return None, "no_qr"

    detected_png, detected_image, error_code = crop_qr(selected_image, points)
    if error_code:
        return None, error_code

    if not decoded_text:
        decoded_text = decode_from_options(image, selected_image, detected_image)
    if not decoded_text:
        generated_png, error_code = normalize_detected_qr(detected_image)
        if error_code:
            return None, error_code
        return {"detected_png": detected_png, "generated_png": generated_png}, "decode_text_failed"

    generated_png, error_code = _generate_wechat_qr_png(decoded_text, cv2)
    if error_code:
        return None, error_code
    return {"detected_png": detected_png, "generated_png": generated_png}, ""


def _generate_wechat_qr_png(decoded_text: str, cv2) -> tuple[bytes | None, str]:
    try:
        import numpy as np
    except Exception:
        return None, "missing_dependency"
    try:
        qr_image = cv2.QRCodeEncoder_create().encode(decoded_text)
    except Exception:
        return None, "encode_failed"
    qr_image = qr_image.astype(np.uint8)
    scale = max(1, 360 // max(1, int(qr_image.shape[0])))
    if scale > 1:
        qr_image = cv2.resize(qr_image, None, fx=scale, fy=scale, interpolation=cv2.INTER_NEAREST)
    border = max(16, int(qr_image.shape[0] * 0.08))
    qr_image = cv2.copyMakeBorder(
        qr_image,
        border,
        border,
        border,
        border,
        cv2.BORDER_CONSTANT,
        value=255,
    )
    encoded, output = cv2.imencode(".png", qr_image)
    if not encoded:
        return None, "encode_failed"
    return output.tobytes(), ""


def _extract_wechat_qr_png(upload) -> tuple[bytes | None, str]:
    preview_pair, error_code = _extract_wechat_qr_preview_pair(upload)
    if error_code:
        return None, error_code
    return preview_pair["generated_png"], ""


def _save_activity_wechat_qr(activity_id: int, upload) -> tuple[str | None, Path | None, str]:
    if not upload or not (upload.filename or "").strip():
        return None, None, ""
    media_dir = Path(current_app.config["MEDIA_UPLOAD_FOLDER"])
    media_dir.mkdir(parents=True, exist_ok=True)
    original_name = secure_filename(Path(upload.filename).name)
    ext = Path(original_name).suffix.lower()
    allowed_image_exts = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
    max_media_bytes = int(current_app.config.get("MAX_MEDIA_BYTES", 10 * 1024 * 1024))
    if not original_name or ext not in allowed_image_exts:
        return None, None, "invalid_type"
    if not file_size_ok(upload, max_media_bytes):
        return None, None, "too_large"
    qr_png, error_code = _extract_wechat_qr_png(upload)
    if not qr_png:
        return None, None, error_code or "qr_extract_failed"
    filename = f"wechat_group_qr_{activity_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}_{secrets.token_hex(4)}.png"
    target_path = media_dir / filename
    target_path.write_bytes(qr_png)
    return filename, target_path, ""


def _wechat_qr_upload_error_message(error_code: str) -> str:
    messages = {
        "invalid_type": "微信群二维码上传失败：仅支持 jpg/jpeg/png/webp/gif 图片",
        "too_large": "微信群二维码上传失败：图片超过大小限制",
        "missing_dependency": "微信群二维码上传失败：当前环境缺少 OpenCV 依赖，请先安装 requirements.txt",
        "decode_failed": "微信群二维码上传失败：无法读取该图片",
        "no_qr": "微信群二维码上传失败：未能在图片中识别出二维码，请上传清晰、无遮挡的二维码照片",
        "decode_text_failed": "已裁切出疑似二维码，但未能解码内容，暂时无法生成标准黑白二维码",
        "encode_failed": "微信群二维码上传失败：二维码图片处理失败",
        "empty_file": "微信群二维码上传失败：上传文件为空",
    }
    return messages.get(error_code or "", "微信群二维码上传失败：请上传清晰的二维码图片")


def _required_page_permission_for_request(path: str, method: str) -> tuple[str, str] | None:
    if not path.startswith("/manage") or path in {"/manage", "/manage/"}:
        return None
    if path.startswith("/manage/login") or path.startswith("/manage/logout"):
        return None

    normalized = path.rstrip("/")
    is_post = method.upper() == "POST"

    if normalized.startswith("/manage/analytics"):
        return PAGE_ANALYTICS, PERMISSION_READ
    if normalized.startswith("/manage/security"):
        return PAGE_SECURITY, PERMISSION_READ
    if normalized.startswith("/manage/audit-logs"):
        return PAGE_AUDIT_LOGS, PERMISSION_READ
    if normalized.startswith("/manage/users"):
        if is_post and ("/delete" in normalized or "/deactivate" in normalized):
            return PAGE_ACCOUNTS, PERMISSION_ADMIN
        if is_post:
            return PAGE_ACCOUNTS, PERMISSION_ADMIN
        if normalized.endswith("/new") or normalized.endswith("/edit"):
            return PAGE_ACCOUNTS, PERMISSION_ADMIN
        return PAGE_ACCOUNTS, PERMISSION_READ

    content_prefixes = (
        ("/manage/routes", PAGE_ROUTES),
        ("/manage/activities", PAGE_ACTIVITIES),
        ("/manage/kit-preorders", PAGE_KIT_PREORDERS),
        ("/manage/announcements", PAGE_ANNOUNCEMENTS),
    )
    for prefix, page_key in content_prefixes:
        if normalized.startswith(prefix):
            if any(item in normalized for item in ("/delete", "/rollback", "/restore", "/recycle")):
                return page_key, PERMISSION_ADMIN
            if is_post or normalized.endswith("/new") or normalized.endswith("/edit"):
                return page_key, PERMISSION_WRITE
            return page_key, PERMISSION_READ

    if normalized.startswith("/manage/bulk-import") or normalized.startswith("/manage/import-report"):
        return PAGE_ROUTES, PERMISSION_WRITE
    if (
        normalized.startswith("/manage/feedback")
        or normalized.startswith("/manage/route-feedback")
        or normalized.startswith("/manage/site-feedback")
    ):
        if is_post and "/delete" in normalized:
            return PAGE_FEEDBACK, PERMISSION_ADMIN
        if is_post:
            return PAGE_FEEDBACK, PERMISSION_WRITE
        return PAGE_FEEDBACK, PERMISSION_READ
    return None


@bp.before_app_request
def _load_user():
    attach_current_user()


@bp.before_request
def _verify_csrf_for_post():
    if request.method != "POST":
        return
    token = request.form.get("csrf_token")
    if not validate_csrf_token(token):
        abort(400, description="Invalid CSRF token")


@bp.before_request
def _enforce_permission_matrix():
    if not getattr(g, "current_user", None):
        return
    path = request.path or ""
    if path in {"/manage/login"}:
        return
    user = g.current_user
    required = _required_page_permission_for_request(path, request.method)
    if not required:
        return
    page_key, required_level = required
    if required_level == PERMISSION_ADMIN and not can_admin_page(user, page_key):
        abort(403)
    if required_level == PERMISSION_WRITE and not can_write_page(user, page_key):
        abort(403)
    if required_level == PERMISSION_READ and not can_read_page(user, page_key):
        abort(403)


@bp.app_context_processor
def _inject_csrf_token():
    user = getattr(g, "current_user", None)
    nav_items = []
    if user:
        nav_items.append({"label": "总览", "endpoint": "admin.dashboard", "prefix": "/manage"})
        if can_read_page(user, PAGE_ROUTES):
            nav_items.append({"label": "路线", "endpoint": "admin.routes_page", "prefix": "/manage/routes"})
        if can_read_page(user, PAGE_ACTIVITIES):
            nav_items.append({"label": "活动", "endpoint": "admin.activities_page", "prefix": "/manage/activities"})
        if can_read_page(user, PAGE_KIT_PREORDERS):
            nav_items.append({"label": "预定", "endpoint": "admin.merch_preorders_page", "prefix": "/manage/kit-preorders"})
        if can_read_page(user, PAGE_ANNOUNCEMENTS):
            nav_items.append({"label": "公告", "endpoint": "admin.announcements_page", "prefix": "/manage/announcements"})
        if can_read_page(user, PAGE_FEEDBACK):
            nav_items.append({"label": "反馈", "endpoint": "admin.site_feedback_page", "prefix": "/manage/site-feedback"})
        if can_read_page(user, PAGE_ANALYTICS):
            nav_items.append({"label": "流量", "endpoint": "admin.analytics_page", "prefix": "/manage/analytics"})
        if can_read_page(user, PAGE_SECURITY):
            nav_items.append({"label": "安全", "endpoint": "admin.security_page", "prefix": "/manage/security"})
        if can_read_page(user, PAGE_ACCOUNTS):
            nav_items.append({"label": "账号", "endpoint": "admin.users_page", "prefix": "/manage/users"})
    return {
        "csrf_token": get_csrf_token,
        "to_local_time": _to_local_time,
        "app_version": _display_app_version(),
        "manage_nav_items": nav_items,
        "audit_action_label": _audit_action_label,
    }


@bp.post("/activities/wechat-qr/preview")
@login_required
def preview_activity_wechat_qr():
    upload = request.files.get("wechat_qr_file") or request.files.get("insurance_qr_file")
    if not upload or not (upload.filename or "").strip():
        return jsonify({"ok": False, "message": "请先选择一张包含二维码的图片。"}), 400
    preview_pair, error_code = _extract_wechat_qr_preview_pair(upload)
    if not preview_pair:
        return jsonify(
            {
                "ok": False,
                "message": _wechat_qr_upload_error_message(error_code),
            }
        ), 400
    detected_url = "data:image/png;base64," + base64.b64encode(preview_pair["detected_png"]).decode("ascii")
    generated_png = preview_pair.get("generated_png")
    if not generated_png:
        return jsonify(
            {
                "ok": True,
                "partial": True,
                "detected_image_data_url": detected_url,
                "generated_image_data_url": "",
                "image_data_url": detected_url,
                "message": _wechat_qr_upload_error_message(error_code),
            }
        )
    message = (
        "未能解码二维码内容，已根据识别结果二值化并生成黑白兜底图，提交活动后将保存右侧结果。"
        if error_code == "decode_text_failed"
        else "已识别二维码内容并生成标准黑白二维码，提交活动后将保存右侧标准黑白结果。"
    )
    return jsonify(
        {
            "ok": True,
            "detected_image_data_url": detected_url,
            "generated_image_data_url": "data:image/png;base64," + base64.b64encode(generated_png).decode("ascii"),
            "image_data_url": "data:image/png;base64," + base64.b64encode(generated_png).decode("ascii"),
            "message": message,
        }
    )


@bp.get("/login")
def login():
    if g.current_user:
        return redirect(url_for("admin.dashboard"))
    return render_template("manage_login.html")


@bp.post("/login")
def login_submit():
    username = (request.form.get("username") or "").strip()
    password = (request.form.get("password") or "").strip()
    source_ip = client_ip()
    user_agent = (request.user_agent.string or "")[:255]
    request_path = request.path
    login_subject = f"{source_ip}:{username.lower() or '_'}"

    retry_after = check_lock("admin_login", login_subject, LOGIN_WINDOW_SECONDS)
    if retry_after > 0:
        flash(f"登录过于频繁，请 {retry_after} 秒后再试", "error")
        return redirect(url_for("admin.login"))

    user = User.query.filter_by(username=username, is_active=True).first()
    if not user or not check_password_hash(user.password, password):
        retry_after = register_failure(
            "admin_login",
            login_subject,
            max_attempts=LOGIN_MAX_FAILURES,
            window_seconds=LOGIN_WINDOW_SECONDS,
            lock_seconds=LOGIN_LOCK_SECONDS,
        )
        if retry_after > 0:
            flash(f"登录过于频繁，请 {retry_after} 秒后再试", "error")
        else:
            flash("用户名或密码错误", "error")
        write_audit_log(
            None,
            "auth.login_failed",
            "user",
            username or None,
            f'ip={source_ip};path={request_path};ua="{user_agent}";retry_after={retry_after}',
        )
        return redirect(url_for("admin.login"))

    clear_state("admin_login", login_subject)
    session["user_id"] = user.id
    write_audit_log(
        user.id,
        "auth.login",
        "user",
        str(user.id),
        f'ip={source_ip};path={request_path};ua="{user_agent}"',
    )
    return redirect(url_for("admin.dashboard"))


@bp.post("/logout")
@login_required
def logout():
    actor_id = g.current_user.id if g.current_user else None
    session.pop("user_id", None)
    write_audit_log(actor_id, "auth.logout", "user", str(actor_id) if actor_id else None, "logout")
    return redirect(url_for("admin.login"))


@bp.get("")
@login_required
def dashboard():
    log_page = max(1, request.args.get("log_page", default=1, type=int))
    now = utcnow()
    start_24h = now - timedelta(hours=24)
    user = g.current_user
    can_view_routes_flag = can_read_page(user, PAGE_ROUTES)
    can_view_activities_flag = can_read_page(user, PAGE_ACTIVITIES)
    can_view_kit_preorders_flag = can_read_page(user, PAGE_KIT_PREORDERS)
    can_view_announcements_flag = can_read_page(user, PAGE_ANNOUNCEMENTS)
    can_view_feedback_flag = can_read_page(user, PAGE_FEEDBACK)
    can_view_analytics_flag = can_read_page(user, PAGE_ANALYTICS)
    can_view_security_flag = can_read_page(user, PAGE_SECURITY)
    can_manage_users_flag = can_read_page(user, PAGE_ACCOUNTS)
    can_view_audit_logs_flag = can_read_page(user, PAGE_AUDIT_LOGS)
    can_write_routes_flag = can_write_page(user, PAGE_ROUTES)
    can_write_activities_flag = can_write_page(user, PAGE_ACTIVITIES)
    can_edit_flag = any(
        can_write_page(user, page_key)
        for page_key in (PAGE_ROUTES, PAGE_ACTIVITIES, PAGE_KIT_PREORDERS, PAGE_ANNOUNCEMENTS)
    )

    pending_feedback_count = RouteFeedback.query.filter_by(status=FEEDBACK_PENDING).count() if can_view_feedback_flag else 0
    pending_site_feedback_count = SiteFeedback.query.filter_by(status=SITE_FEEDBACK_PENDING).count() if can_view_feedback_flag else 0
    audit_logs_pagination = (
        AuditLog.query.order_by(AuditLog.created_at.desc()).paginate(page=log_page, per_page=3, error_out=False)
        if can_view_audit_logs_flag
        else None
    )
    latest_routes = (
        Route.query.filter_by(is_deleted=False).order_by(Route.updated_at.desc()).limit(3).all()
        if can_view_routes_flag
        else []
    )
    latest_activities = (
        Activity.query.order_by(Activity.activity_time.desc()).limit(3).all() if can_view_activities_flag else []
    )
    latest_merch_batches = (
        MerchPreorderBatch.query.order_by(MerchPreorderBatch.updated_at.desc(), MerchPreorderBatch.id.desc()).limit(3).all()
        if can_view_kit_preorders_flag
        else []
    )
    latest_announcements = (
        Announcement.query.order_by(
            Announcement.is_pinned.desc(),
            Announcement.sort_order.desc(),
            db.func.coalesce(Announcement.published_at, Announcement.updated_at).desc(),
        )
        .limit(3)
        .all()
        if can_view_announcements_flag
        else []
    )
    latest_feedback = (
        RouteFeedback.query.order_by(RouteFeedback.created_at.desc()).limit(3).all() if can_view_feedback_flag else []
    )
    latest_site_feedback = (
        SiteFeedback.query.order_by(SiteFeedback.created_at.desc()).limit(3).all() if can_view_feedback_flag else []
    )
    summary = {
        "route_total": Route.query.filter_by(is_deleted=False).count() if can_view_routes_flag else 0,
        "route_deleted": Route.query.filter_by(is_deleted=True).count() if can_view_routes_flag else 0,
        "activity_total": Activity.query.count() if can_view_activities_flag else 0,
        "merch_batch_total": MerchPreorderBatch.query.count() if can_view_kit_preorders_flag else 0,
        "announcement_total": Announcement.query.count() if can_view_announcements_flag else 0,
        "feedback_pending": pending_feedback_count,
        "site_feedback_pending": pending_site_feedback_count,
    }
    security_summary = {"probe_24h": 0, "watchlist_24h": 0, "blocked_429_24h": 0, "probe_ip_24h": 0}
    if can_view_security_flag:
        probe_filter_24h = (
            AccessLog.created_at >= start_24h,
            ~AccessLog.path.like("/manage%"),
            _probe_expression(),
        )
        security_summary = {
            "probe_24h": AccessLog.query.filter(*probe_filter_24h).count(),
            "watchlist_24h": AccessLog.query.filter(
                AccessLog.created_at >= start_24h,
                ~AccessLog.path.like("/manage%"),
                _watchlist_probe_expression(),
            ).count(),
            "blocked_429_24h": AccessLog.query.filter(
                AccessLog.created_at >= start_24h,
                ~AccessLog.path.like("/manage%"),
                AccessLog.status_code == 429,
            ).count(),
            "probe_ip_24h": int(
                db.session.query(db.func.count(db.distinct(AccessLog.ip_address)))
                .filter(*probe_filter_24h)
                .scalar()
                or 0
            ),
        }
    analytics_summary = {"pv_24h": 0, "uv_24h": 0, "active_5m": 0}
    if can_view_analytics_flag:
        analytics_summary = {
            "pv_24h": AccessLog.query.filter(
                AccessLog.created_at >= start_24h,
                *build_non_probe_filters(AccessLog),
            ).count(),
            "uv_24h": int(
                db.session.query(db.func.count(db.distinct(AccessLog.ip_address)))
                .filter(
                    AccessLog.created_at >= start_24h,
                    *build_non_probe_filters(AccessLog),
                )
                .scalar()
                or 0
            ),
            "active_5m": int(
                db.session.query(db.func.count(db.distinct(AccessLog.ip_address)))
                .filter(
                    AccessLog.created_at >= (now - timedelta(minutes=5)),
                    *build_non_probe_filters(AccessLog),
                )
                .scalar()
                or 0
            ),
        }
    return render_template(
        "manage.html",
        summary=summary,
        analytics_summary=analytics_summary,
        audit_logs=audit_logs_pagination.items if audit_logs_pagination else [],
        audit_logs_pagination=audit_logs_pagination,
        latest_routes=latest_routes,
        latest_activities=latest_activities,
        latest_merch_batches=latest_merch_batches,
        merch_batch_status_labels=MERCH_BATCH_STATUS_LABELS,
        latest_announcements=latest_announcements,
        latest_feedback=latest_feedback,
        latest_site_feedback=latest_site_feedback,
        security_summary=security_summary,
        can_view_routes=can_view_routes_flag,
        can_view_activities=can_view_activities_flag,
        can_view_kit_preorders=can_view_kit_preorders_flag,
        can_view_announcements=can_view_announcements_flag,
        can_review=can_view_feedback_flag,
        can_manage_users=can_manage_users_flag,
        can_view_analytics=can_view_analytics_flag,
        can_view_security=can_view_security_flag,
        can_view_audit_logs=can_view_audit_logs_flag,
        can_write_routes=can_write_routes_flag,
        can_write_activities=can_write_activities_flag,
        can_edit=can_edit_flag,
    )


@bp.get("/audit-logs")
@login_required
def audit_logs_page():
    page = max(1, request.args.get("page", default=1, type=int))
    pagination = AuditLog.query.order_by(AuditLog.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    return render_template("manage_audit_logs.html", logs=pagination.items, pagination=pagination)


@bp.get("/analytics")
@login_required
def analytics_page():
    days = request.args.get("days", default=7, type=int)
    if days not in {1, 7, 30}:
        days = 7
    scope = (request.args.get("scope") or "recent").strip()
    if scope not in {"recent", "post_deploy"}:
        scope = "recent"

    now = utcnow()
    scope, start_recent, period_label = _resolve_period_start(days, scope, now)
    start_recent_date = _to_local_time(start_recent).date()
    today_local = _to_local_time(now).date()

    recent_base_filter = (AccessLog.created_at >= start_recent, *build_non_probe_filters(AccessLog))
    recent_raw_filter = (
        AccessLog.created_at >= start_recent,
        ~AccessLog.path.like("/manage%"),
    )
    total_base_filter = (
        AccessLog.created_at >= start_recent,
        *build_non_probe_filters(AccessLog),
    ) if scope == "post_deploy" else build_non_probe_filters(AccessLog)
    total_raw_filter = (
        AccessLog.created_at >= start_recent,
        ~AccessLog.path.like("/manage%"),
    ) if scope == "post_deploy" else (~AccessLog.path.like("/manage%"),)
    recent_pv = AccessLog.query.filter(*recent_base_filter).count()
    recent_uv = (
        db.session.query(db.func.count(db.distinct(AccessLog.ip_address)))
        .filter(*recent_base_filter)
        .scalar()
        or 0
    )
    recent_errors_4xx = AccessLog.query.filter(
        *recent_base_filter,
        AccessLog.status_code >= 400,
        AccessLog.status_code < 500,
    ).count()
    recent_errors_5xx = AccessLog.query.filter(
        *recent_base_filter,
        AccessLog.status_code >= 500,
    ).count()
    recent_downloads = AccessLog.query.filter(
        *recent_base_filter,
        AccessLog.path.like("/download/%"),
        AccessLog.status_code == 200,
    ).count()
    recent_raw_requests = AccessLog.query.filter(*recent_raw_filter).count()
    recent_probe_excluded = max(0, recent_raw_requests - recent_pv)
    recent_probe_ratio = round((recent_probe_excluded / recent_raw_requests) * 100, 2) if recent_raw_requests else 0.0

    total_pv = AccessLog.query.filter(*total_base_filter).count()
    total_uv = (
        db.session.query(db.func.count(db.distinct(AccessLog.ip_address)))
        .filter(*total_base_filter)
        .scalar()
        or 0
    )
    total_errors_4xx = AccessLog.query.filter(
        *total_base_filter,
        AccessLog.status_code >= 400,
        AccessLog.status_code < 500,
    ).count()
    total_errors_5xx = AccessLog.query.filter(
        *total_base_filter,
        AccessLog.status_code >= 500,
    ).count()
    total_downloads = AccessLog.query.filter(
        *total_base_filter,
        AccessLog.path.like("/download/%"),
        AccessLog.status_code == 200,
    ).count()
    total_raw_requests = AccessLog.query.filter(*total_raw_filter).count()
    total_probe_excluded = max(0, total_raw_requests - total_pv)
    total_probe_ratio = round((total_probe_excluded / total_raw_requests) * 100, 2) if total_raw_requests else 0.0

    top_pages_rows = (
        db.session.query(
            AccessLog.path,
            db.func.count(AccessLog.id).label("hits"),
        )
        .filter(*recent_base_filter)
        .group_by(AccessLog.path)
        .order_by(db.text("hits DESC"))
        .limit(10)
        .all()
    )
    top_pages = [{"path": row[0], "hits": int(row[1] or 0)} for row in top_pages_rows]

    logs_recent = (
        db.session.query(AccessLog.created_at, AccessLog.path, AccessLog.status_code)
        .filter(*recent_base_filter)
        .all()
    )
    daily_map: dict = defaultdict(lambda: {"pv": 0, "downloads": 0})
    for created_at, path, status_code in logs_recent:
        local_day = _to_local_time(created_at).date()
        daily_map[local_day]["pv"] += 1
        if (path or "").startswith("/download/") and status_code == 200:
            daily_map[local_day]["downloads"] += 1

    daily_stats = []
    current = start_recent_date
    while current <= today_local:
        row = daily_map[current]
        daily_stats.append({"date": current.strftime("%Y-%m-%d"), "pv": row["pv"], "downloads": row["downloads"]})
        current += timedelta(days=1)

    return render_template(
        "manage_analytics.html",
        days=days,
        scope=scope,
        period_label=period_label,
        can_view_security=can_view_security(g.current_user),
        recent={
            "pv": recent_pv,
            "uv": int(recent_uv),
            "errors_4xx": recent_errors_4xx,
            "errors_5xx": recent_errors_5xx,
            "downloads": recent_downloads,
            "raw_requests": recent_raw_requests,
            "probe_excluded": recent_probe_excluded,
            "probe_ratio": recent_probe_ratio,
        },
        total={
            "pv": total_pv,
            "uv": int(total_uv),
            "errors_4xx": total_errors_4xx,
            "errors_5xx": total_errors_5xx,
            "downloads": total_downloads,
            "raw_requests": total_raw_requests,
            "probe_excluded": total_probe_excluded,
            "probe_ratio": total_probe_ratio,
        },
        top_pages=top_pages,
        daily_stats=daily_stats,
    )


@bp.get("/security")
@login_required
def security_page():
    days = request.args.get("days", default=7, type=int)
    if days not in {1, 7, 30}:
        days = 7
    scope = (request.args.get("scope") or "recent").strip()
    if scope not in {"recent", "post_deploy"}:
        scope = "recent"

    now = utcnow()
    scope, start_recent, period_label = _resolve_period_start(days, scope, now)
    start_recent_date = _to_local_time(start_recent).date()
    today_local = _to_local_time(now).date()
    base_filter = (
        AccessLog.created_at >= start_recent,
        ~AccessLog.path.like("/manage%"),
    )
    probe_expr = _probe_expression()
    watchlist_expr = _watchlist_probe_expression()
    security_expr = or_(probe_expr, AccessLog.status_code == 429)

    recent_probe = AccessLog.query.filter(*base_filter, probe_expr).count()
    recent_watchlist = AccessLog.query.filter(*base_filter, watchlist_expr).count()
    recent_throttled = AccessLog.query.filter(*base_filter, AccessLog.status_code == 429).count()
    recent_errors_4xx = AccessLog.query.filter(
        *base_filter,
        AccessLog.status_code >= 400,
        AccessLog.status_code < 500,
    ).count()
    recent_errors_5xx = AccessLog.query.filter(*base_filter, AccessLog.status_code >= 500).count()
    recent_probe_ip = int(
        db.session.query(db.func.count(db.distinct(AccessLog.ip_address)))
        .filter(*base_filter, security_expr)
        .scalar()
        or 0
    )

    top_paths_rows = (
        db.session.query(AccessLog.path, db.func.count(AccessLog.id).label("hits"))
        .filter(*base_filter, security_expr)
        .group_by(AccessLog.path)
        .order_by(db.text("hits DESC"))
        .limit(10)
        .all()
    )
    top_ips_rows = (
        db.session.query(AccessLog.ip_address, db.func.count(AccessLog.id).label("hits"))
        .filter(*base_filter, security_expr)
        .group_by(AccessLog.ip_address)
        .order_by(db.text("hits DESC"))
        .limit(10)
        .all()
    )
    top_uas_rows = (
        db.session.query(AccessLog.user_agent, db.func.count(AccessLog.id).label("hits"))
        .filter(*base_filter, security_expr, AccessLog.user_agent != "")
        .group_by(AccessLog.user_agent)
        .order_by(db.text("hits DESC"))
        .limit(10)
        .all()
    )

    top_paths = [{"path": row[0], "hits": int(row[1] or 0)} for row in top_paths_rows]
    top_ips = [{"ip": row[0] or "unknown", "hits": int(row[1] or 0)} for row in top_ips_rows]
    top_uas = [{"ua": row[0] or "-", "hits": int(row[1] or 0)} for row in top_uas_rows]

    event_page = max(1, request.args.get("event_page", default=1, type=int))
    event_type = (request.args.get("event_type") or "all").strip().lower()
    if event_type not in {"all", "watchlist", "probe", "throttled"}:
        event_type = "all"
    event_status = (request.args.get("event_status") or "all").strip().lower()
    if event_status not in {"all", "4xx", "5xx", "429"}:
        event_status = "all"
    event_q = (request.args.get("event_q") or "").strip()

    event_query = AccessLog.query.filter(*base_filter, security_expr)
    if event_type == "watchlist":
        event_query = event_query.filter(watchlist_expr)
    elif event_type == "probe":
        event_query = event_query.filter(probe_expr)
    elif event_type == "throttled":
        event_query = event_query.filter(AccessLog.status_code == 429)

    if event_status == "4xx":
        event_query = event_query.filter(AccessLog.status_code >= 400, AccessLog.status_code < 500)
    elif event_status == "5xx":
        event_query = event_query.filter(AccessLog.status_code >= 500)
    elif event_status == "429":
        event_query = event_query.filter(AccessLog.status_code == 429)

    if event_q:
        pattern = f"%{event_q}%"
        event_query = event_query.filter(
            or_(
                AccessLog.path.ilike(pattern),
                AccessLog.ip_address.ilike(pattern),
                AccessLog.user_agent.ilike(pattern),
            )
        )

    events_pagination = event_query.order_by(AccessLog.created_at.desc()).paginate(
        page=event_page, per_page=50, error_out=False
    )
    logs_recent = events_pagination.items
    logs_for_trend = (
        db.session.query(AccessLog.created_at, AccessLog.path, AccessLog.status_code)
        .filter(*base_filter, security_expr)
        .all()
    )
    watchlist_set = {item.lower() for item in WATCHLIST_PROBE_PATHS}
    event_rows = []
    for item in logs_recent:
        normalized = (item.path or "").lower()
        is_watchlist = normalized in watchlist_set or normalized.startswith("/wordpress/wp-admin/")
        event_rows.append(
            {
                "created_at": item.created_at,
                "path": item.path,
                "ip": item.ip_address or "unknown",
                "status_code": item.status_code,
                "user_agent": item.user_agent or "-",
                "event_type": "watchlist" if is_watchlist else ("throttled" if item.status_code == 429 else "probe"),
            }
        )

    daily_map: dict = defaultdict(lambda: {"probe": 0, "watchlist": 0, "throttled": 0})
    for created_at, path, status_code in logs_for_trend:
        local_day = _to_local_time(created_at).date()
        normalized = (path or "").lower()
        if normalized in watchlist_set or normalized.startswith("/wordpress/wp-admin/"):
            daily_map[local_day]["watchlist"] += 1
        elif status_code == 429:
            daily_map[local_day]["throttled"] += 1
        else:
            daily_map[local_day]["probe"] += 1

    daily_stats = []
    current = start_recent_date
    while current <= today_local:
        row = daily_map[current]
        daily_stats.append(
            {
                "date": current.strftime("%Y-%m-%d"),
                "probe": row["probe"],
                "watchlist": row["watchlist"],
                "throttled": row["throttled"],
            }
        )
        current += timedelta(days=1)

    return render_template(
        "manage_security.html",
        days=days,
        scope=scope,
        period_label=period_label,
        can_view_analytics=can_view_analytics(g.current_user),
        recent={
            "probe": recent_probe,
            "watchlist": recent_watchlist,
            "throttled": recent_throttled,
            "probe_ip": recent_probe_ip,
            "errors_4xx": recent_errors_4xx,
            "errors_5xx": recent_errors_5xx,
        },
        top_paths=top_paths,
        top_ips=top_ips,
        top_uas=top_uas,
        daily_stats=daily_stats,
        events=event_rows,
        events_pagination=events_pagination,
        event_filters={
            "event_type": event_type,
            "event_status": event_status,
            "event_q": event_q,
        },
    )


@bp.get("/site-feedback")
@login_required
def site_feedback_page():
    keyword = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status") or "all").strip()
    category_filter = (request.args.get("category") or "all").strip()
    page = max(1, request.args.get("page", default=1, type=int))
    start_date = (request.args.get("start_date") or "").strip()
    end_date = (request.args.get("end_date") or "").strip()

    if status_filter not in {"all", SITE_FEEDBACK_PENDING, SITE_FEEDBACK_DONE}:
        status_filter = "all"
    if category_filter not in {"all", "bug", "suggestion", "data", "other"}:
        category_filter = "all"

    query = SiteFeedback.query
    if keyword:
        pattern = f"%{keyword}%"
        query = query.filter(
            or_(
                SiteFeedback.content.ilike(pattern),
                SiteFeedback.contact.ilike(pattern),
                SiteFeedback.source_page.ilike(pattern),
            )
        )
    if status_filter != "all":
        query = query.filter(SiteFeedback.status == status_filter)
    if category_filter != "all":
        query = query.filter(SiteFeedback.category == category_filter)

    if start_date:
        try:
            start_local = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=SH_TZ)
            query = query.filter(SiteFeedback.created_at >= start_local.astimezone(timezone.utc))
        except ValueError:
            start_date = ""
    if end_date:
        try:
            end_local = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=SH_TZ) + timedelta(days=1)
            query = query.filter(SiteFeedback.created_at < end_local.astimezone(timezone.utc))
        except ValueError:
            end_date = ""

    pagination = query.order_by(SiteFeedback.created_at.desc()).paginate(page=page, per_page=10, error_out=False)
    return render_template(
        "manage_site_feedback.html",
        feedback_list=pagination.items,
        pagination=pagination,
        filters={
            "q": keyword,
            "status": status_filter,
            "category": category_filter,
            "start_date": start_date,
            "end_date": end_date,
        },
        can_edit=can_write_page(g.current_user, PAGE_FEEDBACK),
    )


@bp.post("/site-feedback/<int:feedback_id>/status")
@login_required
def site_feedback_update_status(feedback_id: int):
    target_status = (request.form.get("status") or "").strip()
    if target_status not in {SITE_FEEDBACK_PENDING, SITE_FEEDBACK_DONE}:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": "状态无效"}), 400
        flash("状态无效", "error")
        return redirect(url_for("admin.site_feedback_page"))

    feedback = SiteFeedback.query.filter_by(id=feedback_id).first()
    if not feedback:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": "反馈不存在"}), 404
        flash("反馈不存在", "error")
        return redirect(url_for("admin.site_feedback_page"))

    old_status = feedback.status
    feedback.status = target_status
    feedback.updated_at = utcnow()
    db.session.commit()
    write_audit_log(
        g.current_user.id if g.current_user else None,
        "site_feedback.status_update",
        "site_feedback",
        str(feedback.id),
        f"{old_status}->{target_status}",
    )
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"ok": True, "status": feedback.status})

    next_url = (request.form.get("next") or "").strip()
    if next_url.startswith(url_for("admin.site_feedback_page")):
        return redirect(next_url)
    return redirect(url_for("admin.site_feedback_page"))


@bp.get("/feedback")
@login_required
def feedback_page():
    return redirect(url_for("admin.site_feedback_page", **request.args.to_dict(flat=True)))


@bp.get("/route-feedback")
@login_required
def route_feedback_page():
    keyword = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status") or "all").strip()
    start_date = (request.args.get("start_date") or "").strip()
    end_date = (request.args.get("end_date") or "").strip()
    if status_filter not in {"all", FEEDBACK_PENDING, FEEDBACK_APPROVED, FEEDBACK_REJECTED}:
        status_filter = "all"

    base_query = RouteFeedback.query.outerjoin(Route, Route.id == RouteFeedback.route_id)
    if keyword:
        pattern = f"%{keyword}%"
        base_query = base_query.filter(
            or_(
                Route.route_name.ilike(pattern),
                RouteFeedback.comment.ilike(pattern),
                RouteFeedback.road_condition_update.ilike(pattern),
            )
        )
    if start_date:
        try:
            start_local = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=SH_TZ)
            base_query = base_query.filter(RouteFeedback.created_at >= start_local.astimezone(timezone.utc))
        except ValueError:
            start_date = ""
    if end_date:
        try:
            end_local = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=SH_TZ) + timedelta(days=1)
            base_query = base_query.filter(RouteFeedback.created_at < end_local.astimezone(timezone.utc))
        except ValueError:
            end_date = ""

    pending_page = max(1, request.args.get("pending_page", default=1, type=int))
    reviewed_page = max(1, request.args.get("reviewed_page", default=1, type=int))

    pending_query = base_query.filter(RouteFeedback.status == FEEDBACK_PENDING)
    if status_filter in {FEEDBACK_APPROVED, FEEDBACK_REJECTED}:
        pending_query = pending_query.filter(db.text("1=0"))

    reviewed_query = base_query.filter(RouteFeedback.status.in_([FEEDBACK_APPROVED, FEEDBACK_REJECTED]))
    if status_filter in {FEEDBACK_APPROVED, FEEDBACK_REJECTED}:
        reviewed_query = reviewed_query.filter(RouteFeedback.status == status_filter)
    elif status_filter == FEEDBACK_PENDING:
        reviewed_query = reviewed_query.filter(db.text("1=0"))

    pending_pagination = pending_query.order_by(RouteFeedback.created_at.desc()).paginate(page=pending_page, per_page=12, error_out=False)
    reviewed_pagination = reviewed_query.order_by(RouteFeedback.reviewed_at.desc(), RouteFeedback.created_at.desc()).paginate(
        page=reviewed_page, per_page=12, error_out=False
    )
    return render_template(
        "manage_feedback.html",
        pending_feedback=pending_pagination.items,
        reviewed_feedback=reviewed_pagination.items,
        pending_pagination=pending_pagination,
        reviewed_pagination=reviewed_pagination,
        filters={"q": keyword, "status": status_filter, "start_date": start_date, "end_date": end_date},
        can_edit=can_write_page(g.current_user, PAGE_FEEDBACK),
        can_admin=can_admin_page(g.current_user, PAGE_FEEDBACK),
    )


@bp.get("/routes")
@login_required
def routes_page():
    query, filters = query_routes_from_request(include_unpublished=True)
    pagination = query.paginate(page=filters["page"], per_page=filters["per_page"], error_out=False)
    recycle_count = Route.query.filter_by(is_deleted=True).count()
    return render_template(
        "manage_routes.html",
        routes=pagination.items,
        recycle_count=recycle_count,
        pagination=pagination,
        filters=filters,
        statuses=ROUTE_STATUSES,
        can_edit=can_write_page(g.current_user, PAGE_ROUTES),
        can_admin=can_admin_page(g.current_user, PAGE_ROUTES),
    )


@bp.get("/routes/recycle")
@login_required
def routes_recycle_page():
    recycle_routes = Route.query.filter_by(is_deleted=True).order_by(Route.deleted_at.desc()).all()
    return render_template(
        "manage_routes_recycle.html",
        recycle_routes=recycle_routes,
        can_edit=can_admin_page(g.current_user, PAGE_ROUTES),
    )


@bp.get("/routes/new")
@login_required
def route_new_page():
    return render_template(
        "manage_route_form.html",
        route=None,
        statuses=ROUTE_STATUSES,
        can_edit=can_write_page(g.current_user, PAGE_ROUTES),
        can_admin=can_admin_page(g.current_user, PAGE_ROUTES),
        manual_stats={},
        manual_stat_summary="",
    )


@bp.get("/routes/<int:route_id>/edit")
@login_required
def route_edit_page(route_id: int):
    route = Route.query.filter_by(id=route_id, is_deleted=False).first()
    if not route:
        flash("路线不存在", "error")
        return redirect(url_for("admin.routes_page"))
    versions = (
        RouteVersion.query.filter_by(route_id=route_id)
        .order_by(RouteVersion.version_no.desc())
        .limit(20)
        .all()
    )
    return render_template(
        "manage_route_form.html",
        route=route,
        statuses=ROUTE_STATUSES,
        versions=versions,
        can_edit=can_write_page(g.current_user, PAGE_ROUTES),
        can_admin=can_admin_page(g.current_user, PAGE_ROUTES),
        manual_stats=_route_manual_stat_overrides(route),
        manual_stat_summary=_manual_stat_summary(_route_manual_stat_overrides(route)),
    )


@bp.get("/routes/<int:route_id>/view")
@login_required
def route_detail_manage(route_id: int):
    route = Route.query.filter_by(id=route_id, is_deleted=False).first()
    if not route:
        abort(404, description="Route not found")
    avg_rating, rating_count = approved_rating_summary(route.id)
    return render_template(
        "route_detail.html",
        route=route,
        rating_info={"avg_rating": avg_rating, "rating_count": rating_count},
        preview_endpoint=url_for("admin.route_preview_manage", route_id=route.id),
        back_url=url_for("admin.routes_page"),
        back_label="返回列表",
    )



@bp.get("/routes/<int:route_id>/download")
@login_required
def download_route_gpx(route_id: int):
    route = Route.query.filter_by(id=route_id, is_deleted=False).first()
    if not route:
        abort(404, description="Route not found")

    file_path = _resolve_route_gpx_path(route)
    if not file_path.exists() or not file_path.is_file():
        flash("GPX 文件不存在", "error")
        return redirect(url_for("admin.routes_page"))

    return send_file(
        file_path,
        as_attachment=True,
        download_name=route.gpx_filename,
        mimetype="application/gpx+xml",
    )


@bp.get("/routes/<int:route_id>/preview")
@login_required
def route_preview_manage(route_id: int):
    route = Route.query.filter_by(id=route_id, is_deleted=False).first()
    if not route:
        return jsonify({"error": "route_not_found"}), 404

    file_path = _resolve_route_gpx_path(route)
    if not file_path.exists():
        return jsonify({"error": "gpx_missing"}), 404

    try:
        points, stats, elevation_profile = parse_gpx_points_and_stats(file_path)
        waypoints = parse_gpx_waypoints(file_path)
    except Exception:
        return jsonify({"error": "gpx_parse_failed"}), 400

    if not points:
        return jsonify(
            {
                "route_id": route.id,
                "points": [],
                "bounds": None,
                "stats": stats,
                "elevation_profile": elevation_profile,
                "waypoints": waypoints,
            }
        )

    lats = [item[0] for item in points]
    lons = [item[1] for item in points]
    bounds = {
        "min_lat": min(lats),
        "max_lat": max(lats),
        "min_lon": min(lons),
        "max_lon": max(lons),
    }
    return jsonify(
        {
            "route_id": route.id,
            "points": points,
            "bounds": bounds,
            "stats": stats,
            "elevation_profile": elevation_profile,
            "waypoints": waypoints,
        }
    )

@bp.get("/activities")
@login_required
def activities_page():
    page = max(1, request.args.get("page", default=1, type=int))
    activities_pagination = Activity.query.order_by(Activity.activity_time.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    activity_ids = [item.id for item in activities_pagination.items]
    registration_counts = {activity_id: 0 for activity_id in activity_ids}
    if activity_ids:
        rows = (
            db.session.query(EventRegistration.activity_id, db.func.count(EventRegistration.id))
            .filter(
                EventRegistration.activity_id.in_(activity_ids),
                EventRegistration.status.in_([REGISTRATION_PENDING, REGISTRATION_CONFIRMED]),
            )
            .group_by(EventRegistration.activity_id)
            .all()
        )
        registration_counts.update({int(activity_id): int(count or 0) for activity_id, count in rows})
    return render_template(
        "manage_activities.html",
        activities=activities_pagination.items,
        pagination=activities_pagination,
        can_edit=can_write_page(g.current_user, PAGE_ACTIVITIES),
        can_admin=can_admin_page(g.current_user, PAGE_ACTIVITIES),
        registration_counts=registration_counts,
    )


@bp.get("/activities/new")
@login_required
def activity_new_page():
    routes = Route.query.filter_by(is_deleted=False).order_by(Route.updated_at.desc()).all()
    return render_template(
        "manage_activity_form.html",
        activity=None,
        routes=routes,
        media_assets=[],
        route_option_items=[],
        route_option_map={},
        legacy_selected_ids=[],
        can_edit=can_write_page(g.current_user, PAGE_ACTIVITIES),
        can_admin=can_admin_page(g.current_user, PAGE_ACTIVITIES),
    )


@bp.get("/activities/<int:activity_id>/edit")
@login_required
def activity_edit_page(activity_id: int):
    activity = Activity.query.filter_by(id=activity_id).first()
    if not activity:
        flash("活动不存在", "error")
        return redirect(url_for("admin.activities_page"))
    routes = Route.query.filter_by(is_deleted=False).order_by(Route.updated_at.desc()).all()
    media_assets = MediaAsset.query.filter_by(activity_id=activity.id).order_by(MediaAsset.created_at.desc()).all()
    route_option_items = (
        ActivityRouteOption.query.filter_by(activity_id=activity.id)
        .order_by(ActivityRouteOption.sort_order.asc(), ActivityRouteOption.id.asc())
        .all()
    )
    route_option_map = {
        item.level_key: {
            "route_id": item.route_id,
            "participant_count": int(item.participant_count or 0),
            "activity_time": item.activity_time,
        }
        for item in route_option_items
    }
    if not route_option_map:
        legacy_levels = [key for key, _label in ACTIVITY_ROUTE_LEVELS]
        for index, route in enumerate(activity.routes[: len(legacy_levels)]):
            route_option_map[legacy_levels[index]] = {
                "route_id": route.id,
                "participant_count": int(activity.participant_count or 0),
                "activity_time": activity.activity_time,
            }
    return render_template(
        "manage_activity_form.html",
        activity=activity,
        routes=routes,
        media_assets=media_assets,
        route_option_items=route_option_items,
        route_option_map=route_option_map,
        legacy_selected_ids=[route.id for route in activity.routes],
        can_edit=can_write_page(g.current_user, PAGE_ACTIVITIES),
        can_admin=can_admin_page(g.current_user, PAGE_ACTIVITIES),
    )


@bp.get("/activities/<int:activity_id>/registrations")
@login_required
def activity_registrations_page(activity_id: int):
    activity = Activity.query.filter_by(id=activity_id).first()
    if not activity:
        flash("活动不存在", "error")
        return redirect(url_for("admin.activities_page"))
    page = max(1, request.args.get("page", default=1, type=int))
    pagination = EventRegistration.query.filter_by(activity_id=activity.id).order_by(
        EventRegistration.created_at.desc()
    ).paginate(page=page, per_page=50, error_out=False)
    registration_rows = []
    for item in pagination.items:
        note_map = _parse_registration_notes(item.notes or "")
        route_label = note_map.get("route_label", "")
        route_name = note_map.get("route_name", "")
        route_display = "-"
        if route_label and route_name:
            route_display = f"{route_label} · {route_name}"
        elif route_name:
            route_display = route_name
        elif route_label:
            route_display = route_label
        image_consent_display = "同意" if note_map.get("image_consent") == "1" else "未同意"
        registration_rows.append(
            {
                "id": item.id,
                "name": item.name,
                "student_id": item.student_id,
                "route_display": route_display,
                "status": item.status,
                "created_at": item.created_at,
                "source_ip": item.source_ip,
                "image_consent_display": image_consent_display,
            }
        )

    return render_template(
        "manage_activity_registrations.html",
        activity=activity,
        registrations=registration_rows,
        pagination=pagination,
        can_edit=can_write_page(g.current_user, PAGE_ACTIVITIES),
    )


@bp.get("/activities/<int:activity_id>/registrations/export.xlsx")
@login_required
def export_activity_registrations_excel(activity_id: int):
    activity = Activity.query.filter_by(id=activity_id).first()
    if not activity:
        flash("活动不存在", "error")
        return redirect(url_for("admin.activities_page"))
    try:
        from openpyxl import Workbook
    except Exception:
        flash("导出失败：缺少 openpyxl 依赖，请先安装后重试。", "error")
        return redirect(url_for("admin.activity_registrations_page", activity_id=activity_id))

    rows = (
        EventRegistration.query.filter_by(activity_id=activity.id)
        .order_by(EventRegistration.created_at.asc(), EventRegistration.id.asc())
        .all()
    )

    export_rows: list[dict] = []
    by_route: dict[str, list[dict]] = defaultdict(list)
    for item in rows:
        note_map = _parse_registration_notes(item.notes or "")
        route_label = note_map.get("route_label", "")
        route_name = note_map.get("route_name", "")
        route_display = "未指定路线"
        if route_label and route_name:
            route_display = f"{route_label} · {route_name}"
        elif route_name:
            route_display = route_name
        elif route_label:
            route_display = route_label
        image_consent_display = "同意" if note_map.get("image_consent") == "1" else "未同意"
        row = {
            "route": route_display,
            "name": item.name,
            "image_consent": image_consent_display,
        }
        export_rows.append(row)
        by_route[route_display].append(row)

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "总表"
    headers = ["路线", "姓名", "是否同意影像"]
    ws_all.append(headers)
    for row in export_rows:
        ws_all.append([row["route"], row["name"], row["image_consent"]])
    ws_all.column_dimensions["A"].width = 28
    ws_all.column_dimensions["B"].width = 18
    ws_all.column_dimensions["C"].width = 16

    used_sheet_names = {"总表"}
    for route_key, route_rows in by_route.items():
        title = _excel_safe_sheet_name(route_key, used_sheet_names)
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for row in route_rows:
            ws.append([row["route"], row["name"], row["image_consent"]])
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"activity_{activity.id}_registrations.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp.get("/kit-preorders")
@login_required
def merch_preorders_page():
    page = max(1, request.args.get("page", default=1, type=int))
    pagination = MerchPreorderBatch.query.order_by(
        MerchPreorderBatch.updated_at.desc(),
        MerchPreorderBatch.id.desc(),
    ).paginate(page=page, per_page=20, error_out=False)
    for batch in pagination.items:
        _apply_merch_batch_status(batch)
    count_map = {item.id: _merch_registration_count(item.id) for item in pagination.items}
    return render_template(
        "manage_kit_preorders.html",
        batches=pagination.items,
        pagination=pagination,
        count_map=count_map,
        status_labels=MERCH_BATCH_STATUS_LABELS,
        can_edit=can_write_page(g.current_user, PAGE_KIT_PREORDERS),
        can_admin=can_admin_page(g.current_user, PAGE_KIT_PREORDERS),
    )


@bp.get("/kit-preorders/new")
@login_required
def merch_preorder_new_page():
    default_start_date, default_deadline_date = _default_merch_preorder_dates()
    return render_template(
        "manage_kit_preorder_form.html",
        batch=None,
        gallery_images=[],
        size_images=[],
        default_start_date=default_start_date,
        default_deadline_date=default_deadline_date,
        status_labels=MERCH_BATCH_STATUS_LABELS,
        can_edit=can_write_page(g.current_user, PAGE_KIT_PREORDERS),
        can_admin=can_admin_page(g.current_user, PAGE_KIT_PREORDERS),
    )


@bp.get("/kit-preorders/<int:batch_id>/edit")
@login_required
def merch_preorder_edit_page(batch_id: int):
    batch = MerchPreorderBatch.query.filter_by(id=batch_id).first()
    if not batch:
        flash("预报名批次不存在", "error")
        return redirect(url_for("admin.merch_preorders_page"))
    _apply_merch_batch_status(batch)
    default_start_date, default_deadline_date = _default_merch_preorder_dates()
    return render_template(
        "manage_kit_preorder_form.html",
        batch=batch,
        gallery_images=[item for item in batch.images if item.image_kind == "gallery"],
        size_images=[item for item in batch.images if item.image_kind == "size_chart"],
        default_start_date=default_start_date,
        default_deadline_date=default_deadline_date,
        status_labels=MERCH_BATCH_STATUS_LABELS,
        can_edit=can_write_page(g.current_user, PAGE_KIT_PREORDERS),
        can_admin=can_admin_page(g.current_user, PAGE_KIT_PREORDERS),
    )


@bp.get("/kit-preorders/<int:batch_id>/registrations")
@login_required
def merch_preorder_registrations_page(batch_id: int):
    batch = MerchPreorderBatch.query.filter_by(id=batch_id).first()
    if not batch:
        flash("预报名批次不存在", "error")
        return redirect(url_for("admin.merch_preorders_page"))
    page = max(1, request.args.get("page", default=1, type=int))
    status_filter = (request.args.get("status") or "").strip()
    query = MerchPreorderRegistration.query.filter_by(batch_id=batch.id)
    if status_filter in MERCH_ORDER_STATUSES:
        query = query.filter(MerchPreorderRegistration.status == status_filter)
    pagination = query.order_by(
        MerchPreorderRegistration.created_at.desc(),
        MerchPreorderRegistration.id.desc(),
    ).paginate(page=page, per_page=50, error_out=False)
    return render_template(
        "manage_kit_preorder_registrations.html",
        batch=batch,
        registrations=pagination.items,
        pagination=pagination,
        status_filter=status_filter,
        order_statuses=MERCH_ORDER_STATUSES,
        status_labels=MERCH_ORDER_STATUS_LABELS,
        registration_count=_merch_registration_count(batch.id),
        can_edit=can_write_page(g.current_user, PAGE_KIT_PREORDERS),
    )


@bp.get("/kit-preorders/<int:batch_id>/registrations/export.xlsx")
@login_required
def export_merch_preorder_registrations_excel(batch_id: int):
    batch = MerchPreorderBatch.query.filter_by(id=batch_id).first()
    if not batch:
        flash("预报名批次不存在", "error")
        return redirect(url_for("admin.merch_preorders_page"))
    try:
        from openpyxl import Workbook
    except Exception:
        flash("导出失败：缺少 openpyxl 依赖，请先安装后重试。", "error")
        return redirect(url_for("admin.merch_preorder_registrations_page", batch_id=batch.id))

    rows = (
        MerchPreorderRegistration.query.filter_by(batch_id=batch.id)
        .order_by(MerchPreorderRegistration.created_at.asc(), MerchPreorderRegistration.id.asc())
        .all()
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "预报名"
    headers = ["姓名", "学号", "手机号", "性别", "尺码", "件数", "状态", "备注", "提交时间"]
    sheet.append(headers)
    for item in rows:
        sheet.append(
            [
                item.name,
                item.student_id,
                item.phone,
                item.gender,
                item.size,
                item.quantity,
                MERCH_ORDER_STATUS_LABELS.get(item.status, item.status),
                item.notes,
                _to_local_time(item.created_at).strftime("%Y-%m-%d %H:%M") if item.created_at else "",
            ]
        )
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 28)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"kit_preorder_{batch.id}_registrations.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.get("/announcements")
@login_required
def announcements_page():
    page = max(1, request.args.get("page", default=1, type=int))
    pagination = Announcement.query.order_by(
        Announcement.is_pinned.desc(),
        Announcement.sort_order.desc(),
        db.func.coalesce(Announcement.published_at, Announcement.updated_at).desc(),
    ).paginate(page=page, per_page=20, error_out=False)
    return render_template(
        "manage_announcements.html",
        announcements=pagination.items,
        pagination=pagination,
        can_edit=can_write_page(g.current_user, PAGE_ANNOUNCEMENTS),
        can_admin=can_admin_page(g.current_user, PAGE_ANNOUNCEMENTS),
    )


def _announcement_link_targets() -> list[dict]:
    activities = Activity.query.order_by(Activity.activity_time.desc(), Activity.id.desc()).limit(80).all()
    routes = Route.query.filter_by(is_deleted=False).order_by(Route.updated_at.desc(), Route.id.desc()).limit(80).all()
    batches = MerchPreorderBatch.query.order_by(
        MerchPreorderBatch.deadline_at.desc(),
        MerchPreorderBatch.updated_at.desc(),
        MerchPreorderBatch.id.desc(),
    ).limit(80).all()
    return [
        *[
            {"type": "活动", "label": item.title, "path": f"/events/{item.id}"}
            for item in activities
        ],
        *[
            {"type": "路线", "label": item.route_name, "path": f"/routes/{item.id}"}
            for item in routes
        ],
        *[
            {"type": "预定", "label": item.title, "path": f"/kit/{item.id}"}
            for item in batches
        ],
    ]


@bp.get("/announcements/new")
@login_required
def announcement_new_page():
    routes = Route.query.filter_by(is_deleted=False).order_by(Route.updated_at.desc()).all()
    activities = Activity.query.order_by(Activity.activity_time.desc()).all()
    return render_template(
        "manage_announcement_form.html",
        announcement=None,
        routes=routes,
        activities=activities,
        announcement_link_targets=_announcement_link_targets(),
        default_published_at=_next_local_hour(),
        can_edit=can_write_page(g.current_user, PAGE_ANNOUNCEMENTS),
        can_admin=can_admin_page(g.current_user, PAGE_ANNOUNCEMENTS),
    )


@bp.get("/announcements/<int:announcement_id>/edit")
@login_required
def announcement_edit_page(announcement_id: int):
    announcement = Announcement.query.filter_by(id=announcement_id).first()
    if not announcement:
        flash("公告不存在", "error")
        return redirect(url_for("admin.announcements_page"))
    return render_template(
        "manage_announcement_form.html",
        announcement=announcement,
        routes=Route.query.filter_by(is_deleted=False).order_by(Route.updated_at.desc()).all(),
        activities=Activity.query.order_by(Activity.activity_time.desc()).all(),
        announcement_link_targets=_announcement_link_targets(),
        default_published_at=_next_local_hour(),
        can_edit=can_write_page(g.current_user, PAGE_ANNOUNCEMENTS),
        can_admin=can_admin_page(g.current_user, PAGE_ANNOUNCEMENTS),
    )


@bp.get("/users")
@login_required
def users_page():
    page = max(1, request.args.get("page", default=1, type=int))
    pagination = User.query.order_by(User.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    for user in pagination.items:
        ensure_user_page_permissions(user)
    db.session.commit()
    return render_template(
        "manage_users.html",
        users=pagination.items,
        pagination=pagination,
        role_labels=ROLE_LABELS,
        page_labels=PAGE_PERMISSION_LABELS,
        permission_level_labels=PERMISSION_LEVEL_LABELS,
        page_keys=PAGE_KEYS,
        user_page_permissions=_user_page_permissions,
        can_admin_accounts=can_admin_page(g.current_user, PAGE_ACCOUNTS),
    )


@bp.get("/users/new")
@login_required
def user_new_page():
    return render_template(
        "manage_user_form.html",
        user=None,
        roles=ROLES,
        role_labels=ROLE_LABELS,
        page_labels=PAGE_PERMISSION_LABELS,
        permission_level_labels=PERMISSION_LEVEL_LABELS,
        permission_levels=PAGE_PERMISSION_LEVELS,
        page_keys=PAGE_KEYS,
        page_permissions=_default_page_permissions_for_role(ROLE_CONTENT_ADMIN),
    )


@bp.get("/users/<int:user_id>/edit")
@login_required
def user_edit_page(user_id: int):
    user = User.query.filter_by(id=user_id).first()
    if not user:
        flash("管理员不存在", "error")
        return redirect(url_for("admin.users_page"))
    ensure_user_page_permissions(user)
    db.session.commit()
    return render_template(
        "manage_user_form.html",
        user=user,
        roles=ROLES,
        role_labels=ROLE_LABELS,
        page_labels=PAGE_PERMISSION_LABELS,
        permission_level_labels=PERMISSION_LEVEL_LABELS,
        permission_levels=PAGE_PERMISSION_LEVELS,
        page_keys=PAGE_KEYS,
        page_permissions=_user_page_permissions(user),
    )


def _route_from_form(route: Route | None = None) -> dict:
    raw_difficulty = (request.form.get("difficulty") or (route.difficulty if route else "3")).strip()
    normalized_difficulty = _normalize_difficulty(raw_difficulty)
    return {
        "route_name": (request.form.get("route_name") or (route.route_name if route else "")).strip(),
        "distance_km": route.distance_km if route else 0.0,
        "difficulty": normalized_difficulty,
        "category": "cycling",
        "description": (request.form.get("description") or (route.description if route else "")).strip(),
        "status": (request.form.get("status") or (route.status if route else STATUS_PENDING_REVIEW)).strip(),
        "suggested_duration_hours": parse_distance(
            request.form.get("suggested_duration_hours") or (route.suggested_duration_hours if route else "0")
        ),
        "supply_points": (request.form.get("supply_points") or (route.supply_points if route else "")).strip(),
        "risk_warning": (request.form.get("risk_warning") or (route.risk_warning if route else "")).strip(),
    }


def _compute_route_stats(gpx_path: Path) -> dict:
    _points, stats, _elevation_profile = parse_gpx_points_and_stats(gpx_path)
    return {
        "distance_km": float(stats.get("distance_km") or 0.0),
        "ascent_m": stats.get("ascent_m"),
        "descent_m": stats.get("descent_m"),
        "min_ele_m": stats.get("min_ele_m"),
        "max_ele_m": stats.get("max_ele_m"),
    }


def _suggested_speed_for_difficulty(difficulty: str | None) -> float:
    speeds = {
        "1": 12.0,
        "2": 15.0,
        "3": 18.0,
        "4": 22.0,
        "5": 25.0,
    }
    normalized = _normalize_difficulty(difficulty or "3")
    return speeds.get(normalized, 18.0)


def _calculate_suggested_duration_hours(distance_km: float | None, ascent_m: float | None, difficulty: str | None) -> float:
    speed = _suggested_speed_for_difficulty(difficulty)
    distance_hours = float(distance_km or 0.0) / speed
    ascent_hours = float(ascent_m or 0.0) / 600.0
    return round((distance_hours + ascent_hours) * 1.15, 1)


ROUTE_MANUAL_STAT_FIELDS = {
    "distance_km": {"label": "里程", "unit": "km", "digits": 1},
    "ascent_m": {"label": "累计爬升", "unit": "m", "digits": 0},
    "descent_m": {"label": "累计下降", "unit": "m", "digits": 0},
    "min_ele_m": {"label": "最低海拔", "unit": "m", "digits": 0},
    "max_ele_m": {"label": "最高海拔", "unit": "m", "digits": 0},
    "suggested_duration_hours": {"label": "建议用时", "unit": "h", "digits": 1},
}


def _clean_manual_stat_overrides(raw: dict | None) -> dict:
    overrides = {}
    for key in ROUTE_MANUAL_STAT_FIELDS:
        if not raw or key not in raw:
            continue
        try:
            overrides[key] = float(raw[key])
        except (TypeError, ValueError):
            continue
    return overrides


def _route_manual_stat_overrides(route: Route | None) -> dict:
    if not route or not route.manual_stat_overrides:
        return {}
    try:
        raw = json.loads(route.manual_stat_overrides)
    except (TypeError, ValueError):
        return {}
    if not isinstance(raw, dict):
        return {}
    return _clean_manual_stat_overrides(raw)


def _manual_stat_overrides_from_form(route: Route | None = None) -> tuple[dict, str | None]:
    overrides = {}
    for key in ROUTE_MANUAL_STAT_FIELDS:
        raw = (request.form.get(f"manual_{key}") or "").strip()
        if not raw:
            continue
        value = parse_distance(raw)
        if value is None:
            return {}, ROUTE_MANUAL_STAT_FIELDS[key]["label"]
        overrides[key] = value
    return overrides, None


def _format_manual_stat_value(key: str, value: float) -> str:
    field = ROUTE_MANUAL_STAT_FIELDS[key]
    digits = field["digits"]
    number = f"{float(value):.{digits}f}"
    return f"{field['label']}已经人工指定为 {number} {field['unit']}"


def _manual_stat_summary(overrides: dict) -> str:
    return "；".join(_format_manual_stat_value(key, overrides[key]) for key in ROUTE_MANUAL_STAT_FIELDS if key in overrides)


def _apply_route_manual_stat_overrides(route: Route, overrides: dict, difficulty: str | None) -> None:
    for key in ("distance_km", "ascent_m", "descent_m", "min_ele_m", "max_ele_m"):
        if key in overrides:
            setattr(route, key, overrides[key])

    if "suggested_duration_hours" in overrides:
        route.suggested_duration_hours = overrides["suggested_duration_hours"]
    else:
        route.suggested_duration_hours = _calculate_suggested_duration_hours(
            route.distance_km,
            route.ascent_m,
            difficulty,
        )


def _resolve_route_gpx_path(route: Route) -> Path:
    upload_folder = Path(current_app.config["UPLOAD_FOLDER"])
    filename = route.gpx_filename or ""
    candidates = [filename]

    match = re.match(r"^\d{14}_(.+)$", filename)
    if match:
        candidates.append(match.group(1))

    for candidate in list(candidates):
        if "_" in candidate:
            candidates.append(candidate.replace("_", " "))

    seen = set()
    for candidate in candidates:
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        path = upload_folder / candidate
        if path.exists() and path.is_file():
            return path

    return upload_folder / filename


def _route_stats_payload(route: Route) -> dict:
    manual_overrides = _route_manual_stat_overrides(route)
    return {
        "distance_km": round(float(route.distance_km or 0), 1),
        "ascent_m": round(float(route.ascent_m or 0)),
        "descent_m": round(float(route.descent_m or 0)),
        "min_ele_m": round(float(route.min_ele_m or 0)),
        "max_ele_m": round(float(route.max_ele_m or 0)),
        "suggested_duration_hours": round(float(route.suggested_duration_hours or 0), 1),
        "manual_summary": _manual_stat_summary(manual_overrides),
    }


def _apply_route_stats(route: Route, gpx_path: Path) -> dict:
    stats = _compute_route_stats(gpx_path)
    route.distance_km = stats["distance_km"]
    route.ascent_m = stats["ascent_m"]
    route.descent_m = stats["descent_m"]
    route.min_ele_m = stats["min_ele_m"]
    route.max_ele_m = stats["max_ele_m"]
    return stats


def _normalize_difficulty(raw: str) -> str:
    if raw in {"1", "2", "3", "4", "5"}:
        return raw
    # Backward compatibility for legacy values.
    # We map `hard` to 4 to avoid over-stating routes as 5-star by default.
    if raw == "easy":
        return "2"
    if raw == "medium":
        return "3"
    if raw == "hard":
        return "4"
    return "3"


def _parse_activity_time(value: str | None) -> datetime | None:
    raw = (value or "").strip()
    if not raw:
        return None

    try:
        parsed = datetime.fromisoformat(raw)
    except ValueError:
        parsed = None

    if parsed is None:
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
            try:
                parsed = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue

    if parsed is None:
        return None

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=SH_TZ)

    return parsed.astimezone(timezone.utc)


def _parse_activity_date(value: str | None) -> date | None:
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        parsed = _parse_activity_time(raw)
        return _to_local_time(parsed).date() if parsed else None


def _parse_activity_option_time(activity_date: date | None, value: str | None) -> datetime | None:
    raw = (value or "").strip()
    if not raw:
        return None
    if "T" in raw or re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}", raw):
        return _parse_activity_time(raw)
    if activity_date is None:
        return None
    try:
        parsed_time = time.fromisoformat(raw)
    except ValueError:
        return None
    return datetime.combine(activity_date, parsed_time).replace(tzinfo=SH_TZ).astimezone(timezone.utc)


def _parse_int_field(value: str | None) -> int | None:
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        parsed = int(raw)
    except (TypeError, ValueError):
        return None
    return parsed if parsed >= 0 else None


def _next_month_same_or_next_existing(today_value: date) -> date:
    year = today_value.year
    month = today_value.month + 1
    if month > 12:
        year += 1
        month = 1
    try:
        return date(year, month, today_value.day)
    except ValueError:
        if month == 12:
            return date(year + 1, 1, 1)
        return date(year, month + 1, 1)


def _default_merch_preorder_dates() -> tuple[date, date]:
    today_value = _to_local_time(utcnow()).date()
    return today_value, _next_month_same_or_next_existing(today_value)


def _next_local_hour(value: datetime | None = None) -> datetime:
    local_value = _to_local_time(value or utcnow())
    if local_value.minute or local_value.second or local_value.microsecond:
        local_value = local_value.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
    else:
        local_value = local_value.replace(second=0, microsecond=0)
    return local_value


def _parse_merch_date(value: str | None, *, end_of_day: bool = False) -> datetime | None:
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        if "T" in raw:
            parsed_dt = _parse_activity_time(raw)
            parsed_date = _to_local_time(parsed_dt).date() if parsed_dt else None
        else:
            parsed_date = date.fromisoformat(raw)
    except ValueError:
        return None
    if parsed_date is None:
        return None
    local_time = time(23, 59, 59) if end_of_day else time(0, 0)
    return datetime.combine(parsed_date, local_time).replace(tzinfo=SH_TZ).astimezone(timezone.utc)


def _apply_merch_batch_status(batch: MerchPreorderBatch) -> MerchPreorderBatch:
    batch.status = merch_batch_status_for_window(batch.start_at, batch.deadline_at)
    return batch


def _merch_batch_from_form(batch: MerchPreorderBatch | None = None) -> tuple[dict, str | None]:
    title = (request.form.get("title") or (batch.title if batch else "")).strip()
    default_start_date, default_deadline_date = _default_merch_preorder_dates()
    start_raw = request.form.get("start_date") or request.form.get("start_at")
    deadline_raw = request.form.get("deadline_date") or request.form.get("deadline_at")
    start_at = _parse_merch_date(start_raw) or (
        batch.start_at if batch else _parse_merch_date(default_start_date.isoformat())
    )
    deadline_at = _parse_merch_date(deadline_raw, end_of_day=True) or (
        batch.deadline_at if batch else _parse_merch_date(default_deadline_date.isoformat(), end_of_day=True)
    )
    price_min = _parse_int_field(request.form.get("price_min"))
    price_max = _parse_int_field(request.form.get("price_max"))
    price_note = (request.form.get("price_note") or "").strip()
    description = (request.form.get("description") or "").strip()
    size_note = (request.form.get("size_note") or "").strip()
    is_visible = (request.form.get("is_visible") or "0").strip() == "1"

    if not title:
        return {}, "预报名标题不能为空"
    if deadline_at is None:
        return {}, "请填写最终截止日期"
    if start_at and start_at > deadline_at:
        return {}, "开始日期不能晚于最终截止日期"
    if price_min is not None and price_max is not None and price_min > price_max:
        return {}, "价格下限不能高于价格上限"

    if not price_note:
        price_note = "最终价格将根据实际预报名人数确定，但不会超过上限。"
    if not size_note:
        size_note = "尺码偏小\n不追求极致贴身的非公路车玩家，建议买大一码或大两码"

    return {
        "title": title,
        "status": merch_batch_status_for_window(start_at, deadline_at),
        "start_at": start_at,
        "deadline_at": deadline_at,
        "price_min": price_min,
        "price_max": price_max,
        "price_note": price_note,
        "description": description,
        "size_note": size_note,
        "is_visible": is_visible,
    }, None


def _announcement_from_form(announcement: Announcement | None = None) -> dict:
    title = (request.form.get("title") or (announcement.title if announcement else "")).strip()
    content = (request.form.get("content") or (announcement.content if announcement else "")).strip()
    status = (request.form.get("status") or (announcement.status if announcement else CONTENT_STATUS_DRAFT)).strip()
    if status not in {CONTENT_STATUS_DRAFT, CONTENT_STATUS_PUBLISHED, CONTENT_STATUS_OFFLINE}:
        status = CONTENT_STATUS_DRAFT
    sort_order_raw = (request.form.get("sort_order") or (announcement.sort_order if announcement else "0")).strip()
    try:
        sort_order = int(sort_order_raw)
    except (TypeError, ValueError):
        sort_order = 0
    is_pinned = (request.form.get("is_pinned") or "0").strip() == "1"
    published_at = _parse_activity_time(request.form.get("published_at"))
    offline_at = _parse_activity_time(request.form.get("offline_at"))
    if published_at and offline_at and offline_at <= published_at:
        offline_at = published_at + timedelta(minutes=1)
    preserve_associations = (request.form.get("association_mode") or "").strip() == "preserve"
    activity_ids = None if preserve_associations and "activity_ids" not in request.form else [
        item for item in request.form.getlist("activity_ids") if str(item).strip()
    ]
    route_ids = None if preserve_associations and "route_ids" not in request.form else [
        item for item in request.form.getlist("route_ids") if str(item).strip()
    ]
    return {
        "title": title,
        "content": content,
        "status": status,
        "sort_order": sort_order,
        "is_pinned": is_pinned,
        "published_at": published_at,
        "offline_at": offline_at,
        "activity_ids": activity_ids,
        "route_ids": route_ids,
    }


def _activity_route_options_from_form() -> list[dict]:
    items: list[dict] = []
    activity_date = _parse_activity_date(request.form.get("activity_date"))
    for sort_order, (level_key, level_label) in enumerate(ACTIVITY_ROUTE_LEVELS, start=1):
        raw = (request.form.get(f"route_option_{level_key}") or "").strip()
        if not raw:
            continue
        try:
            route_id = int(raw)
        except ValueError:
            continue
        participant_count = parse_distance(request.form.get(f"route_option_{level_key}_participants") or "0")
        option_time = _parse_activity_option_time(activity_date, request.form.get(f"route_option_{level_key}_time"))
        items.append(
            {
                "level_key": level_key,
                "level_label": level_label,
                "sort_order": sort_order,
                "route_id": route_id,
                "participant_count": int(participant_count or 0),
                "activity_time": option_time,
            }
        )
    return items


def _activity_registration_from_form(activity: Activity | None = None) -> tuple[bool, datetime | None, int | None, str | None]:
    _ = activity  # keep signature aligned for future extension
    needs_registration = (request.form.get("needs_registration") or "").strip() == "1"
    raw_deadline = request.form.get("registration_deadline")
    deadline = _parse_activity_time(raw_deadline) if raw_deadline is not None else None
    raw_limit = (request.form.get("registration_limit") or "").strip()
    if not needs_registration:
        return False, None, None, None
    if not raw_limit:
        return True, deadline, None, None
    try:
        limit = int(raw_limit)
    except (TypeError, ValueError):
        return True, deadline, None, "报名人数限制必须是整数"
    if limit <= 0:
        return True, deadline, None, "报名人数限制必须大于 0"
    return True, deadline, limit, None


def _validate_registration_deadline(
    needs_registration: bool,
    registration_deadline: datetime | None,
    activity_time: datetime | None,
) -> str | None:
    if not needs_registration:
        return None
    if registration_deadline is None:
        return "请填写报名截止时间"
    if activity_time is None:
        return "活动开始时间缺失，无法校验报名截止时间"
    if registration_deadline >= activity_time:
        return "报名截止时间必须早于活动开始时间"
    return None


def _sync_activity_route_options(activity: Activity, option_items: list[dict]) -> None:
    existing_options = ActivityRouteOption.query.filter_by(activity_id=activity.id).all()
    existing_by_level = {item.level_key: item for item in existing_options}
    next_levels = {item["level_key"] for item in option_items}

    for item in option_items:
        option = existing_by_level.get(item["level_key"])
        if option is None:
            db.session.add(
                ActivityRouteOption(
                    activity_id=activity.id,
                    route_id=item["route_id"],
                    level_key=item["level_key"],
                    level_label=item["level_label"],
                    activity_time=item.get("activity_time"),
                    participant_count=int(item.get("participant_count") or 0),
                    sort_order=item["sort_order"],
                )
            )
            continue
        option.route_id = item["route_id"]
        option.level_label = item["level_label"]
        option.activity_time = item.get("activity_time")
        option.participant_count = int(item.get("participant_count") or 0)
        option.sort_order = item["sort_order"]

    for option in existing_options:
        if option.level_key in next_levels:
            continue
        MediaAsset.query.filter_by(activity_route_option_id=option.id).update(
            {MediaAsset.activity_route_option_id: None},
            synchronize_session=False,
        )
        db.session.delete(option)


@bp.post("/routes/create")
@login_required
def create_route():
    payload = _route_from_form()
    manual_overrides, manual_error = _manual_stat_overrides_from_form()
    gpx_file = request.files.get("gpx_file")
    submit_action = (request.form.get("submit_action") or "save").strip()

    if not payload["route_name"] or not gpx_file:
        flash("参数错误：请填写路线名并上传 GPX 文件", "error")
        return redirect(url_for("admin.routes_page"))
    if payload["status"] not in ROUTE_STATUSES:
        flash("参数错误：状态无效", "error")
        return redirect(url_for("admin.routes_page"))
    if payload["suggested_duration_hours"] is None:
        flash("参数错误：预计用时格式错误", "error")
        return redirect(url_for("admin.routes_page"))
    if manual_error:
        flash(f"参数错误：{manual_error}人工指定值格式错误", "error")
        return redirect(url_for("admin.routes_page"))
    if not allowed_file(gpx_file.filename or "", {".gpx"}):
        flash("参数错误：仅支持 .gpx 文件", "error")
        return redirect(url_for("admin.routes_page"))
    if not file_size_ok(gpx_file, current_app.config.get("MAX_GPX_BYTES", 5 * 1024 * 1024)):
        flash("参数错误：GPX 文件过大", "error")
        return redirect(url_for("admin.routes_page"))

    filename, path = save_gpx_file(gpx_file)
    if not filename:
        flash("参数错误：仅支持 .gpx 文件", "error")
        return redirect(url_for("admin.routes_page"))

    try:
        computed_stats = _compute_route_stats(path)
    except Exception:
        if path and path.exists():
            path.unlink(missing_ok=True)
        flash("参数错误：GPX 解析失败，无法计算路线统计信息", "error")
        return redirect(url_for("admin.routes_page"))

    try:
        suggested_duration_hours = _calculate_suggested_duration_hours(
            computed_stats["distance_km"],
            computed_stats["ascent_m"],
            payload["difficulty"],
        )
        route = Route(
            route_name=payload["route_name"],
            gpx_filename=filename,
            uploaded_at=utcnow(),
            distance_km=computed_stats["distance_km"],
            difficulty=payload["difficulty"],
            category=payload["category"],
            description=payload["description"],
            status=payload["status"],
            is_active=(payload["status"] == STATUS_PUBLISHED),
            suggested_duration_hours=suggested_duration_hours,
            supply_points=payload["supply_points"],
            risk_warning=payload["risk_warning"],
            ascent_m=computed_stats["ascent_m"],
            descent_m=computed_stats["descent_m"],
            min_ele_m=computed_stats["min_ele_m"],
            max_ele_m=computed_stats["max_ele_m"],
            manual_stat_overrides=json.dumps(manual_overrides, ensure_ascii=False),
            created_by=g.current_user.id,
            updated_by=g.current_user.id,
        )
        _apply_route_manual_stat_overrides(route, manual_overrides, payload["difficulty"])
        db.session.add(route)
        db.session.flush()
        create_route_version(route, g.current_user.id, change_note="create")
        db.session.commit()
        write_audit_log(g.current_user.id, "route.create", "route", str(route.id), route.route_name)
        if submit_action == "save_view":
            flash("路线创建成功", "success")
            return redirect(url_for("admin.route_detail_manage", route_id=route.id))
        flash("路线创建成功", "success")
    except Exception:
        if path and path.exists():
            path.unlink(missing_ok=True)
        db.session.rollback()
        flash("系统错误：保存路线失败", "error")

    return redirect(url_for("admin.routes_page"))


@bp.post("/routes/<int:route_id>/update")
@login_required
def update_route(route_id: int):
    route = Route.query.filter_by(id=route_id, is_deleted=False).first()
    if not route:
        flash("路线不存在", "error")
        return redirect(url_for("admin.routes_page"))

    submit_action = (request.form.get("submit_action") or "save").strip()
    payload = _route_from_form(route)
    manual_overrides, manual_error = _manual_stat_overrides_from_form(route)
    if not payload["route_name"] or payload["status"] not in ROUTE_STATUSES:
        flash("参数错误：请检查必填项", "error")
        return redirect(url_for("admin.routes_page"))
    if payload["suggested_duration_hours"] is None:
        flash("参数错误：预计用时格式错误", "error")
        return redirect(url_for("admin.routes_page"))

    if manual_error:
        flash(f"参数错误：{manual_error}人工指定值格式错误", "error")
        return redirect(url_for("admin.routes_page"))

    before = route_snapshot(route)
    gpx_file = request.files.get("gpx_file")
    old_filename = route.gpx_filename
    saved_path = None
    stats_path = _resolve_route_gpx_path(route)
    if gpx_file and gpx_file.filename:
        if not allowed_file(gpx_file.filename, {".gpx"}):
            flash("参数错误：仅支持 .gpx 文件", "error")
            return redirect(url_for("admin.routes_page"))
        if not file_size_ok(gpx_file, current_app.config.get("MAX_GPX_BYTES", 5 * 1024 * 1024)):
            flash("参数错误：GPX 文件过大", "error")
            return redirect(url_for("admin.routes_page"))
        new_filename, saved_path = save_gpx_file(gpx_file)
        if not new_filename:
            flash("参数错误：仅支持 .gpx 文件", "error")
            return redirect(url_for("admin.routes_page"))
        route.gpx_filename = new_filename
        stats_path = saved_path

    try:
        stats = _apply_route_stats(route, stats_path)
        route.route_name = payload["route_name"]
        route.difficulty = payload["difficulty"]
        route.category = payload["category"]
        route.description = payload["description"]
        route.status = payload["status"]
        route.is_active = payload["status"] == STATUS_PUBLISHED
        route.uploaded_at = utcnow()
        route.updated_by = g.current_user.id
        route.manual_stat_overrides = json.dumps(manual_overrides, ensure_ascii=False)
        _apply_route_manual_stat_overrides(route, manual_overrides, payload["difficulty"])
        route.supply_points = payload["supply_points"]
        route.risk_warning = payload["risk_warning"]
        create_route_version(route, g.current_user.id, change_note="update")
        db.session.commit()

        if old_filename != route.gpx_filename:
            old_path = Path(current_app.config["UPLOAD_FOLDER"]) / old_filename
            old_path.unlink(missing_ok=True)

        write_field_audit_log(g.current_user.id, "route", str(route.id), before, route_snapshot(route))
        write_audit_log(g.current_user.id, "route.update", "route", str(route.id), route.route_name)
        if submit_action == "save_view":
            flash("路线更新成功", "success")
            return redirect(url_for("admin.route_detail_manage", route_id=route.id))
        flash("路线更新成功", "success")
    except Exception:
        db.session.rollback()
        if saved_path and saved_path.exists():
            saved_path.unlink(missing_ok=True)
        flash("系统错误：更新路线失败", "error")

    return redirect(url_for("admin.routes_page"))


@bp.post("/routes/<int:route_id>/recalculate-stats")
@login_required
def recalculate_route_stats(route_id: int):
    route = Route.query.filter_by(id=route_id, is_deleted=False).first()
    wants_json = request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html
    if not route:
        if wants_json:
            return jsonify({"ok": False, "message": "路线不存在"}), 404
        flash("路线不存在", "error")
        return redirect(url_for("admin.routes_page"))

    gpx_path = _resolve_route_gpx_path(route)
    if not gpx_path.exists() or not gpx_path.is_file():
        if wants_json:
            return jsonify({"ok": False, "message": "GPX 文件不存在，无法更新统计"}), 404
        flash("GPX 文件不存在，无法更新统计", "error")
        return redirect(url_for("admin.route_edit_page", route_id=route_id))

    manual_overrides, manual_error = _manual_stat_overrides_from_form(route)
    if manual_error:
        if wants_json:
            return jsonify({"ok": False, "message": f"{manual_error}人工指定值格式错误"}), 400
        flash(f"参数错误：{manual_error}人工指定值格式错误", "error")
        return redirect(url_for("admin.route_edit_page", route_id=route_id))

    before = route_snapshot(route)
    try:
        stats = _apply_route_stats(route, gpx_path)
        difficulty_for_duration = request.form.get("difficulty") or route.difficulty
        route.manual_stat_overrides = json.dumps(manual_overrides, ensure_ascii=False)
        _apply_route_manual_stat_overrides(route, manual_overrides, difficulty_for_duration)
        route.updated_by = g.current_user.id
        route.updated_at = utcnow()
        create_route_version(route, g.current_user.id, change_note="recalculate_stats")
        db.session.commit()
        write_field_audit_log(g.current_user.id, "route", str(route.id), before, route_snapshot(route))
        write_audit_log(
            g.current_user.id,
            "route.recalculate_stats",
            "route",
            str(route.id),
            f"distance_km={stats['distance_km']}",
        )
        if wants_json:
            return jsonify({
                "ok": True,
                "message": "已根据 GPX 自动更新里程与爬升统计",
                "stats": _route_stats_payload(route),
            })
        flash("已根据 GPX 自动更新里程与爬升统计", "success")
    except Exception:
        db.session.rollback()
        if wants_json:
            return jsonify({"ok": False, "message": "统计更新失败：GPX 解析异常"}), 400
        flash("统计更新失败：GPX 解析异常", "error")
    return redirect(url_for("admin.route_edit_page", route_id=route_id))


@bp.post("/routes/<int:route_id>/delete")
@login_required
def delete_route(route_id: int):
    route = Route.query.filter_by(id=route_id).first()
    if not route:
        flash("路线不存在", "error")
        return redirect(url_for("admin.routes_page"))

    try:
        route.is_deleted = True
        route.deleted_at = utcnow()
        route.deleted_by = g.current_user.id
        route.status = STATUS_OFFLINE
        route.is_active = False
        route.updated_by = g.current_user.id
        create_route_version(route, g.current_user.id, change_note="soft_delete")
        db.session.commit()
        write_audit_log(g.current_user.id, "route.soft_delete", "route", str(route_id), route.gpx_filename)
        flash("路线已移入回收站", "success")
    except Exception:
        db.session.rollback()
        flash("操作失败", "error")

    return redirect(url_for("admin.routes_page"))


@bp.post("/routes/<int:route_id>/restore")
@login_required
def restore_route(route_id: int):
    route = Route.query.filter_by(id=route_id, is_deleted=True).first()
    if not route:
        flash("回收站中未找到路线", "error")
        return redirect(url_for("admin.routes_page"))

    route.is_deleted = False
    route.deleted_at = None
    route.deleted_by = None
    route.status = STATUS_DRAFT
    route.is_active = False
    route.updated_by = g.current_user.id
    create_route_version(route, g.current_user.id, change_note="restore")
    db.session.commit()
    write_audit_log(g.current_user.id, "route.restore", "route", str(route.id), route.route_name)
    flash("路线已恢复（状态为草稿）", "success")
    return redirect(url_for("admin.routes_page"))


@bp.post("/routes/<int:route_id>/status")
@login_required
def update_status(route_id: int):
    route = Route.query.filter_by(id=route_id, is_deleted=False).first()
    if not route:
        flash("路线不存在", "error")
        return redirect(url_for("admin.routes_page"))

    status = (request.form.get("status") or "").strip()
    if status not in ROUTE_STATUSES:
        flash("状态无效", "error")
        return redirect(url_for("admin.routes_page"))

    before = route_snapshot(route)
    route.status = status
    route.is_active = status == STATUS_PUBLISHED
    route.updated_by = g.current_user.id
    create_route_version(route, g.current_user.id, change_note=f"status:{status}")
    db.session.commit()
    write_field_audit_log(g.current_user.id, "route", str(route_id), before, route_snapshot(route))
    write_audit_log(g.current_user.id, "route.status", "route", str(route_id), status)
    flash("状态更新成功", "success")
    return redirect(url_for("admin.routes_page"))


@bp.post("/routes/<int:route_id>/rollback")
@login_required
def rollback_route(route_id: int):
    route = Route.query.filter_by(id=route_id).first()
    if not route:
        flash("路线不存在", "error")
        return redirect(url_for("admin.routes_page"))

    version_no_raw = (request.form.get("version_no") or "").strip()
    try:
        version_no = int(version_no_raw)
    except ValueError:
        flash("版本号无效", "error")
        return redirect(url_for("admin.routes_page"))

    version = RouteVersion.query.filter_by(route_id=route_id, version_no=version_no).first()
    if not version:
        flash("目标版本不存在", "error")
        return redirect(url_for("admin.routes_page"))

    try:
        rollback_route_to_version(route, version, g.current_user.id)
        db.session.commit()
        write_audit_log(g.current_user.id, "route.rollback", "route", str(route.id), f"to_v{version_no}")
        flash(f"已回滚到版本 {version_no}", "success")
    except ValueError as exc:
        db.session.rollback()
        if str(exc).startswith("gpx_not_found:"):
            flash("回滚失败：对应 GPX 文件不存在", "error")
        else:
            flash("回滚失败：版本数据异常", "error")
    return redirect(url_for("admin.routes_page"))


@bp.post("/feedback/<int:feedback_id>/review")
@login_required
def review_feedback(feedback_id: int):
    feedback = RouteFeedback.query.filter_by(id=feedback_id).first()
    if not feedback:
        flash("反馈不存在", "error")
        return redirect(url_for("admin.dashboard"))

    status = (request.form.get("status") or "").strip()
    note = (request.form.get("reviewer_note") or "").strip()
    if status not in (FEEDBACK_APPROVED, FEEDBACK_REJECTED):
        flash("审核状态无效", "error")
        return redirect(url_for("admin.dashboard"))

    feedback.status = status
    feedback.reviewer_note = note
    feedback.reviewer_id = g.current_user.id
    feedback.reviewed_at = utcnow()
    db.session.commit()

    write_audit_log(g.current_user.id, "feedback.review", "route_feedback", str(feedback.id), status)
    flash("审核完成", "success")
    next_page = (request.form.get("next") or "").strip()
    if next_page == "feedback":
        return redirect(url_for("admin.feedback_page"))
    return redirect(url_for("admin.dashboard"))


@bp.post("/feedback/<int:feedback_id>/reopen")
@login_required
def reopen_feedback(feedback_id: int):
    feedback = RouteFeedback.query.filter_by(id=feedback_id).first()
    if not feedback:
        flash("反馈不存在", "error")
        return redirect(url_for("admin.feedback_page"))

    feedback.status = FEEDBACK_PENDING
    feedback.reviewer_note = ""
    feedback.reviewer_id = None
    feedback.reviewed_at = None
    db.session.commit()
    write_audit_log(g.current_user.id, "feedback.reopen", "route_feedback", str(feedback.id), "reopen_to_pending")
    flash("已重新打开反馈，状态改为待审核", "success")
    return redirect(url_for("admin.feedback_page"))


@bp.post("/feedback/<int:feedback_id>/delete")
@login_required
def delete_feedback(feedback_id: int):
    if not can_admin_page(g.current_user, PAGE_FEEDBACK):
        abort(403)
    feedback = RouteFeedback.query.filter_by(id=feedback_id).first()
    if not feedback:
        flash("反馈不存在", "error")
        return redirect(url_for("admin.feedback_page"))

    route_id = feedback.route_id
    db.session.delete(feedback)
    db.session.commit()
    write_audit_log(g.current_user.id, "feedback.delete", "route_feedback", str(feedback_id), f"route={route_id}")
    flash("反馈已删除", "success")
    return redirect(url_for("admin.feedback_page"))


@bp.post("/activities/create")
@login_required
def create_activity():
    title = (request.form.get("title") or "").strip()
    submit_action = (request.form.get("submit_action") or "save").strip()
    option_items = _activity_route_options_from_form()
    needs_registration, registration_deadline, registration_limit, registration_error = _activity_registration_from_form()
    if registration_error:
        flash(registration_error, "error")
        return redirect(url_for("admin.activity_new_page"))
    selected_route_ids = [item["route_id"] for item in option_items]
    selected_route_ids = list(dict.fromkeys(selected_route_ids))

    if not title:
        flash("参数错误：活动标题不能为空", "error")
        return redirect(url_for("admin.activities_page"))

    activity = Activity(
        title=title,
        participant_count=0,
        needs_registration=needs_registration,
        registration_deadline=registration_deadline,
        registration_limit=registration_limit,
        insurance_qr_path=None,
        weather="",
        summary="",
        created_by=g.current_user.id,
    )

    selected_routes = (
        Route.query.filter(Route.id.in_(selected_route_ids), Route.is_deleted.is_(False)).all()
        if selected_route_ids
        else []
    )
    activity.routes = selected_routes

    db.session.add(activity)
    db.session.flush()
    valid_route_ids = {route.id for route in selected_routes}
    option_items = [item for item in option_items if item["route_id"] in valid_route_ids]
    option_times = [item.get("activity_time") for item in option_items if item.get("activity_time")]
    activity.activity_time = min(option_times) if option_times else activity.activity_time
    activity.participant_count = int(sum(int(item.get("participant_count") or 0) for item in option_items))
    registration_error = _validate_registration_deadline(
        needs_registration=activity.needs_registration,
        registration_deadline=activity.registration_deadline,
        activity_time=activity.activity_time,
    )
    if registration_error:
        db.session.rollback()
        flash(registration_error, "error")
        return redirect(url_for("admin.activity_new_page"))
    _sync_activity_route_options(activity, option_items)
    db.session.flush()
    option_map = {item.level_key: item for item in ActivityRouteOption.query.filter_by(activity_id=activity.id).all()}
    uploaded_count = 0
    for level_key, _label in ACTIVITY_ROUTE_LEVELS:
        option = option_map.get(level_key)
        if not option:
            continue
        uploaded_count += _save_activity_media(
            activity.id,
            request.files.getlist(f"media_files_{level_key}"),
            activity_route_option_id=option.id,
        )
    wechat_upload = request.files.get("wechat_qr_file") or request.files.get("insurance_qr_file")
    qr_name, _qr_path, qr_error = _save_activity_wechat_qr(activity.id, wechat_upload)
    if wechat_upload and (wechat_upload.filename or "").strip() and not qr_name:
        db.session.rollback()
        flash(_wechat_qr_upload_error_message(qr_error), "error")
        return redirect(url_for("admin.activity_new_page"))
    if qr_name:
        activity.insurance_qr_path = qr_name
    db.session.commit()
    write_audit_log(g.current_user.id, "activity.create", "activity", str(activity.id), activity.title)
    if uploaded_count > 0:
        flash(f"活动创建成功，已上传媒体文件 {uploaded_count} 个", "success")
    else:
        flash("活动创建成功", "success")
    if submit_action == "save_view":
        return redirect(url_for("web.activity_detail", activity_id=activity.id, source="manage"))
    return redirect(url_for("admin.activities_page"))


@bp.post("/activities/<int:activity_id>/update")
@login_required
def update_activity(activity_id: int):
    activity = Activity.query.filter_by(id=activity_id).first()
    if not activity:
        flash("活动不存在", "error")
        return redirect(url_for("admin.activities_page"))

    title = (request.form.get("title") or "").strip()
    submit_action = (request.form.get("submit_action") or "save").strip()
    if not title:
        flash("参数错误：活动标题不能为空", "error")
        return redirect(url_for("admin.activities_page"))

    option_items = _activity_route_options_from_form()
    needs_registration, registration_deadline, registration_limit, registration_error = _activity_registration_from_form(activity)
    if registration_error:
        flash(registration_error, "error")
        return redirect(url_for("admin.activity_edit_page", activity_id=activity.id))
    selected_route_ids = [item["route_id"] for item in option_items]
    selected_route_ids = list(dict.fromkeys(selected_route_ids))

    activity.title = title
    activity.needs_registration = needs_registration
    activity.registration_deadline = registration_deadline
    activity.registration_limit = registration_limit
    activity.weather = ""
    activity.summary = ""

    selected_routes = (
        Route.query.filter(Route.id.in_(selected_route_ids), Route.is_deleted.is_(False)).all()
        if selected_route_ids
        else []
    )
    activity.routes = selected_routes

    valid_route_ids = {route.id for route in selected_routes}
    option_items = [item for item in option_items if item["route_id"] in valid_route_ids]
    option_times = [item.get("activity_time") for item in option_items if item.get("activity_time")]
    activity.activity_time = min(option_times) if option_times else activity.activity_time
    activity.participant_count = int(sum(int(item.get("participant_count") or 0) for item in option_items))
    registration_error = _validate_registration_deadline(
        needs_registration=activity.needs_registration,
        registration_deadline=activity.registration_deadline,
        activity_time=activity.activity_time,
    )
    if registration_error:
        db.session.rollback()
        flash(registration_error, "error")
        return redirect(url_for("admin.activity_edit_page", activity_id=activity.id))
    _sync_activity_route_options(activity, option_items)
    db.session.flush()
    option_map = {item.level_key: item for item in ActivityRouteOption.query.filter_by(activity_id=activity.id).all()}
    uploaded_count = 0
    for level_key, _label in ACTIVITY_ROUTE_LEVELS:
        option = option_map.get(level_key)
        if not option:
            continue
        uploaded_count += _save_activity_media(
            activity.id,
            request.files.getlist(f"media_files_{level_key}"),
            activity_route_option_id=option.id,
        )
    clear_wechat_qr = (request.form.get("clear_wechat_qr") or request.form.get("clear_insurance_qr") or "").strip() == "1"
    old_qr_path = (activity.insurance_qr_path or "").strip()
    qr_cleanup_paths: list[Path] = []
    if clear_wechat_qr and old_qr_path:
        qr_cleanup_paths.append(Path(current_app.config["MEDIA_UPLOAD_FOLDER"]) / old_qr_path)
        activity.insurance_qr_path = None
    wechat_upload = request.files.get("wechat_qr_file") or request.files.get("insurance_qr_file")
    qr_name, _qr_path, qr_error = _save_activity_wechat_qr(activity.id, wechat_upload)
    if wechat_upload and (wechat_upload.filename or "").strip() and not qr_name:
        db.session.rollback()
        flash(_wechat_qr_upload_error_message(qr_error), "error")
        return redirect(url_for("admin.activity_edit_page", activity_id=activity.id))
    if qr_name:
        if old_qr_path:
            qr_cleanup_paths.append(Path(current_app.config["MEDIA_UPLOAD_FOLDER"]) / old_qr_path)
        activity.insurance_qr_path = qr_name
    db.session.commit()
    if qr_cleanup_paths:
        _cleanup_paths(qr_cleanup_paths)
    write_audit_log(g.current_user.id, "activity.update", "activity", str(activity.id), activity.title)
    if uploaded_count > 0:
        flash(f"活动更新成功，新增媒体文件 {uploaded_count} 个", "success")
    else:
        flash("活动更新成功", "success")
    if submit_action == "save_view":
        return redirect(url_for("web.activity_detail", activity_id=activity.id, source="manage"))
    return redirect(url_for("admin.activities_page"))


@bp.post("/activities/<int:activity_id>/copy")
@login_required
def copy_activity(activity_id: int):
    activity = Activity.query.filter_by(id=activity_id).first()
    if not activity:
        flash("活动不存在", "error")
        return redirect(url_for("admin.activities_page"))

    base_title = f"{activity.title} 副本"
    title = base_title[:128]
    counter = 2
    while Activity.query.filter_by(title=title).first():
        suffix = f" {counter}"
        title = f"{base_title[:128 - len(suffix)]}{suffix}"
        counter += 1

    copied = Activity(
        title=title,
        activity_time=activity.activity_time,
        needs_registration=activity.needs_registration,
        registration_deadline=activity.registration_deadline,
        registration_limit=activity.registration_limit,
        insurance_qr_path=None,
        participant_count=activity.participant_count,
        weather=activity.weather,
        summary=activity.summary,
        created_by=g.current_user.id,
    )
    copied.routes = list(activity.routes)
    db.session.add(copied)
    db.session.flush()

    for option in sorted(activity.route_options, key=lambda item: item.sort_order):
        db.session.add(
            ActivityRouteOption(
                activity_id=copied.id,
                route_id=option.route_id,
                level_key=option.level_key,
                level_label=option.level_label,
                activity_time=option.activity_time,
                participant_count=option.participant_count,
                sort_order=option.sort_order,
            )
        )

    db.session.commit()
    write_audit_log(g.current_user.id, "activity.copy", "activity", str(copied.id), f"from={activity.id}")
    flash("活动已复制，请检查时间、报名截止时间和路线信息", "success")
    return redirect(url_for("admin.activity_edit_page", activity_id=copied.id))


@bp.post("/activities/<int:activity_id>/delete")
@login_required
def delete_activity(activity_id: int):
    activity = Activity.query.filter_by(id=activity_id).first()
    if not activity:
        flash("活动不存在", "error")
        return redirect(url_for("admin.activities_page"))

    title = activity.title
    media_assets = MediaAsset.query.filter_by(activity_id=activity.id).all()
    media_dir = Path(current_app.config["MEDIA_UPLOAD_FOLDER"])
    media_paths = [media_dir / item.storage_path for item in media_assets if item.storage_path]
    if activity.insurance_qr_path:
        media_paths.append(media_dir / activity.insurance_qr_path)
    for item in media_assets:
        db.session.delete(item)
    db.session.delete(activity)
    db.session.commit()
    _cleanup_paths(media_paths)
    write_audit_log(g.current_user.id, "activity.delete", "activity", str(activity_id), title)
    flash("活动删除成功", "success")
    return redirect(url_for("admin.activities_page"))


@bp.post("/activities/<int:activity_id>/media/<int:asset_id>/delete")
@login_required
def delete_activity_media(activity_id: int, asset_id: int):
    activity = Activity.query.filter_by(id=activity_id).first()
    if not activity:
        flash("活动不存在", "error")
        return redirect(url_for("admin.activities_page"))

    asset = MediaAsset.query.filter_by(id=asset_id, activity_id=activity_id).first()
    if not asset:
        flash("媒体文件不存在", "error")
        return redirect(url_for("admin.activity_edit_page", activity_id=activity_id))

    target_path = Path(current_app.config["MEDIA_UPLOAD_FOLDER"]) / (asset.storage_path or "")
    db.session.delete(asset)
    db.session.commit()
    _cleanup_paths([target_path])
    write_audit_log(
        g.current_user.id,
        "activity.media.delete",
        "media_asset",
        str(asset_id),
        f"activity={activity_id}",
    )
    flash("媒体文件已删除", "success")
    return redirect(url_for("admin.activity_edit_page", activity_id=activity_id))


@bp.post("/activities/<int:activity_id>/media/<int:asset_id>/assign")
@login_required
def assign_activity_media(activity_id: int, asset_id: int):
    activity = Activity.query.filter_by(id=activity_id).first()
    if not activity:
        flash("活动不存在", "error")
        return redirect(url_for("admin.activities_page"))

    asset = MediaAsset.query.filter_by(id=asset_id, activity_id=activity_id).first()
    if not asset:
        flash("媒体文件不存在", "error")
        return redirect(url_for("admin.activity_edit_page", activity_id=activity_id))

    option_raw = (request.form.get("activity_route_option_id") or "").strip()
    option_id = None
    if option_raw:
        try:
            option_id = int(option_raw)
        except ValueError:
            flash("路线分配参数无效", "error")
            return redirect(url_for("admin.activity_edit_page", activity_id=activity_id))
        option = ActivityRouteOption.query.filter_by(id=option_id, activity_id=activity_id).first()
        if not option:
            flash("目标路线不存在或不属于当前活动", "error")
            return redirect(url_for("admin.activity_edit_page", activity_id=activity_id))

    asset.activity_route_option_id = option_id
    db.session.commit()
    write_audit_log(
        g.current_user.id,
        "activity.media.assign",
        "media_asset",
        str(asset_id),
        f"activity={activity_id},route_option={option_id or 'none'}",
    )
    flash("媒体路线归属已更新", "success")
    return redirect(url_for("admin.activity_edit_page", activity_id=activity_id))


@bp.post("/kit-preorders/create")
@login_required
def create_merch_preorder():
    payload, error = _merch_batch_from_form()
    if error:
        flash(error, "error")
        return redirect(url_for("admin.merch_preorder_new_page"))
    batch = MerchPreorderBatch(**payload, created_by=g.current_user.id, updated_by=g.current_user.id)
    db.session.add(batch)
    db.session.flush()
    uploaded_count = _save_merch_preorder_images(batch.id, request.files.getlist("gallery_images"), "gallery")
    uploaded_count += _save_merch_preorder_images(batch.id, request.files.getlist("size_chart_images"), "size_chart")
    db.session.commit()
    write_audit_log(g.current_user.id, "merch_preorder.create", "merch_preorder_batch", str(batch.id), batch.title)
    flash(f"预报名批次创建成功，已上传图片 {uploaded_count} 张", "success")
    if (request.form.get("submit_action") or "").strip() == "save_view":
        return redirect(url_for("web.kit_preorder_detail", batch_id=batch.id))
    return redirect(url_for("admin.merch_preorders_page"))


@bp.post("/kit-preorders/<int:batch_id>/update")
@login_required
def update_merch_preorder(batch_id: int):
    batch = MerchPreorderBatch.query.filter_by(id=batch_id).first()
    if not batch:
        flash("预报名批次不存在", "error")
        return redirect(url_for("admin.merch_preorders_page"))
    payload, error = _merch_batch_from_form(batch)
    if error:
        flash(error, "error")
        return redirect(url_for("admin.merch_preorder_edit_page", batch_id=batch.id))
    for key, value in payload.items():
        setattr(batch, key, value)
    batch.updated_by = g.current_user.id
    batch.updated_at = utcnow()
    uploaded_count = _save_merch_preorder_images(batch.id, request.files.getlist("gallery_images"), "gallery")
    uploaded_count += _save_merch_preorder_images(batch.id, request.files.getlist("size_chart_images"), "size_chart")
    db.session.commit()
    write_audit_log(g.current_user.id, "merch_preorder.update", "merch_preorder_batch", str(batch.id), batch.title)
    flash(f"预报名批次已更新，新增图片 {uploaded_count} 张", "success")
    if (request.form.get("submit_action") or "").strip() == "save_view":
        return redirect(url_for("web.kit_preorder_detail", batch_id=batch.id))
    return redirect(url_for("admin.merch_preorder_edit_page", batch_id=batch.id))


@bp.post("/kit-preorders/<int:batch_id>/visibility")
@login_required
def toggle_merch_preorder_visibility(batch_id: int):
    batch = MerchPreorderBatch.query.filter_by(id=batch_id).first()
    if not batch:
        flash("预报名批次不存在", "error")
        return redirect(url_for("admin.merch_preorders_page"))
    batch.is_visible = (request.form.get("is_visible") or "0").strip() == "1"
    batch.updated_by = g.current_user.id
    batch.updated_at = utcnow()
    db.session.commit()
    write_audit_log(
        g.current_user.id,
        "merch_preorder.visibility",
        "merch_preorder_batch",
        str(batch.id),
        "online" if batch.is_visible else "offline",
    )
    flash("预定批次已上线" if batch.is_visible else "预定批次已下线", "success")
    return redirect(url_for("admin.merch_preorders_page"))


@bp.post("/kit-preorders/<int:batch_id>/delete")
@login_required
def delete_merch_preorder(batch_id: int):
    batch = MerchPreorderBatch.query.filter_by(id=batch_id).first()
    if not batch:
        flash("预报名批次不存在", "error")
        return redirect(url_for("admin.merch_preorders_page"))
    title = batch.title
    image_paths = [
        Path(current_app.config["MEDIA_UPLOAD_FOLDER"]) / (image.storage_path or "")
        for image in batch.images
        if image.storage_path
    ]
    db.session.delete(batch)
    db.session.commit()
    _cleanup_paths(image_paths)
    write_audit_log(g.current_user.id, "merch_preorder.delete", "merch_preorder_batch", str(batch_id), title)
    flash("预定批次已删除", "success")
    return redirect(url_for("admin.merch_preorders_page"))


@bp.post("/kit-preorders/<int:batch_id>/images/<int:image_id>/delete")
@login_required
def delete_merch_preorder_image(batch_id: int, image_id: int):
    batch = MerchPreorderBatch.query.filter_by(id=batch_id).first()
    image = MerchPreorderImage.query.filter_by(id=image_id, batch_id=batch_id).first()
    if not batch or not image:
        flash("图片不存在", "error")
        return redirect(url_for("admin.merch_preorders_page"))
    target_path = Path(current_app.config["MEDIA_UPLOAD_FOLDER"]) / (image.storage_path or "")
    db.session.delete(image)
    db.session.commit()
    _cleanup_paths([target_path])
    write_audit_log(g.current_user.id, "merch_preorder.image.delete", "merch_preorder_image", str(image_id), f"batch={batch_id}")
    flash("图片已删除", "success")
    return redirect(url_for("admin.merch_preorder_edit_page", batch_id=batch.id))


@bp.post("/kit-preorders/<int:batch_id>/registrations/<int:registration_id>/status")
@login_required
def update_merch_preorder_registration_status(batch_id: int, registration_id: int):
    registration = MerchPreorderRegistration.query.filter_by(id=registration_id, batch_id=batch_id).first()
    if not registration:
        flash("预报名记录不存在", "error")
        return redirect(url_for("admin.merch_preorders_page"))
    status = (request.form.get("status") or "").strip()
    if status not in MERCH_ORDER_STATUSES:
        flash("记录状态无效", "error")
        return redirect(url_for("admin.merch_preorder_registrations_page", batch_id=batch_id))
    registration.status = status
    if status == MERCH_ORDER_CANCELLED and not registration.cancelled_at:
        registration.cancelled_at = utcnow()
    registration.updated_at = utcnow()
    db.session.commit()
    write_audit_log(
        g.current_user.id,
        "merch_preorder.registration.status",
        "merch_preorder_registration",
        str(registration.id),
        status,
    )
    return redirect(url_for("admin.merch_preorder_registrations_page", batch_id=batch_id))


@bp.post("/announcements/create")
@login_required
def create_announcement():
    payload = _announcement_from_form()
    if not payload["title"]:
        flash("参数错误：公告标题不能为空", "error")
        return redirect(url_for("admin.announcements_page"))

    announcement = Announcement(
        title=payload["title"],
        content=payload["content"],
        status=payload["status"],
        is_pinned=payload["is_pinned"],
        sort_order=payload["sort_order"],
        created_by=g.current_user.id,
        updated_by=g.current_user.id,
    )
    selected_activities = (
        Activity.query.filter(Activity.id.in_(payload["activity_ids"])).all()
        if payload["activity_ids"]
        else []
    )
    selected_routes = (
        Route.query.filter(Route.id.in_(payload["route_ids"]), Route.is_deleted.is_(False)).all()
        if payload["route_ids"]
        else []
    )
    announcement.activities = selected_activities
    announcement.routes = selected_routes
    if payload["status"] == CONTENT_STATUS_PUBLISHED:
        announcement.published_at = payload["published_at"] or utcnow()
    else:
        announcement.published_at = payload["published_at"]
    announcement.offline_at = payload["offline_at"]

    db.session.add(announcement)
    db.session.commit()
    write_audit_log(g.current_user.id, "announcement.create", "announcement", str(announcement.id), announcement.title)
    flash("公告创建成功", "success")
    if (request.form.get("submit_action") or "").strip() == "save_view":
        return redirect(url_for("web.announcement_detail", announcement_id=announcement.id))
    return redirect(url_for("admin.announcements_page"))


@bp.post("/announcements/<int:announcement_id>/update")
@login_required
def update_announcement(announcement_id: int):
    announcement = Announcement.query.filter_by(id=announcement_id).first()
    if not announcement:
        flash("公告不存在", "error")
        return redirect(url_for("admin.announcements_page"))

    payload = _announcement_from_form(announcement)
    if not payload["title"]:
        flash("参数错误：公告标题不能为空", "error")
        return redirect(url_for("admin.announcement_edit_page", announcement_id=announcement_id))

    announcement.title = payload["title"]
    announcement.content = payload["content"]
    announcement.status = payload["status"]
    announcement.is_pinned = payload["is_pinned"]
    announcement.sort_order = payload["sort_order"]
    announcement.updated_by = g.current_user.id
    if payload["activity_ids"] is not None:
        announcement.activities = (
            Activity.query.filter(Activity.id.in_(payload["activity_ids"])).all()
            if payload["activity_ids"]
            else []
        )
    if payload["route_ids"] is not None:
        announcement.routes = (
            Route.query.filter(Route.id.in_(payload["route_ids"]), Route.is_deleted.is_(False)).all()
            if payload["route_ids"]
            else []
        )
    if payload["status"] == CONTENT_STATUS_PUBLISHED:
        announcement.published_at = payload["published_at"] or announcement.published_at or utcnow()
    else:
        announcement.published_at = payload["published_at"]
    announcement.offline_at = payload["offline_at"]

    db.session.commit()
    write_audit_log(g.current_user.id, "announcement.update", "announcement", str(announcement.id), announcement.title)
    flash("公告更新成功", "success")
    if (request.form.get("submit_action") or "").strip() == "save_view":
        return redirect(url_for("web.announcement_detail", announcement_id=announcement.id))
    return redirect(url_for("admin.announcements_page"))


@bp.post("/announcements/<int:announcement_id>/status")
@login_required
def update_announcement_status(announcement_id: int):
    announcement = Announcement.query.filter_by(id=announcement_id).first()
    if not announcement:
        flash("公告不存在", "error")
        return redirect(url_for("admin.announcements_page"))

    status = (request.form.get("status") or "").strip()
    if status not in {CONTENT_STATUS_PUBLISHED, CONTENT_STATUS_OFFLINE}:
        flash("公告状态无效", "error")
        return redirect(url_for("admin.announcements_page"))

    announcement.status = status
    announcement.updated_by = g.current_user.id
    announcement.updated_at = utcnow()
    if status == CONTENT_STATUS_PUBLISHED and not announcement.published_at:
        announcement.published_at = utcnow()

    db.session.commit()
    write_audit_log(g.current_user.id, "announcement.status", "announcement", str(announcement.id), status)
    flash("公告已上线" if status == CONTENT_STATUS_PUBLISHED else "公告已下线", "success")
    return redirect(url_for("admin.announcements_page"))


@bp.post("/announcements/<int:announcement_id>/delete")
@login_required
def delete_announcement(announcement_id: int):
    announcement = Announcement.query.filter_by(id=announcement_id).first()
    if not announcement:
        flash("公告不存在", "error")
        return redirect(url_for("admin.announcements_page"))

    title = announcement.title
    db.session.delete(announcement)
    db.session.commit()
    write_audit_log(g.current_user.id, "announcement.delete", "announcement", str(announcement_id), title)
    flash("公告删除成功", "success")
    return redirect(url_for("admin.announcements_page"))


@bp.post("/users/create")
@login_required
def create_user():
    username = (request.form.get("username") or "").strip()
    password = (request.form.get("password") or "").strip()
    role = (request.form.get("role") or ROLE_CONTENT_ADMIN).strip()

    if not username or not password or role not in ROLES:
        flash("参数错误：请检查用户名/密码/角色", "error")
        return redirect(url_for("admin.users_page"))
    if len(password) < 6:
        flash("参数错误：密码长度至少 6 位", "error")
        return redirect(url_for("admin.users_page"))

    if User.query.filter_by(username=username).first():
        flash("参数错误：用户名已存在", "error")
        return redirect(url_for("admin.users_page"))

    user = User(
        username=username,
        password=generate_password_hash(password),
        role=role,
        is_active=True,
    )
    db.session.add(user)
    _apply_user_page_permissions(user, _page_permissions_from_form(role))
    db.session.commit()
    write_audit_log(g.current_user.id, "user.create", "user", str(user.id), username)
    flash("管理员创建成功", "success")
    return redirect(url_for("admin.users_page"))


@bp.post("/users/<int:user_id>/update")
@login_required
def update_user(user_id: int):
    user = User.query.filter_by(id=user_id).first()
    if not user:
        flash("管理员不存在", "error")
        return redirect(url_for("admin.users_page"))

    username = (request.form.get("username") or "").strip()
    role = (request.form.get("role") or ROLE_CONTENT_ADMIN).strip()
    is_active = (request.form.get("is_active") or "0").strip() == "1"
    password = (request.form.get("password") or "").strip()

    if not username or role not in ROLES:
        flash("参数错误：请检查用户名和角色", "error")
        return redirect(url_for("admin.user_edit_page", user_id=user_id))
    if User.query.filter(User.id != user_id, User.username == username).first():
        flash("参数错误：用户名已存在", "error")
        return redirect(url_for("admin.user_edit_page", user_id=user_id))
    if password and len(password) < 6:
        flash("参数错误：密码长度至少 6 位", "error")
        return redirect(url_for("admin.user_edit_page", user_id=user_id))
    if user.id == g.current_user.id and not is_active:
        flash("不能停用当前登录账号", "error")
        return redirect(url_for("admin.user_edit_page", user_id=user_id))

    user.username = username
    user.role = role
    user.is_active = is_active
    _apply_user_page_permissions(user, _page_permissions_from_form(role))
    if password:
        user.password = generate_password_hash(password)

    db.session.commit()
    write_audit_log(g.current_user.id, "user.update", "user", str(user.id), username)
    flash("管理员更新成功", "success")
    return redirect(url_for("admin.users_page"))


@bp.post("/users/<int:user_id>/delete")
@login_required
def delete_user(user_id: int):
    user = User.query.filter_by(id=user_id).first()
    if not user:
        flash("管理员不存在", "error")
        return redirect(url_for("admin.users_page"))
    if user.id == g.current_user.id:
        flash("不能删除当前登录账号", "error")
        return redirect(url_for("admin.user_edit_page", user_id=user_id))

    user.is_active = False
    db.session.commit()
    write_audit_log(g.current_user.id, "user.deactivate", "user", str(user.id), user.username)
    flash("管理员已停用", "success")
    return redirect(url_for("admin.users_page"))


def _cleanup_paths(paths: list[Path]) -> None:
    for path in paths:
        path.unlink(missing_ok=True)


@bp.post("/bulk-import")
@login_required
def bulk_import():
    csv_file = request.files.get("csv_file")
    if not csv_file or not (csv_file.filename or "").lower().endswith(".csv"):
        flash("参数错误：请上传 CSV 文件", "error")
        return redirect(url_for("admin.routes_page"))

    csv_text = csv_file.stream.read().decode("utf-8-sig")
    reader = csv.DictReader(StringIO(csv_text))
    required = {
        "route_name",
        "gpx_filename",
        "difficulty",
        "category",
        "description",
        "status",
        "suggested_duration_hours",
        "supply_points",
        "risk_warning",
    }
    if not required.issubset(set(reader.fieldnames or [])):
        flash("参数错误：CSV 缺少必要列", "error")
        return redirect(url_for("admin.routes_page"))
    rows = list(reader)

    extra_gpx_files = request.files.getlist("gpx_files")
    uploaded_name_map: dict[str, str] = {}
    uploaded_paths: dict[str, Path] = {}
    for gpx in extra_gpx_files:
        if not gpx or not gpx.filename:
            continue
        if not allowed_file(gpx.filename, {".gpx"}):
            continue
        if not file_size_ok(gpx, current_app.config.get("MAX_GPX_BYTES", 5 * 1024 * 1024)):
            continue

        original_name = Path(gpx.filename).name
        saved_name, saved_path = save_gpx_file(gpx)
        if not saved_name or not saved_path:
            continue
        uploaded_name_map[original_name] = saved_name
        uploaded_paths[saved_name] = saved_path

    upload_folder = Path(current_app.config["UPLOAD_FOLDER"])
    created = 0
    skipped = 0
    used_uploaded_names: set[str] = set()
    report_rows: list[dict] = []

    for index, row in enumerate(rows, start=1):
        route_name = (row.get("route_name") or "").strip()
        gpx_filename = (row.get("gpx_filename") or "").strip()
        gpx_filename = uploaded_name_map.get(gpx_filename, gpx_filename)
        difficulty = (row.get("difficulty") or "medium").strip()
        category = (row.get("category") or "hiking").strip()
        description = (row.get("description") or "").strip()
        status = (row.get("status") or STATUS_OFFLINE).strip()
        suggested_duration_hours = parse_distance(row.get("suggested_duration_hours") or "0")
        supply_points = (row.get("supply_points") or "").strip()
        risk_warning = (row.get("risk_warning") or "").strip()

        reason = ""
        if not route_name or not gpx_filename.lower().endswith(".gpx") or status not in ROUTE_STATUSES:
            reason = "invalid_fields"
        elif suggested_duration_hours is None:
            reason = "invalid_duration"
        elif Route.query.filter((Route.gpx_filename == gpx_filename) | (Route.route_name == route_name)).first():
            reason = "duplicated"
        elif not (upload_folder / gpx_filename).exists():
            reason = "gpx_not_found"

        if reason:
            skipped += 1
            report_rows.append(
                {
                    "row": index,
                    "route_name": route_name,
                    "gpx_filename": gpx_filename,
                    "status": "failed",
                    "reason": reason,
                }
            )
            continue

        gpx_path = upload_folder / gpx_filename
        try:
            computed_stats = _compute_route_stats(gpx_path)
        except Exception:
            skipped += 1
            report_rows.append(
                {
                    "row": index,
                    "route_name": route_name,
                    "gpx_filename": gpx_filename,
                    "status": "failed",
                    "reason": "gpx_parse_error",
                }
            )
            continue

        route = Route(
            route_name=route_name,
            gpx_filename=gpx_filename,
            distance_km=computed_stats["distance_km"],
            difficulty=difficulty,
            category=category,
            description=description,
            status=status,
            is_active=(status == STATUS_PUBLISHED),
            uploaded_at=utcnow(),
            suggested_duration_hours=suggested_duration_hours,
            supply_points=supply_points,
            risk_warning=risk_warning,
            ascent_m=computed_stats["ascent_m"],
            descent_m=computed_stats["descent_m"],
            min_ele_m=computed_stats["min_ele_m"],
            max_ele_m=computed_stats["max_ele_m"],
            created_by=g.current_user.id,
            updated_by=g.current_user.id,
        )
        db.session.add(route)
        db.session.flush()
        create_route_version(route, g.current_user.id, change_note="bulk_import")
        created += 1
        report_rows.append(
            {
                "row": index,
                "route_name": route_name,
                "gpx_filename": gpx_filename,
                "status": "success",
                "reason": "",
            }
        )
        if gpx_filename in uploaded_paths:
            used_uploaded_names.add(gpx_filename)

    try:
        db.session.commit()
        report = save_import_report(g.current_user.id, report_rows, created, skipped)
    except Exception:
        db.session.rollback()
        _cleanup_paths(list(uploaded_paths.values()))
        flash("批量导入失败：数据库写入异常", "error")
        return redirect(url_for("admin.routes_page"))

    orphan_paths = [path for name, path in uploaded_paths.items() if name not in used_uploaded_names]
    _cleanup_paths(orphan_paths)

    write_audit_log(g.current_user.id, "route.bulk_import", "route", None, f"created={created},skipped={skipped}")
    flash(
        f"批量导入完成：成功 {created}，跳过 {skipped}。报告：{url_for('admin.download_import_report', token=report.report_token)}",
        "success",
    )
    return redirect(url_for("admin.routes_page"))


@bp.get("/import-report/<string:token>")
@login_required
def download_import_report(token: str):
    report = ImportReport.query.filter_by(report_token=token).first()
    if not report:
        abort(404, description="报告不存在")

    report_dir = Path(current_app.instance_path) / "import_reports"
    report_path = report_dir / report.report_filename
    if not report_path.exists():
        abort(404, description="报告文件不存在")

    return send_from_directory(
        str(report_dir),
        report.report_filename,
        as_attachment=True,
        download_name=report.report_filename,
        mimetype="text/csv",
    )






