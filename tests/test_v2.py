from datetime import date, datetime, timedelta, timezone
from datetime import time
from io import BytesIO
import json
from pathlib import Path
import re

from alembic import command
from alembic.config import Config
import pytest
from sqlalchemy import UniqueConstraint, create_engine, inspect
from werkzeug.security import check_password_hash, generate_password_hash

from app import create_app
from app.membership_application_options import APPLICATION_STATUSES, BICYCLE_STATUS_VALUES, COMPETITION_INTEREST_VALUES, CURRENT_MEMBERSHIP_APPLICATION_FORM_VERSION, CYCLING_EXPERIENCE_VALUES
from app.models import AccessLog, Activity, ActivityRouteOption, Announcement, AuditLog, EventRegistration, MediaAsset, MEMBER_ACCOUNT_ACTIVE, MEMBER_ACCOUNT_DISABLED, MERCH_BATCH_ACTIVE, MERCH_BATCH_ENDED, MERCH_BATCH_UPCOMING, MERCH_ORDER_CANCELLED, MERCH_ORDER_PENDING, MembershipApplication, MemberProfile, MemberUser, MerchPreorderBatch, MerchPreorderRegistration, PAGE_ACCOUNTS, PAGE_ANALYTICS, PAGE_AUDIT_LOGS, PAGE_FEEDBACK, PAGE_MEMBERS, PAGE_ROUTES, PAGE_SECURITY, PERMISSION_ADMIN, PERMISSION_NONE, PERMISSION_READ, PERMISSION_WRITE, ROLE_OPS_ADMIN, ROLE_VIEWER, Route, RouteFeedback, RouteVersion, SiteFeedback, User, UserPagePermission, db
from app.security_monitor import is_watchlist_probe_path
from app.services import is_membership_application_enabled, set_membership_application_enabled


def _extract_csrf(html: str) -> str:
    match = re.search(r'name="csrf_token"\s+value="([^"]+)"', html)
    assert match, "csrf token not found"
    return match.group(1)


def _project_version() -> str:
    return Path("VERSION").read_text(encoding="utf-8-sig").strip()


def grant_page_permission(user: User, page_key: str, level: str) -> None:
    db.session.add(UserPagePermission(user=user, page_key=page_key, permission_level=level))


@pytest.fixture()
def app_and_client(tmp_path, monkeypatch):
    db_path = tmp_path / "test.db"
    upload_dir = tmp_path / "uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)

    monkeypatch.setenv("FLASK_ENV", "development")
    monkeypatch.setenv("SECRET_KEY", "test-secret")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{db_path.as_posix()}")
    monkeypatch.setenv("UPLOAD_FOLDER", str(upload_dir))
    monkeypatch.setenv("DEFAULT_ADMIN_USERNAME", "admin")
    monkeypatch.setenv("DEFAULT_ADMIN_PASSWORD", "admin123456789")
    monkeypatch.setenv("APP_DEPLOYED_AT", "2026-03-17T18:30:00+08:00")
    monkeypatch.setenv("ACCESS_LOG_ASYNC", "false")

    app = create_app()
    app.config.update(TESTING=True, SEED_DEMO_DATA=False)

    with app.test_client() as test_client:
        yield app, test_client


def login(client, username: str, password: str):
    login_page = client.get("/manage/login")
    token = _extract_csrf(login_page.get_data(as_text=True))
    return client.post(
        "/manage/login",
        data={"username": username, "password": password, "csrf_token": token},
        follow_redirects=True,
    )


def _clear_manage_session(client):
    with client.session_transaction() as sess:
        sess.pop("user_id", None)


def login_admin(client):
    return login(client, "admin", "admin123456789")


def register_member(client, student_id: str, nickname: str, password: str, follow_redirects: bool = True):
    register_page = client.get("/member/register")
    token = _extract_csrf(register_page.get_data(as_text=True))
    return client.post(
        "/member/register",
        data={
            "student_id": student_id,
            "nickname": nickname,
            "password": password,
            "password_confirm": password,
            "csrf_token": token,
        },
        follow_redirects=follow_redirects,
    )


def login_member(client, student_id: str, password: str, follow_redirects: bool = True):
    login_page = client.get("/member/login")
    token = _extract_csrf(login_page.get_data(as_text=True))
    return client.post(
        "/member/login",
        data={"student_id": student_id, "password": password, "csrf_token": token},
        follow_redirects=follow_redirects,
    )


def get_manage_csrf(client) -> str:
    page = client.get("/manage")
    assert page.status_code == 200
    return _extract_csrf(page.get_data(as_text=True))


def create_route(
    client,
    csrf_token: str,
    name: str = "Test Route",
    follow_redirects: bool = True,
    distance_km: str = "3.2",
    gpx_content: bytes | None = None,
):
    if gpx_content is None:
        gpx_content = b"<?xml version='1.0'?><gpx version='1.1'></gpx>"
    return client.post(
        "/manage/routes/create",
        data={
            "csrf_token": csrf_token,
            "route_name": name,
            "distance_km": distance_km,
            "difficulty": "easy",
            "category": "hiking",
            "description": "for test",
            "status": "published",
            "suggested_duration_hours": "1.5",
            "supply_points": "shop",
            "risk_warning": "slippery",
            "gpx_file": (BytesIO(gpx_content), f"{name}.gpx"),
        },
        content_type="multipart/form-data",
        follow_redirects=follow_redirects,
    )


def test_api_v1_list_available(app_and_client):
    _app, client = app_and_client
    resp = client.get("/api/v1/routes")
    assert resp.status_code == 200
    payload = resp.get_json()
    assert "items" in payload


def test_csrf_required_for_login(app_and_client):
    _app, client = app_and_client
    resp = client.post(
        "/manage/login",
        data={"username": "admin", "password": "admin123456789"},
        follow_redirects=False,
    )
    assert resp.status_code == 400


def test_member_register_creates_account_and_logs_in(app_and_client):
    app, client = app_and_client
    resp = register_member(client, "12345678", "Rider One", "memberpass123")
    assert resp.status_code == 200
    assert "Rider One" in resp.get_data(as_text=True)
    account_page = client.get("/member/account")
    assert "Rider One" in account_page.get_data(as_text=True)

    with app.app_context():
        member = MemberUser.query.filter_by(student_id="12345678").first()
        assert member is not None
        assert member.nickname == "Rider One"
        assert member.account_status == MEMBER_ACCOUNT_ACTIVE
        assert check_password_hash(member.password_hash, "memberpass123")
        assert member.last_login_at is None


def test_member_register_rejects_duplicate_student_id(app_and_client):
    app, client = app_and_client
    assert register_member(client, "sid001", "First Rider", "memberpass123").status_code == 200
    with client.session_transaction() as sess:
        sess.pop("member_user_id", None)
    resp = register_member(client, "SID001", "Second Rider", "memberpass123")
    assert resp.status_code == 409
    assert "该学号已注册账号" in resp.get_data(as_text=True)

    with app.app_context():
        assert MemberUser.query.filter_by(student_id="SID001").count() == 1


def test_member_login_updates_last_login_and_logout_clears_session(app_and_client):
    app, client = app_and_client
    assert register_member(client, "20260001", "Login Rider", "memberpass123").status_code == 200
    page = client.get("/member/account")
    logout_token = _extract_csrf(page.get_data(as_text=True))
    client.post("/member/logout", data={"csrf_token": logout_token}, follow_redirects=True)

    resp = login_member(client, "20260001", "memberpass123")
    assert resp.status_code == 200
    assert "Login Rider" in resp.get_data(as_text=True)
    account_page = client.get("/member/account")
    assert "Login Rider" in account_page.get_data(as_text=True)

    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20260001").first()
        assert member.last_login_at is not None

    page = client.get("/member/account")
    logout_token = _extract_csrf(page.get_data(as_text=True))
    resp = client.post("/member/logout", data={"csrf_token": logout_token}, follow_redirects=True)
    assert resp.status_code == 200
    assert "Login Rider" not in resp.get_data(as_text=True)


def test_member_password_page_requires_login(app_and_client):
    _app, client = app_and_client
    resp = client.get("/member/password", follow_redirects=False)
    assert resp.status_code == 302
    assert "/member/login" in (resp.headers.get("Location") or "")
    assert "next=/member/password" in (resp.headers.get("Location") or "")


def test_member_account_page_requires_login(app_and_client):
    _app, client = app_and_client
    resp = client.get("/member/account", follow_redirects=False)
    assert resp.status_code == 302
    assert "/member/login" in (resp.headers.get("Location") or "")
    assert "next=/member/account" in (resp.headers.get("Location") or "")


def test_member_profile_page_requires_login(app_and_client):
    _app, client = app_and_client
    resp = client.get("/member/profile", follow_redirects=False)
    assert resp.status_code == 302
    assert "/member/login" in (resp.headers.get("Location") or "")
    assert "next=/member/profile" in (resp.headers.get("Location") or "")


def test_member_account_page_shows_account_summary_and_password_entry(app_and_client):
    _app, client = app_and_client
    assert register_member(client, "20260050", "Account Rider", "memberpass123").status_code == 200

    page = client.get("/member/account")
    html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert "我的账号" in html
    assert "20260050" in html
    assert "Account Rider" in html
    assert "修改密码" in html
    assert "/member/password" in html
    assert "账号状态" not in html
    assert "注册时间" not in html
    assert "最近登录" not in html
    assert "社员资料" in html
    assert "/member/profile" in html
    assert "租车与装备记录" in html
    assert "退出登录" in html

    homepage = client.get("/")
    homepage_html = homepage.get_data(as_text=True)
    assert "Account Rider" in homepage_html
    assert "<span>账号</span>" not in homepage_html
    assert "改密" not in homepage_html
    assert "退出登录" not in homepage_html
    assert "/member/account" in homepage_html


def test_member_register_auto_links_matching_profile(app_and_client):
    app, client = app_and_client
    with app.app_context():
        profile = MemberProfile(
            student_id="20260100",
            full_name="Linked Person",
            gender="女",
            entry_year=2026,
            school="理工学院",
            college="逸夫书院",
            phone="13800000000",
            last_confirmed_at=date(2026, 7, 11),
        )
        db.session.add(profile)
        db.session.commit()
        profile_id = profile.id

    assert register_member(client, "20260100", "Linked Rider", "memberpass123").status_code == 200

    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20260100").first()
        profile = db.session.get(MemberProfile, profile_id)
        assert member is not None
        assert profile is not None
        assert profile.member_user_id == member.id

    page = client.get("/member/profile")
    html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert "Linked Person" in html
    assert "理工学院" in html
    assert "13800000000" in html
    assert "2026-07-11" in html


def test_member_account_auto_links_profile_imported_after_registration(app_and_client):
    app, client = app_and_client
    assert register_member(client, "20260101", "Late Profile Rider", "memberpass123").status_code == 200

    with app.app_context():
        profile = MemberProfile(
            student_id="20260101",
            full_name="Late Profile Person",
            school="经管学院",
            last_confirmed_at=date(2026, 7, 11),
        )
        db.session.add(profile)
        db.session.commit()
        profile_id = profile.id
        assert profile.member_user_id is None

    page = client.get("/member/account")
    html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert "已匹配到 Late Profile Person 的社员档案" in html

    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20260101").first()
        profile = db.session.get(MemberProfile, profile_id)
        assert member is not None
        assert profile.member_user_id == member.id


def test_member_profile_page_handles_missing_profile(app_and_client):
    _app, client = app_and_client
    assert register_member(client, "20260102", "No Profile Rider", "memberpass123").status_code == 200

    page = client.get("/member/profile")
    html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert "暂未匹配到社员档案" in html
    assert "20260102" in html


def test_member_can_update_own_nickname(app_and_client):
    app, client = app_and_client
    assert register_member(client, "20260052", "Old Nick", "memberpass123").status_code == 200

    page = client.get("/member/account")
    token = _extract_csrf(page.get_data(as_text=True))
    blank = client.post(
        "/member/account/nickname",
        data={"csrf_token": token, "nickname": "   "},
        follow_redirects=True,
    )
    assert blank.status_code == 400
    assert "请填写昵称" in blank.get_data(as_text=True)

    updated = client.post(
        "/member/account/nickname",
        data={"csrf_token": token, "nickname": "New Nick"},
        follow_redirects=True,
    )
    updated_html = updated.get_data(as_text=True)
    assert updated.status_code == 200
    assert "昵称已更新" in updated_html
    assert "New Nick" in updated_html

    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20260052").first()
        assert member is not None
        assert member.nickname == "New Nick"

    homepage = client.get("/")
    assert "New Nick" in homepage.get_data(as_text=True)


def test_member_can_change_own_password(app_and_client):
    app, client = app_and_client
    assert register_member(client, "20260051", "Password Rider", "oldpass123").status_code == 200

    page = client.get("/member/password")
    assert page.status_code == 200
    token = _extract_csrf(page.get_data(as_text=True))
    wrong_current = client.post(
        "/member/password",
        data={
            "csrf_token": token,
            "current_password": "wrongpass123",
            "new_password": "newpass123",
            "password_confirm": "newpass123",
        },
        follow_redirects=True,
    )
    assert wrong_current.status_code == 400
    assert "当前密码不正确" in wrong_current.get_data(as_text=True)

    updated = client.post(
        "/member/password",
        data={
            "csrf_token": token,
            "current_password": "oldpass123",
            "new_password": "newpass123",
            "password_confirm": "newpass123",
        },
        follow_redirects=True,
    )
    assert updated.status_code == 200
    assert "密码已更新" in updated.get_data(as_text=True)

    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20260051").first()
        assert member is not None
        assert check_password_hash(member.password_hash, "newpass123")
        assert not check_password_hash(member.password_hash, "oldpass123")

    page = client.get("/member/account")
    logout_token = _extract_csrf(page.get_data(as_text=True))
    client.post("/member/logout", data={"csrf_token": logout_token}, follow_redirects=True)
    assert login_member(client, "20260051", "oldpass123").status_code == 401
    assert login_member(client, "20260051", "newpass123").status_code == 200


def test_member_login_rejects_disabled_account(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            MemberUser(
                student_id="DISABLED001",
                nickname="Disabled Rider",
                password_hash=generate_password_hash("memberpass123"),
                account_status=MEMBER_ACCOUNT_DISABLED,
            )
        )
        db.session.commit()

    resp = login_member(client, "DISABLED001", "memberpass123")
    assert resp.status_code == 401
    assert "学号或密码不正确" in resp.get_data(as_text=True)


def test_manage_member_users_page_lists_member_accounts(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add_all(
            [
                MemberUser(
                    student_id="20260011",
                    nickname="Visible Rider",
                    password_hash=generate_password_hash("memberpass123"),
                    account_status=MEMBER_ACCOUNT_ACTIVE,
                ),
                MemberUser(
                    student_id="20260012",
                    nickname="Disabled Rider",
                    password_hash=generate_password_hash("memberpass123"),
                    account_status=MEMBER_ACCOUNT_DISABLED,
                ),
            ]
        )
        db.session.commit()

    assert login_admin(client).status_code == 200
    resp = client.get("/manage/members")
    assert resp.status_code == 200
    html = resp.get_data(as_text=True)
    assert "20260011" in html
    assert "Visible Rider" in html
    assert "20260012" in html
    assert "Disabled Rider" in html
    assert "重置密码" in html
    assert "删除账户" in html

    filtered = client.get("/manage/members?q=Visible")
    assert filtered.status_code == 200
    filtered_html = filtered.get_data(as_text=True)
    assert "Visible Rider" in filtered_html
    assert "Disabled Rider" not in filtered_html


def test_manage_member_profiles_page_lists_and_filters_profiles(app_and_client):
    app, client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20260201",
            nickname="Bound Rider",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        db.session.add_all(
            [
                member,
                MemberProfile(
                    member_user=member,
                    student_id="20260201",
                    full_name="Bound Person",
                    gender="男",
                    entry_year=2026,
                    school="理工学院",
                    college="逸夫书院",
                    phone="13800000001",
                    last_confirmed_at=date(2026, 7, 11),
                ),
                MemberProfile(
                    student_id="20260202",
                    full_name="Unbound Person",
                    school="经管学院",
                    phone="13800000002",
                    last_confirmed_at=date(2026, 7, 11),
                ),
            ]
        )
        db.session.commit()

    assert login_admin(client).status_code == 200
    resp = client.get("/manage/member-profiles")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "Bound Person" in html
    assert "Unbound Person" in html
    assert "理工学院" in html
    assert "13800000001" in html
    assert "已绑定" in html
    assert "未绑定" in html
    assert "Bound Rider" in html

    filtered = client.get("/manage/member-profiles?q=经管")
    filtered_html = filtered.get_data(as_text=True)
    assert filtered.status_code == 200
    assert "Unbound Person" in filtered_html
    assert "Bound Person" not in filtered_html

    linked = client.get("/manage/member-profiles?link_status=linked")
    linked_html = linked.get_data(as_text=True)
    assert linked.status_code == 200
    assert "Bound Person" in linked_html
    assert "Unbound Person" not in linked_html

    unlinked = client.get("/manage/member-profiles?link_status=unlinked")
    unlinked_html = unlinked.get_data(as_text=True)
    assert unlinked.status_code == 200
    assert "Unbound Person" in unlinked_html
    assert "Bound Person" not in unlinked_html


def test_manage_member_profiles_page_allows_member_read_permission(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            MemberProfile(
                student_id="20260203",
                full_name="Reader Visible Person",
                school="人文社科学院",
                last_confirmed_at=date(2026, 7, 11),
            )
        )
        reader = User(
            username="member_profile_reader",
            password=generate_password_hash("reader123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(reader)
        db.session.flush()
        grant_page_permission(reader, PAGE_MEMBERS, PERMISSION_READ)
        db.session.commit()

    assert login(client, "member_profile_reader", "reader123456789").status_code == 200
    page = client.get("/manage/member-profiles")
    html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert "Reader Visible Person" in html
    assert "社员档案" in html
    assert "编辑</a>" not in html


def test_manage_member_profile_edit_updates_profile_and_writes_audit_log(app_and_client):
    app, client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20260204",
            nickname="Editable Bound Rider",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        profile = MemberProfile(
            student_id="20260204",
            full_name="Editable Person",
            school="SSE",
            phone="13800000004",
            last_confirmed_at=date(2026, 7, 11),
        )
        db.session.add_all([member, profile])
        db.session.commit()
        member_id = member.id
        profile_id = profile.id

    assert login_admin(client).status_code == 200
    edit_page = client.get(f"/manage/member-profiles/{profile_id}/edit")
    assert edit_page.status_code == 200
    token = _extract_csrf(edit_page.get_data(as_text=True))
    updated = client.post(
        f"/manage/member-profiles/{profile_id}/edit",
        data={
            "csrf_token": token,
            "student_id": "20260204",
            "full_name": "Editable Person Updated",
            "entry_year": "2026",
            "gender": "女",
            "school": "SDS",
            "college": "muse",
            "phone": "13900000004",
            "last_confirmed_at": "2026-07-11",
            "member_user_id": str(member_id),
        },
        follow_redirects=True,
    )
    html = updated.get_data(as_text=True)
    assert updated.status_code == 200
    assert "社员档案已更新" in html

    with app.app_context():
        profile = db.session.get(MemberProfile, profile_id)
        assert profile.full_name == "Editable Person Updated"
        assert profile.school == "SDS"
        assert profile.college == "muse"
        assert profile.member_user_id == member_id
        log = AuditLog.query.filter_by(action="member_profile.admin_update", target_type="member_profile", target_id=str(profile_id)).first()
        assert log is not None
        assert '"source": "admin_update"' in log.detail
        assert '"school": {"before": "SSE", "after": "SDS"}' in log.detail


def _member_profile_import_workbook(rows: list[list]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["学号", "真实姓名", "入学年份", "性别", "学院", "书院", "手机号", "最近确认日期"])
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_manage_member_profile_import_template_downloads(app_and_client):
    _app, client = app_and_client
    assert login_admin(client).status_code == 200
    resp = client.get("/manage/member-profiles/import-template.xlsx")
    assert resp.status_code == 200
    assert resp.headers["Content-Type"].startswith("application/vnd.openxmlformats-officedocument")
    assert len(resp.data) > 100


def test_manage_member_profile_import_requires_write_permission(app_and_client):
    app, client = app_and_client
    with app.app_context():
        reader = User(
            username="member_import_reader",
            password=generate_password_hash("reader123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(reader)
        db.session.flush()
        grant_page_permission(reader, PAGE_MEMBERS, PERMISSION_READ)
        db.session.commit()

    assert login(client, "member_import_reader", "reader123456789").status_code == 200
    page = client.get("/manage/member-profiles")
    assert page.status_code == 200
    resp = client.get("/manage/member-profiles/import-template.xlsx")
    assert resp.status_code in {302, 403}
    assert not resp.headers.get("Content-Type", "").startswith("application/vnd.openxmlformats-officedocument")


def test_manage_member_profile_import_write_permission_skips_duplicates(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            MemberProfile(
                student_id="20260301",
                full_name="Existing Person",
                school="SSE",
                phone="13800000001",
                last_confirmed_at=date(2026, 7, 11),
            )
        )
        writer = User(
            username="member_import_writer",
            password=generate_password_hash("writer123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(writer)
        db.session.flush()
        grant_page_permission(writer, PAGE_MEMBERS, PERMISSION_WRITE)
        db.session.commit()

    assert login(client, "member_import_writer", "writer123456789").status_code == 200
    page = client.get("/manage/member-profiles")
    token = _extract_csrf(page.get_data(as_text=True))
    workbook_bytes = _member_profile_import_workbook(
        [
            ["20260301", "Existing Person Updated", 2026, "女", "数据科学学院", "思廷书院", "13900000001", "2026-07-12"],
            ["20260302", "Created Person", "2022级及以前", "", "理工学院", "", "13800000002", "2026-07-11"],
        ]
    )
    preview = client.post(
        "/manage/member-profiles/import-preview",
        data={"csrf_token": token, "excel_file": (BytesIO(workbook_bytes), "members.xlsx")},
        content_type="multipart/form-data",
    )
    preview_html = preview.get_data(as_text=True)
    assert preview.status_code == 200
    assert "Existing Person Updated" in preview_html
    assert "理工学院 | SSE" in preview_html
    assert "数据科学学院 | SDS" in preview_html
    assert 'value="overwrite"' not in preview_html
    preview_token = re.search(r'name="preview_token" value="([^"]+)"', preview_html).group(1)
    confirm_token = _extract_csrf(preview_html)
    confirmed = client.post(
        "/manage/member-profiles/import-confirm",
        data={"csrf_token": confirm_token, "preview_token": preview_token, "import_mode": "skip"},
        follow_redirects=True,
    )
    assert confirmed.status_code == 200
    assert "新增 1 条，覆盖 0 条，跳过 1 条" in confirmed.get_data(as_text=True)

    with app.app_context():
        existing = MemberProfile.query.filter_by(student_id="20260301").first()
        created = MemberProfile.query.filter_by(student_id="20260302").first()
        assert existing.full_name == "Existing Person"
        assert existing.school == "SSE"
        assert created is not None
        assert created.full_name == "Created Person"
        assert created.entry_year == 2022
        assert created.school == "SSE"
        log = AuditLog.query.filter_by(action="member_profile.import_excel", target_id=str(created.id)).first()
        assert log is not None
        assert '"source": "excel_import"' in log.detail


def test_manage_member_profile_import_admin_can_overwrite_duplicates(app_and_client):
    app, client = app_and_client
    with app.app_context():
        profile = MemberProfile(
            student_id="20260303",
            full_name="Overwrite Person",
            school="SSE",
            phone="13800000003",
            last_confirmed_at=date(2026, 7, 11),
        )
        db.session.add(profile)
        db.session.commit()
        profile_id = profile.id

    assert login_admin(client).status_code == 200
    page = client.get("/manage/member-profiles")
    token = _extract_csrf(page.get_data(as_text=True))
    workbook_bytes = _member_profile_import_workbook(
        [["20260303", "Overwrite Person Updated", 2026, "女", "数据科学学院", "思廷书院", "13900000003", "2026-07-12"]]
    )
    preview = client.post(
        "/manage/member-profiles/import-preview",
        data={"csrf_token": token, "excel_file": (BytesIO(workbook_bytes), "members.xlsx")},
        content_type="multipart/form-data",
    )
    preview_html = preview.get_data(as_text=True)
    assert preview.status_code == 200
    assert 'value="overwrite"' in preview_html
    preview_token = re.search(r'name="preview_token" value="([^"]+)"', preview_html).group(1)
    confirm_token = _extract_csrf(preview_html)
    confirmed = client.post(
        "/manage/member-profiles/import-confirm",
        data={"csrf_token": confirm_token, "preview_token": preview_token, "import_mode": "overwrite"},
        follow_redirects=True,
    )
    assert confirmed.status_code == 200
    assert "新增 0 条，覆盖 1 条，跳过 0 条" in confirmed.get_data(as_text=True)

    with app.app_context():
        profile = db.session.get(MemberProfile, profile_id)
        assert profile.full_name == "Overwrite Person Updated"
        assert profile.school == "SDS"
        assert profile.college == "muse"
        assert profile.phone == "13900000003"
        log = AuditLog.query.filter_by(action="member_profile.import_excel", target_type="member_profile", target_id=str(profile_id)).first()
        assert log is not None
        assert '"mode": "overwrite"' in log.detail
        assert '"school": {"before": "SSE", "after": "SDS"}' in log.detail


def test_member_can_update_own_profile_fields_and_audit_log(app_and_client):
    app, client = app_and_client
    with app.app_context():
        profile = MemberProfile(
            student_id="20260205",
            full_name="Self Editable Person",
            gender="男",
            entry_year=2026,
            school="SSE",
            college="shaw",
            phone="13800000005",
            last_confirmed_at=date(2026, 7, 11),
        )
        db.session.add(profile)
        db.session.commit()
        profile_id = profile.id

    assert register_member(client, "20260205", "Self Edit Rider", "memberpass123").status_code == 200
    edit_page = client.get("/member/profile/edit")
    assert edit_page.status_code == 200
    token = _extract_csrf(edit_page.get_data(as_text=True))
    updated = client.post(
        "/member/profile/edit",
        data={
            "csrf_token": token,
            "gender": "",
            "entry_year": "2022级及以前",
            "school": "SDS",
            "college": "muse",
            "phone": "13900000005",
        },
        follow_redirects=True,
    )
    html = updated.get_data(as_text=True)
    assert updated.status_code == 200
    assert "社员资料已保存并确认" in html
    assert "数据科学学院 | SDS" in html

    with app.app_context():
        profile = db.session.get(MemberProfile, profile_id)
        member = MemberUser.query.filter_by(student_id="20260205").first()
        assert profile.gender is None
        assert profile.entry_year == 2022
        assert profile.school == "SDS"
        assert profile.college == "muse"
        assert profile.phone == "13900000005"
        assert profile.last_confirmed_at is not None
        log = AuditLog.query.filter_by(action="member_profile.self_update", target_type="member_profile", target_id=str(profile_id)).first()
        assert log is not None
        assert f'"actor_member_user_id": {member.id}' in log.detail
        assert '"source": "self_update"' in log.detail


def test_manage_member_write_permission_can_update_but_not_delete(app_and_client):
    app, client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20260021",
            nickname="Writable Rider",
            password_hash=generate_password_hash("old-member-pass"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        writer = User(
            username="member_writer",
            password=generate_password_hash("writer123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add_all([member, writer])
        db.session.flush()
        grant_page_permission(writer, PAGE_MEMBERS, PERMISSION_WRITE)
        db.session.commit()
        member_id = member.id

    assert login_member(client, "20260021", "old-member-pass").status_code == 200
    page = client.get("/member/account")
    logout_token = _extract_csrf(page.get_data(as_text=True))
    client.post("/member/logout", data={"csrf_token": logout_token}, follow_redirects=True)

    assert login(client, "member_writer", "writer123456789").status_code == 200
    page = client.get("/manage/members")
    assert page.status_code == 200
    html = page.get_data(as_text=True)
    assert "重置密码" in html
    assert "删除账户" not in html
    token = _extract_csrf(html)

    disabled = client.post(
        f"/manage/members/{member_id}/status",
        data={"csrf_token": token, "account_status": MEMBER_ACCOUNT_DISABLED},
        follow_redirects=True,
    )
    assert disabled.status_code == 200
    with app.app_context():
        member = db.session.get(MemberUser, member_id)
        assert member.account_status == MEMBER_ACCOUNT_DISABLED

    reset = client.post(
        f"/manage/members/{member_id}/reset-password",
        data={"csrf_token": token},
        follow_redirects=True,
    )
    assert reset.status_code == 200
    reset_html = reset.get_data(as_text=True)
    assert "临时密码" in reset_html
    with app.app_context():
        member = db.session.get(MemberUser, member_id)
        assert not check_password_hash(member.password_hash, "old-member-pass")

    deleted = client.post(
        f"/manage/members/{member_id}/delete",
        data={"csrf_token": token},
        follow_redirects=False,
    )
    assert deleted.status_code == 302
    assert "/manage/login" in (deleted.headers.get("Location") or "")
    with app.app_context():
        assert db.session.get(MemberUser, member_id) is not None


def test_manage_member_read_permission_is_read_only(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            MemberUser(
                student_id="20260031",
                nickname="Read Only Rider",
                password_hash=generate_password_hash("memberpass123"),
                account_status=MEMBER_ACCOUNT_ACTIVE,
            )
        )
        reader = User(
            username="member_reader",
            password=generate_password_hash("reader123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(reader)
        db.session.flush()
        grant_page_permission(reader, PAGE_MEMBERS, PERMISSION_READ)
        db.session.commit()

    assert login(client, "member_reader", "reader123456789").status_code == 200
    page = client.get("/manage/members")
    assert page.status_code == 200
    html = page.get_data(as_text=True)
    assert "Read Only Rider" in html
    assert "重置密码" not in html
    assert "禁用账户" not in html
    assert "删除账户" not in html


def test_manage_member_admin_permission_can_delete_account(app_and_client):
    app, client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20260041",
            nickname="Delete Rider",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        db.session.add(member)
        db.session.commit()
        member_id = member.id

    assert login_admin(client).status_code == 200
    page = client.get("/manage/members")
    token = _extract_csrf(page.get_data(as_text=True))
    deleted = client.post(
        f"/manage/members/{member_id}/delete",
        data={"csrf_token": token},
        follow_redirects=True,
    )
    assert deleted.status_code == 200
    assert "社员账号已删除" in deleted.get_data(as_text=True)
    with app.app_context():
        assert db.session.get(MemberUser, member_id) is None


def test_member_profile_links_to_member_user_and_serializes(app_and_client):
    app, _client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20261001",
            nickname="Profile Rider",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        profile = MemberProfile(
            member_user=member,
            student_id="20261001",
            full_name="Profile Person",
            entry_year=2026,
            school="理工学院",
            college="逸夫书院",
            phone="13800000000",
            last_confirmed_at=date(2026, 7, 11),
        )
        db.session.add(profile)
        db.session.commit()

        saved = MemberProfile.query.filter_by(student_id="20261001").first()
        assert saved is not None
        assert saved.member_user == member
        assert member.profile == saved
        assert saved.as_dict()["last_confirmed_at"] == "2026-07-11"
        assert saved.as_dict()["full_name"] == "Profile Person"


def test_member_profile_survives_member_user_delete(app_and_client):
    app, _client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20261002",
            nickname="Detached Rider",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        profile = MemberProfile(
            member_user=member,
            student_id="20261002",
            full_name="Detached Person",
            last_confirmed_at=date(2026, 7, 11),
        )
        db.session.add(profile)
        db.session.commit()
        member_id = member.id
        profile_id = profile.id

        db.session.delete(member)
        db.session.commit()

        saved = db.session.get(MemberProfile, profile_id)
        assert db.session.get(MemberUser, member_id) is None
        assert saved is not None
        assert saved.member_user_id is None
        assert saved.student_id == "20261002"


def _membership_application(**overrides) -> MembershipApplication:
    values = {
        "student_id": "20261101",
        "full_name": "Application Person",
        "gender": "女",
        "entry_year": 2026,
        "school": "SSE",
        "college": "shaw",
        "phone": "13800000001",
        "competition_interest": "unsure",
        "cycling_experience": "casual",
        "bicycle_status": "road_bike",
    }
    values.update(overrides)
    return MembershipApplication(**values)


def _join_form_data(**overrides) -> dict[str, str]:
    values = {
        "student_id": "20261201",
        "full_name": "Join Person",
        "gender": "女",
        "entry_year": "2026",
        "school": "SSE",
        "college": "shaw",
        "phone": "+86 13800000001",
        "competition_interest": "unsure",
        "cycling_experience": "casual",
        "bicycle_status": "road_bike",
        "other_bicycle_description": "",
        "additional_note": "想了解周末骑行安排。",
        "confirm_info": "1",
    }
    values.update(overrides)
    return values


def _post_join(client, follow_redirects: bool = False, **overrides):
    page = client.get("/join")
    token = _extract_csrf(page.get_data(as_text=True))
    data = _join_form_data(**overrides)
    data["csrf_token"] = token
    return client.post("/join", data=data, follow_redirects=follow_redirects)


def _register_member_with_next(client, student_id: str, nickname: str, password: str, next_url: str = "/account/membership-applications"):
    register_page = client.get(f"/member/register?next={next_url}")
    token = _extract_csrf(register_page.get_data(as_text=True))
    return client.post(
        "/member/register",
        data={
            "student_id": student_id,
            "nickname": nickname,
            "password": password,
            "password_confirm": password,
            "next": next_url,
            "csrf_token": token,
        },
        follow_redirects=True,
    )


def _login_member_with_next(client, student_id: str, password: str, next_url: str = "/account/membership-applications"):
    login_page = client.get(f"/member/login?next={next_url}")
    token = _extract_csrf(login_page.get_data(as_text=True))
    return client.post(
        "/member/login",
        data={"student_id": student_id, "password": password, "next": next_url, "csrf_token": token},
        follow_redirects=True,
    )


def test_membership_application_can_be_created_without_member_user(app_and_client):
    app, _client = app_and_client
    with app.app_context():
        application = _membership_application(member_user_id=None)
        db.session.add(application)
        db.session.commit()

        saved = db.session.get(MembershipApplication, application.id)
        assert saved is not None
        assert saved.member_user_id is None
        assert saved.status == "pending"
        assert saved.form_version == CURRENT_MEMBERSHIP_APPLICATION_FORM_VERSION
        assert saved.submitted_at is not None
        assert saved.created_at is not None
        assert saved.updated_at is not None
        payload = saved.as_dict()
        assert payload["submitted_at"] == saved.submitted_at.isoformat()
        assert payload["created_at"] == saved.created_at.isoformat()
        assert payload["updated_at"] == saved.updated_at.isoformat()


def test_membership_application_relationships_and_non_unique_history(app_and_client):
    app, _client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20261102",
            nickname="Application Rider",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        reviewer = User(
            username="application_reviewer",
            password=generate_password_hash("reviewer123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        profile = MemberProfile(student_id="20261102", full_name="Approved Person")
        db.session.add_all([member, reviewer, profile])
        db.session.flush()

        first = _membership_application(
            member_user=member,
            student_id="20261102",
            full_name="Application Rider",
            reviewer=reviewer,
            approved_profile=profile,
            status="rejected",
        )
        second = _membership_application(
            member_user=member,
            student_id="20261102",
            full_name="Application Rider Again",
            status="pending",
            competition_interest="yes",
            cycling_experience="long_distance",
            bicycle_status="mountain_bike",
        )
        db.session.add_all([first, second])
        db.session.commit()

        saved = MembershipApplication.query.filter_by(student_id="20261102").order_by(MembershipApplication.id.asc()).all()
        assert len(saved) == 2
        assert saved[0].member_user == member
        assert saved[0].reviewer == reviewer
        assert saved[0].approved_profile == profile
        assert saved[1].member_user_id == member.id


def test_membership_application_options_and_foreign_keys_match_contract(app_and_client):
    app, _client = app_and_client
    assert APPLICATION_STATUSES == ("pending", "approved", "rejected")
    assert COMPETITION_INTEREST_VALUES == ("yes", "no", "unsure")
    assert CYCLING_EXPERIENCE_VALUES == ("beginner", "casual", "long_distance", "competition")
    assert BICYCLE_STATUS_VALUES == (
        "no_bicycle",
        "mountain_bike",
        "road_bike",
        "folding_commuter",
        "other_bicycle",
        "off_campus",
    )

    with app.app_context():
        table = MembershipApplication.__table__
        foreign_keys = {
            column.name: (foreign_key.column.table.name, foreign_key.column.name, foreign_key.ondelete)
            for column in table.columns
            for foreign_key in column.foreign_keys
        }
        assert foreign_keys["member_user_id"] == ("member_users", "id", "SET NULL")
        assert foreign_keys["reviewed_by"] == ("users", "id", "SET NULL")
        assert foreign_keys["approved_profile_id"] == ("member_profiles", "id", "SET NULL")
        unique_constraints = [constraint for constraint in table.constraints if isinstance(constraint, UniqueConstraint)]
        assert not any(constraint.columns.keys() == ["student_id"] for constraint in unique_constraints)
        assert not any(constraint.columns.keys() == ["member_user_id"] for constraint in unique_constraints)


def test_join_page_available_for_anonymous_user(app_and_client):
    _app, client = app_and_client
    resp = client.get("/join")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "加入 2Tired 骑行社" in html
    assert 'name="student_id"' in html
    assert "我已阅读上述说明" in html


def test_join_page_for_logged_in_member_uses_readonly_account_student_id(app_and_client):
    _app, client = app_and_client
    assert register_member(client, "20261202", "Join Rider", "memberpass123").status_code == 200
    resp = client.get("/join")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "20261202" in html
    assert "已登录账号将使用当前账号学号" in html


def test_join_page_and_submit_are_blocked_when_application_switch_is_off(app_and_client):
    app, client = app_and_client
    with app.app_context():
        set_membership_application_enabled(False)
        db.session.commit()

    page = client.get("/join")
    page_html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert "入社申请暂时关闭，请稍后关注网站通知。" in page_html
    assert 'name="student_id"' not in page_html

    token = _extract_csrf(client.get("/member/register").get_data(as_text=True))
    closed_post = client.post(
        "/join",
        data={**_join_form_data(student_id="20269001"), "csrf_token": token},
        follow_redirects=False,
    )
    assert closed_post.status_code == 302
    assert "/join" in (closed_post.headers.get("Location") or "")

    with app.app_context():
        assert MembershipApplication.query.filter_by(student_id="20269001").count() == 0


def test_membership_application_switch_requires_write_permission_for_admin_update(app_and_client):
    app, client = app_and_client
    with app.app_context():
        _create_membership_reader("membership_setting_reader")
        _create_membership_admin("membership_setting_admin")
        db.session.commit()

    assert login(client, "membership_setting_reader", "reader123456789").status_code == 200
    reader_token = get_manage_csrf(client)
    reader_close = client.post(
        "/manage/membership-applications/settings",
        data={"csrf_token": reader_token, "application_open": "0"},
        follow_redirects=False,
    )
    assert reader_close.status_code in {302, 403}

    with app.app_context():
        assert MembershipApplication.query.filter_by(status="pending").count() == 0
        assert is_membership_application_enabled() is True

    _clear_manage_session(client)
    assert login(client, "membership_setting_admin", "admin123456789").status_code == 200
    admin_token = get_manage_csrf(client)
    admin_close = client.post(
        "/manage/membership-applications/settings",
        data={"csrf_token": admin_token, "application_open": "0"},
        follow_redirects=True,
    )
    assert admin_close.status_code == 200

    with app.app_context():
        assert is_membership_application_enabled() is False
        close_log = AuditLog.query.filter_by(action="membership_application.close").order_by(AuditLog.id.desc()).first()
        assert close_log is not None


def test_membership_application_switch_reflects_frontend_pages(app_and_client):
    app, client = app_and_client
    with app.app_context():
        set_membership_application_enabled(True)
        db.session.commit()

    home_open = client.get("/")
    about_open = client.get("/about")
    home_open_html = home_open.get_data(as_text=True)
    about_open_html = about_open.get_data(as_text=True)
    assert "想一起骑车？" in home_open_html
    assert "申请加入社团" in about_open_html

    with app.app_context():
        set_membership_application_enabled(False)
        db.session.commit()

    home_closed = client.get("/")
    about_closed = client.get("/about")
    home_closed_html = home_closed.get_data(as_text=True)
    about_closed_html = about_closed.get_data(as_text=True)
    assert "想一起骑车？" not in home_closed_html
    assert "入社申请暂时关闭，请稍后关注网站通知。" in about_closed_html
    assert "申请加入社团" not in about_closed_html
def test_anonymous_join_post_creates_pending_application_and_audit(app_and_client):
    app, client = app_and_client
    resp = _post_join(client, student_id="sid-join-01", follow_redirects=False)
    assert resp.status_code == 302
    assert "/join/success" in (resp.headers.get("Location") or "")

    with app.app_context():
        application = MembershipApplication.query.filter_by(student_id="SID-JOIN-01").first()
        assert application is not None
        assert application.member_user_id is None
        assert application.status == "pending"
        assert application.form_version == CURRENT_MEMBERSHIP_APPLICATION_FORM_VERSION
        assert application.reviewed_at is None
        assert application.reviewed_by is None
        assert application.review_note is None
        assert application.approved_profile_id is None
        assert application.phone == "+86 13800000001"
        audit = AuditLog.query.filter_by(action="membership_application.submit").first()
        assert audit is not None
        assert audit.actor_id is None
        assert audit.target_type == "membership_application"
        assert audit.target_id == str(application.id)
        detail = json.loads(audit.detail)
        assert detail == {
            "actor_type": "anonymous",
            "submission_type": "public",
            "form_version": CURRENT_MEMBERSHIP_APPLICATION_FORM_VERSION,
        }
        assert application.full_name not in audit.detail
        assert application.phone not in audit.detail
        assert "周末骑行" not in audit.detail


def test_join_post_success_page_has_no_personal_data_in_url(app_and_client):
    _app, client = app_and_client
    resp = _post_join(client, student_id="20261203", full_name="No Url Person")
    assert resp.status_code == 302
    location = resp.headers.get("Location") or ""
    assert location.endswith("/join/success")
    assert "20261203" not in location
    assert "No" not in location


def test_logged_in_join_post_binds_member_and_ignores_forged_student_id(app_and_client):
    app, client = app_and_client
    assert register_member(client, "20261204", "Bound Rider", "memberpass123").status_code == 200
    resp = _post_join(client, student_id="FORGED999", full_name="Bound Person")
    assert resp.status_code == 302

    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20261204").first()
        application = MembershipApplication.query.filter_by(full_name="Bound Person").first()
        assert application is not None
        assert application.student_id == "20261204"
        assert application.member_user_id == member.id
        audit = AuditLog.query.filter_by(
            action="membership_application.submit",
            target_id=str(application.id),
        ).first()
        detail = json.loads(audit.detail)
        assert detail["actor_type"] == "member_user"
        assert detail["submission_type"] == "authenticated"
        assert detail["member_user_id"] == member.id


def test_anonymous_submission_does_not_auto_bind_existing_member_user(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            MemberUser(
                student_id="20261205",
                nickname="Existing Account",
                password_hash=generate_password_hash("memberpass123"),
                account_status=MEMBER_ACCOUNT_ACTIVE,
            )
        )
        db.session.commit()

    resp = _post_join(client, student_id="20261205")
    assert resp.status_code == 302
    with app.app_context():
        application = MembershipApplication.query.filter_by(student_id="20261205").first()
        assert application is not None
        assert application.member_user_id is None


def test_join_blocks_existing_formal_member_profile_and_approved_application(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(MemberProfile(student_id="20261206", full_name="Formal Person"))
        db.session.add(_membership_application(student_id="20261207", status="approved"))
        db.session.commit()

    profile_resp = _post_join(client, student_id="20261206")
    approved_resp = _post_join(client, student_id="20261207")
    assert profile_resp.status_code == 409
    assert "该学号已经存在正式社员档案，无需重复申请。如资料有误，请联系管理人员。" in profile_resp.get_data(as_text=True)
    assert approved_resp.status_code == 409
    assert "该学号已经存在正式社员档案，无需重复申请。如资料有误，请联系管理人员。" in approved_resp.get_data(as_text=True)


def test_join_get_blocks_logged_in_account_bound_to_profile(app_and_client):
    app, client = app_and_client
    assert register_member(client, "20261208", "Formal Rider", "memberpass123").status_code == 200
    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20261208").first()
        db.session.add(MemberProfile(member_user_id=member.id, student_id="20261208", full_name="Formal Rider"))
        db.session.commit()

    resp = client.get("/join")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "该学号已经存在正式社员档案，无需重复申请。如资料有误，请联系管理人员。" in html
    assert "提交申请" not in html


def test_join_blocks_pending_by_student_id_and_member_user(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(_membership_application(student_id="20261209", status="pending"))
        db.session.commit()
    student_resp = _post_join(client, student_id="20261209")
    assert student_resp.status_code == 409
    assert "你已有一份待审核的入社申请，请勿重复提交。" in student_resp.get_data(as_text=True)

    assert register_member(client, "20261210", "Pending Rider", "memberpass123").status_code == 200
    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20261210").first()
        db.session.add(_membership_application(member_user_id=member.id, student_id="DIFFERENT10", status="pending"))
        db.session.commit()
    account_page = client.get("/member/account")
    token = _extract_csrf(account_page.get_data(as_text=True))
    data = _join_form_data()
    data["csrf_token"] = token
    member_resp = client.post("/join", data=data)
    assert member_resp.status_code == 409
    assert "你已有一份待审核的入社申请，请勿重复提交。" in member_resp.get_data(as_text=True)


def test_rejected_application_history_allows_new_submission(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(_membership_application(student_id="20261211", status="rejected"))
        db.session.commit()

    get_resp = client.get("/join")
    assert get_resp.status_code == 200
    post_resp = _post_join(client, student_id="20261211", full_name="Second Try")
    assert post_resp.status_code == 302
    with app.app_context():
        applications = MembershipApplication.query.filter_by(student_id="20261211").all()
        assert len(applications) == 2
        assert sorted(item.status for item in applications) == ["pending", "rejected"]


@pytest.mark.parametrize(
    ("overrides", "message"),
    [
        ({"student_id": ""}, "请填写学号。"),
        ({"full_name": "   "}, "请填写真实姓名。"),
        ({"gender": "unknown"}, "性别不在允许范围内。"),
        ({"entry_year": "year"}, "入学年份需为数字"),
        ({"school": "Unknown School"}, "学院不在允许范围内。"),
        ({"college": "Unknown College"}, "书院不在允许范围内。"),
        ({"phone": ""}, "请填写手机号或常用联系电话。"),
        ({"competition_interest": "maybe"}, "请选择参赛意愿。"),
        ({"cycling_experience": "daily"}, "请选择骑行经验。"),
        ({"bicycle_status": "spaceship"}, "请选择车辆情况。"),
        ({"confirm_info": ""}, "请先确认提交的信息真实有效。"),
    ],
)
def test_join_validation_errors_keep_form_and_reject_bad_values(app_and_client, overrides, message):
    _app, client = app_and_client
    data = dict(overrides)
    if "student_id" not in data:
        data["student_id"] = f"VAL{abs(hash(message)) % 1000000:06d}"
    resp = _post_join(client, **data)
    html = resp.get_data(as_text=True)
    assert resp.status_code == 400
    assert message in html
    if data.get("student_id"):
        assert data["student_id"] in html


def test_join_requires_other_bicycle_description_only_for_other_status(app_and_client):
    app, client = app_and_client
    missing = _post_join(client, student_id="20261212", bicycle_status="other_bicycle", other_bicycle_description="")
    assert missing.status_code == 400
    assert "请补充说明自行车类型" in missing.get_data(as_text=True)

    accepted = _post_join(
        client,
        student_id="20261213",
        bicycle_status="road_bike",
        other_bicycle_description="forged hidden value",
    )
    assert accepted.status_code == 302
    with app.app_context():
        application = MembershipApplication.query.filter_by(student_id="20261213").first()
        assert application.other_bicycle_description is None


def test_join_accepts_international_phone_and_limits_note_length(app_and_client):
    app, client = app_and_client
    accepted = _post_join(client, student_id="20261214", phone="+852 9123 4567")
    assert accepted.status_code == 302
    too_long = _post_join(client, student_id="20261215", additional_note="x" * 1001)
    assert too_long.status_code == 400
    assert "补充说明不能超过 1000 个字符" in too_long.get_data(as_text=True)

    with app.app_context():
        application = MembershipApplication.query.filter_by(student_id="20261214").first()
        assert application.phone == "+852 9123 4567"


def test_join_requires_csrf_token(app_and_client):
    _app, client = app_and_client
    resp = client.post("/join", data=_join_form_data())
    assert resp.status_code == 400


def test_join_rate_limit_uses_ip_and_student_id(app_and_client):
    _app, client = app_and_client
    for index in range(5):
        resp = _post_join(client, student_id="20261216", full_name=f"Bad Try {index}", confirm_info="")
        assert resp.status_code == 400
    limited = _post_join(client, student_id="20261216", full_name="Bad Try Limited", confirm_info="")
    assert limited.status_code == 429
    assert "提交过于频繁" in limited.get_data(as_text=True)


def test_membership_application_creation_rolls_back_when_audit_fails(app_and_client, monkeypatch):
    app, _client = app_and_client

    def fail_audit(*_args, **_kwargs):
        raise RuntimeError("audit failed")

    monkeypatch.setattr("app.services.add_audit_log", fail_audit)
    with app.app_context():
        from app.services import create_membership_application

        audit_count_before = AuditLog.query.filter_by(
            action="membership_application.submit",
            target_type="membership_application",
        ).count()
        with pytest.raises(RuntimeError):
            create_membership_application(_join_form_data(student_id="20261217"))

        assert MembershipApplication.query.filter_by(student_id="20261217").count() == 0
        assert (
            AuditLog.query.filter_by(
                action="membership_application.submit",
                target_type="membership_application",
            ).count()
            == audit_count_before
        )


def test_anonymous_join_success_stores_short_lived_link_context(app_and_client):
    app, client = app_and_client
    resp = _post_join(client, student_id="20261701")
    assert resp.status_code == 302
    with client.session_transaction() as sess:
        context = sess.get("membership_application_link_context")
        assert context is not None
        assert isinstance(context.get("application_id"), int)
        assert isinstance(context.get("token"), str)
        assert len(context["token"]) >= 32
        assert isinstance(context.get("created_at"), int)
    with app.app_context():
        application = MembershipApplication.query.filter_by(student_id="20261701").first()
        assert application.member_user_id is None


def test_register_after_anonymous_join_links_matching_application(app_and_client):
    app, client = app_and_client
    assert _post_join(client, student_id="20261702").status_code == 302
    resp = _register_member_with_next(client, "20261702", "Linked Register", "memberpass123")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "刚提交的入社申请已关联到当前账号" in html

    with client.session_transaction() as sess:
        assert "membership_application_link_context" not in sess
    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20261702").first()
        application = MembershipApplication.query.filter_by(student_id="20261702").first()
        assert application.member_user_id == member.id
        log = AuditLog.query.filter_by(
            action="membership_application.link_account",
            target_type="membership_application",
            target_id=str(application.id),
        ).first()
        assert log is not None
        detail = json.loads(log.detail)
        assert detail == {
            "actor_type": "member_user",
            "member_user_id": member.id,
            "link_method": "post_submission_register",
        }
        assert "token" not in log.detail
        assert application.full_name not in log.detail
        assert application.phone not in log.detail


def test_login_after_anonymous_join_links_matching_application(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            MemberUser(
                student_id="20261703",
                nickname="Existing Link Login",
                password_hash=generate_password_hash("memberpass123"),
                account_status=MEMBER_ACCOUNT_ACTIVE,
            )
        )
        db.session.commit()

    assert _post_join(client, student_id="20261703").status_code == 302
    resp = _login_member_with_next(client, "20261703", "memberpass123")
    assert resp.status_code == 200
    assert "刚提交的入社申请已关联到当前账号" in resp.get_data(as_text=True)
    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20261703").first()
        application = MembershipApplication.query.filter_by(student_id="20261703").first()
        assert application.member_user_id == member.id
        log = AuditLog.query.filter_by(action="membership_application.link_account", target_id=str(application.id)).first()
        assert json.loads(log.detail)["link_method"] == "post_submission_login"


def test_post_submission_link_requires_exact_student_id_and_does_not_break_register(app_and_client):
    app, client = app_and_client
    assert _post_join(client, student_id="20261704").status_code == 302
    resp = _register_member_with_next(client, "20261799", "Mismatch Register", "memberpass123")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "不匹配" in html
    with client.session_transaction() as sess:
        assert "membership_application_link_context" not in sess
    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20261799").first()
        application = MembershipApplication.query.filter_by(student_id="20261704").first()
        assert member is not None
        assert application.member_user_id is None


def test_post_submission_link_rejects_already_bound_and_expired_or_forged_context(app_and_client):
    app, client = app_and_client
    with app.app_context():
        owner = MemberUser(
            student_id="20261705",
            nickname="Owner",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        db.session.add(owner)
        db.session.flush()
        owner_id = owner.id
        bound_application = _membership_application(
            member_user_id=owner_id,
            student_id="20261705",
            full_name="Already Bound Link",
            status="pending",
        )
        expired_application = _membership_application(
            student_id="20261706",
            full_name="Expired Link",
            status="pending",
        )
        db.session.add_all([bound_application, expired_application])
        db.session.commit()
        bound_id = bound_application.id
        expired_id = expired_application.id

    with client.session_transaction() as sess:
        sess["membership_application_link_context"] = {
            "application_id": bound_id,
            "token": "x" * 48,
            "created_at": int(datetime.now(timezone.utc).timestamp()),
        }
    assert login_member(client, "20261705", "memberpass123").status_code == 200
    with app.app_context():
        assert db.session.get(MembershipApplication, bound_id).member_user_id == owner_id

    with client.session_transaction() as sess:
        sess.pop("member_user_id", None)
        sess["membership_application_link_context"] = {
            "application_id": expired_id,
            "token": "y" * 48,
            "created_at": int((datetime.now(timezone.utc) - timedelta(hours=2)).timestamp()),
        }
    with app.app_context():
        db.session.add(
            MemberUser(
                student_id="20261706",
                nickname="Expired Login",
                password_hash=generate_password_hash("memberpass123"),
                account_status=MEMBER_ACCOUNT_ACTIVE,
            )
        )
        db.session.commit()
    assert _login_member_with_next(client, "20261706", "memberpass123").status_code == 200
    with app.app_context():
        assert db.session.get(MembershipApplication, expired_id).member_user_id is None

    with client.session_transaction() as sess:
        sess.pop("member_user_id", None)
        sess["membership_application_link_context"] = {"application_id": "bad", "token": "", "created_at": "bad"}
    assert _register_member_with_next(client, "20261707", "Forged Context", "memberpass123").status_code == 200
    with client.session_transaction() as sess:
        assert "membership_application_link_context" not in sess


def test_link_failure_does_not_roll_back_login(app_and_client, monkeypatch):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            MemberUser(
                student_id="20261708",
                nickname="Link Failure Login",
                password_hash=generate_password_hash("memberpass123"),
                account_status=MEMBER_ACCOUNT_ACTIVE,
            )
        )
        db.session.commit()
    assert _post_join(client, student_id="20261708").status_code == 302

    def fail_audit(*_args, **_kwargs):
        raise RuntimeError("audit failed")

    monkeypatch.setattr("app.routes_web.add_audit_log", fail_audit)
    resp = _login_member_with_next(client, "20261708", "memberpass123")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "关联未完成" in html
    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20261708").first()
        application = MembershipApplication.query.filter_by(student_id="20261708").first()
        assert member.last_login_at is not None
        assert application.member_user_id is None


def test_member_applications_page_requires_login_and_shows_only_current_user_history(app_and_client):
    app, client = app_and_client
    with app.app_context():
        first_member = MemberUser(
            student_id="20261709",
            nickname="History One",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        other_member = MemberUser(
            student_id="20261710",
            nickname="History Two",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        db.session.add_all([first_member, other_member])
        db.session.flush()
        db.session.add_all(
            [
                _membership_application(
                    member_user_id=first_member.id,
                    student_id="20261709",
                    full_name="Old Pending History",
                    status="pending",
                    submitted_at=datetime(2026, 7, 10, 1, 0, tzinfo=timezone.utc),
                ),
                _membership_application(
                    member_user_id=first_member.id,
                    student_id="20261709",
                    full_name="New Approved History",
                    status="approved",
                    reviewed_at=datetime(2026, 7, 12, 1, 0, tzinfo=timezone.utc),
                    review_note="欢迎加入",
                    submitted_at=datetime(2026, 7, 12, 1, 0, tzinfo=timezone.utc),
                ),
                _membership_application(
                    member_user_id=first_member.id,
                    student_id="20261709",
                    full_name="Rejected History",
                    status="rejected",
                    reviewed_at=datetime(2026, 7, 11, 1, 0, tzinfo=timezone.utc),
                    review_note="请补充资料",
                    submitted_at=datetime(2026, 7, 11, 1, 0, tzinfo=timezone.utc),
                ),
                _membership_application(
                    member_user_id=other_member.id,
                    student_id="20261710",
                    full_name="Other User History",
                    status="pending",
                    submitted_at=datetime(2026, 7, 13, 1, 0, tzinfo=timezone.utc),
                ),
            ]
        )
        db.session.commit()

    anonymous = client.get("/account/membership-applications", follow_redirects=False)
    assert anonymous.status_code == 302
    assert "/member/login" in (anonymous.headers.get("Location") or "")
    assert login_member(client, "20261709", "memberpass123").status_code == 200
    page = client.get("/account/membership-applications")
    html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert "New Approved History" not in html
    assert "已同意" in html
    assert "待审核" in html
    assert "已拒绝" in html
    assert "欢迎加入" in html
    assert "请补充资料" in html
    assert "Other User History" not in html
    assert html.index("已同意") < html.index("已拒绝") < html.index("待审核")


def test_logged_in_join_application_appears_in_member_history(app_and_client):
    app, client = app_and_client
    assert register_member(client, "20261711", "Logged Join", "memberpass123").status_code == 200
    resp = _post_join(client, student_id="FORGED", full_name="Logged Join Application")
    assert resp.status_code == 302
    page = client.get("/account/membership-applications")
    html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert "待审核" in html
    with app.app_context():
        member = MemberUser.query.filter_by(student_id="20261711").first()
        application = MembershipApplication.query.filter_by(full_name="Logged Join Application").first()
        assert application.member_user_id == member.id


def _create_membership_reader(username: str = "application_reader") -> None:
    reader = User(
        username=username,
        password=generate_password_hash("reader123456789"),
        role=ROLE_VIEWER,
        is_active=True,
    )
    db.session.add(reader)
    db.session.flush()
    grant_page_permission(reader, PAGE_MEMBERS, PERMISSION_READ)


def _create_membership_writer(username: str = "application_writer") -> None:
    writer = User(
        username=username,
        password=generate_password_hash("writer123456789"),
        role=ROLE_VIEWER,
        is_active=True,
    )
    db.session.add(writer)
    db.session.flush()
    grant_page_permission(writer, PAGE_MEMBERS, PERMISSION_WRITE)


def _create_membership_admin(username: str = "application_admin") -> None:
    admin = User(
        username=username,
        password=generate_password_hash("admin123456789"),
        role=ROLE_VIEWER,
        is_active=True,
    )
    db.session.add(admin)
    db.session.flush()
    grant_page_permission(admin, PAGE_MEMBERS, PERMISSION_ADMIN)


def test_manage_membership_applications_requires_backend_login_and_permission(app_and_client):
    app, client = app_and_client
    resp = client.get("/manage/membership-applications", follow_redirects=False)
    assert resp.status_code == 302
    assert "/manage/login" in (resp.headers.get("Location") or "")

    assert register_member(client, "20261300", "Front Member", "memberpass123").status_code == 200
    member_resp = client.get("/manage/membership-applications", follow_redirects=False)
    assert member_resp.status_code == 302
    assert "/manage/login" in (member_resp.headers.get("Location") or "")

    with app.app_context():
        blocked = User(
            username="application_blocked",
            password=generate_password_hash("blocked123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(blocked)
        db.session.flush()
        grant_page_permission(blocked, PAGE_MEMBERS, PERMISSION_NONE)
        db.session.commit()

    assert login(client, "application_blocked", "blocked123456789").status_code == 200
    blocked_page = client.get("/manage/membership-applications")
    assert blocked_page.status_code in {302, 403}
    assert b"Page Applicant" not in blocked_page.data


def test_manage_membership_applications_read_permission_lists_and_details(app_and_client):
    app, client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20261301",
            nickname="Bound Applicant",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        profile = MemberProfile(student_id="20261301", full_name="Approved Profile")
        reviewer = User(
            username="application_review_user",
            password=generate_password_hash("reviewer123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add_all([member, profile, reviewer])
        db.session.flush()
        application = _membership_application(
            member_user_id=member.id,
            approved_profile_id=profile.id,
            reviewed_by=reviewer.id,
            student_id="20261301",
            full_name="Readable Applicant",
            gender="男",
            entry_year=2026,
            school="SSE",
            college="shaw",
            competition_interest="unsure",
            cycling_experience="casual",
            bicycle_status="road_bike",
            additional_note="希望参加周末骑行。",
            status="approved",
            reviewed_at=datetime(2026, 7, 14, 8, 30, tzinfo=timezone.utc),
            review_note="资料通过",
        )
        db.session.add(application)
        _create_membership_reader()
        db.session.commit()
        application_id = application.id

    assert login(client, "application_reader", "reader123456789").status_code == 200
    with app.app_context():
        audit_count_before = AuditLog.query.count()
    list_resp = client.get("/manage/membership-applications?status=all")
    assert list_resp.status_code == 200
    list_html = list_resp.get_data(as_text=True)
    assert "Readable Applicant" in list_html
    assert "还不确定" in list_html
    assert "偶尔休闲骑行" in list_html
    assert "有公路车" in list_html
    assert "已绑定" in list_html

    detail_resp = client.get(f"/manage/membership-applications/{application_id}?status=all")
    detail_html = detail_resp.get_data(as_text=True)
    assert detail_resp.status_code == 200
    assert "Readable Applicant" in detail_html
    assert "理工学院" in detail_html
    assert "逸夫书院" in detail_html
    assert "希望参加周末骑行。" in detail_html
    assert "资料通过" in detail_html
    assert "application_review_user" in detail_html
    assert "Approved Profile" in detail_html
    assert "password_hash" not in detail_html
    with app.app_context():
        assert AuditLog.query.count() == audit_count_before


def test_manage_membership_applications_default_filters_pending_and_orders_desc(app_and_client):
    app, client = app_and_client
    with app.app_context():
        old_pending = _membership_application(
            student_id="20261302",
            full_name="Old Pending Applicant",
            status="pending",
            submitted_at=datetime(2026, 7, 10, 1, 0, tzinfo=timezone.utc),
        )
        new_pending = _membership_application(
            student_id="20261303",
            full_name="New Pending Applicant",
            status="pending",
            submitted_at=datetime(2026, 7, 12, 1, 0, tzinfo=timezone.utc),
        )
        approved = _membership_application(
            student_id="20261304",
            full_name="Approved Hidden Applicant",
            status="approved",
            submitted_at=datetime(2026, 7, 13, 1, 0, tzinfo=timezone.utc),
        )
        db.session.add_all([old_pending, new_pending, approved])
        _create_membership_reader("application_order_reader")
        db.session.commit()

    assert login(client, "application_order_reader", "reader123456789").status_code == 200
    resp = client.get("/manage/membership-applications?q=Pending Applicant")
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "New Pending Applicant" in html
    assert "Old Pending Applicant" in html
    assert "Approved Hidden Applicant" not in html
    assert html.index("New Pending Applicant") < html.index("Old Pending Applicant")


def test_manage_membership_applications_status_search_and_date_filters(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add_all(
            [
                _membership_application(
                    student_id="20261305",
                    full_name="Pending Searchable",
                    status="pending",
                    submitted_at=datetime(2026, 7, 11, 4, 0, tzinfo=timezone.utc),
                ),
                _membership_application(
                    student_id="20261306",
                    full_name="Rejected Searchable",
                    status="rejected",
                    submitted_at=datetime(2026, 7, 12, 4, 0, tzinfo=timezone.utc),
                ),
                _membership_application(
                    student_id="20261307",
                    full_name="Approved Searchable",
                    status="approved",
                    submitted_at=datetime(2026, 7, 13, 4, 0, tzinfo=timezone.utc),
                ),
            ]
        )
        _create_membership_reader("application_filter_reader")
        db.session.commit()

    assert login(client, "application_filter_reader", "reader123456789").status_code == 200
    all_resp = client.get("/manage/membership-applications?status=all&q=Searchable")
    all_html = all_resp.get_data(as_text=True)
    assert "Pending Searchable" in all_html
    assert "Rejected Searchable" in all_html
    assert "Approved Searchable" in all_html

    rejected_resp = client.get("/manage/membership-applications?status=rejected&q=Searchable")
    rejected_html = rejected_resp.get_data(as_text=True)
    assert "Rejected Searchable" in rejected_html
    assert "Pending Searchable" not in rejected_html

    approved_resp = client.get("/manage/membership-applications?status=approved&q=Searchable")
    approved_html = approved_resp.get_data(as_text=True)
    assert "Approved Searchable" in approved_html
    assert "Pending Searchable" not in approved_html

    name_resp = client.get("/manage/membership-applications?status=all&q=Approved")
    name_html = name_resp.get_data(as_text=True)
    assert "Approved Searchable" in name_html
    assert "Rejected Searchable" not in name_html

    sid_resp = client.get("/manage/membership-applications?status=all&q=20261306")
    sid_html = sid_resp.get_data(as_text=True)
    assert "Rejected Searchable" in sid_html
    assert "Approved Searchable" not in sid_html

    date_resp = client.get("/manage/membership-applications?status=all&q=Searchable&start_date=2026-07-12&end_date=2026-07-12")
    date_html = date_resp.get_data(as_text=True)
    assert "Rejected Searchable" in date_html
    assert "Pending Searchable" not in date_html
    assert "Approved Searchable" not in date_html


def test_manage_membership_applications_invalid_date_range_and_missing_detail(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            _membership_application(
                student_id="20261308",
                full_name="Date Range Applicant",
                status="pending",
                submitted_at=datetime(2026, 7, 12, 4, 0, tzinfo=timezone.utc),
            )
        )
        _create_membership_reader("application_date_reader")
        db.session.commit()

    assert login(client, "application_date_reader", "reader123456789").status_code == 200
    invalid = client.get("/manage/membership-applications?start_date=2026-07-13&end_date=2026-07-12")
    invalid_html = invalid.get_data(as_text=True)
    assert invalid.status_code == 200
    assert "开始日期不能晚于结束日期。" in invalid_html
    assert "Date Range Applicant" not in invalid_html

    missing = client.get("/manage/membership-applications/999999")
    assert missing.status_code == 404


def test_manage_membership_applications_pagination_preserves_filters_and_nav_count(app_and_client):
    app, client = app_and_client
    with app.app_context():
        for index in range(21):
            db.session.add(
                _membership_application(
                    student_id=f"202614{index:02d}",
                    full_name=f"Page Applicant {index:02d}",
                    status="pending",
                    submitted_at=datetime(2026, 7, 14, 0, index, tzinfo=timezone.utc),
                )
            )
        db.session.add(_membership_application(student_id="20261500", full_name="Rejected Nav Applicant", status="rejected"))
        _create_membership_reader("application_page_reader")
        db.session.commit()
        pending_count = MembershipApplication.query.filter_by(status="pending").count()

    assert login(client, "application_page_reader", "reader123456789").status_code == 200
    page = client.get("/manage/membership-applications?status=pending&q=Page&page=1")
    html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert f"入社申请（{pending_count}）" in html
    assert "page=2" in html
    assert "q=Page" in html
    assert "status=pending" in html
    assert "Rejected Nav Applicant" not in html


def test_manage_membership_applications_nav_hidden_without_permission(app_and_client):
    app, client = app_and_client
    with app.app_context():
        blocked = User(
            username="application_nav_blocked",
            password=generate_password_hash("blocked123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(blocked)
        db.session.flush()
        grant_page_permission(blocked, PAGE_MEMBERS, PERMISSION_NONE)
        grant_page_permission(blocked, PAGE_ANALYTICS, PERMISSION_READ)
        db.session.commit()

    assert login(client, "application_nav_blocked", "blocked123456789").status_code == 200
    page = client.get("/manage/analytics")
    html = page.get_data(as_text=True)
    assert page.status_code == 200
    assert "入社申请" not in html


def test_membership_application_review_buttons_require_write_permission(app_and_client):
    app, client = app_and_client
    with app.app_context():
        application = _membership_application(student_id="20261601", full_name="Readonly Review", status="pending")
        db.session.add(application)
        _create_membership_reader("application_review_reader")
        db.session.commit()
        application_id = application.id

    assert login(client, "application_review_reader", "reader123456789").status_code == 200
    detail = client.get(f"/manage/membership-applications/{application_id}")
    html = detail.get_data(as_text=True)
    assert detail.status_code == 200
    assert "你的权限为只读" in html
    assert "同意申请" not in html
    token = get_manage_csrf(client)
    approve_resp = client.post(
        f"/manage/membership-applications/{application_id}/approve",
        data={"csrf_token": token},
        follow_redirects=False,
    )
    reject_resp = client.post(
        f"/manage/membership-applications/{application_id}/reject",
        data={"csrf_token": token},
        follow_redirects=False,
    )
    assert approve_resp.status_code in {302, 403}
    assert reject_resp.status_code in {302, 403}
    with app.app_context():
        saved = db.session.get(MembershipApplication, application_id)
        assert saved.status == "pending"
        assert MemberProfile.query.filter_by(student_id="20261601").count() == 0


def test_membership_application_approve_creates_profile_and_bound_account(app_and_client):
    app, client = app_and_client
    submitted_at = datetime(2026, 7, 11, 10, 15, tzinfo=timezone.utc)
    with app.app_context():
        member = MemberUser(
            student_id="20261602",
            nickname="Approve Bound",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        db.session.add(member)
        db.session.flush()
        application = _membership_application(
            member_user_id=member.id,
            student_id="20261602",
            full_name="Approved Bound Applicant",
            gender="女",
            entry_year=2026,
            school="SSE",
            college="shaw",
            phone="+86 13800000002",
            competition_interest="yes",
            cycling_experience="competition",
            bicycle_status="road_bike",
            other_bicycle_description="should stay on application",
            additional_note="private application note",
            status="pending",
            submitted_at=submitted_at,
        )
        db.session.add(application)
        _create_membership_writer()
        db.session.commit()
        application_id = application.id
        member_id = member.id

    assert login(client, "application_writer", "writer123456789").status_code == 200
    detail = client.get(f"/manage/membership-applications/{application_id}")
    detail_html = detail.get_data(as_text=True)
    assert "将绑定申请提交时使用的社员账号" in detail_html
    token = _extract_csrf(detail_html)
    resp = client.post(
        f"/manage/membership-applications/{application_id}/approve",
        data={"csrf_token": token, "review_note": "资料完整"},
        follow_redirects=True,
    )
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "入社申请已同意" in html
    assert "查看档案" in html

    with app.app_context():
        application = db.session.get(MembershipApplication, application_id)
        profile = MemberProfile.query.filter_by(student_id="20261602").first()
        assert profile is not None
        assert profile.member_user_id == member_id
        assert profile.full_name == "Approved Bound Applicant"
        assert profile.gender == "女"
        assert profile.entry_year == 2026
        assert profile.school == "SSE"
        assert profile.college == "shaw"
        assert profile.phone == "+86 13800000002"
        assert profile.last_confirmed_at == submitted_at.date()
        assert not hasattr(profile, "competition_interest")
        assert application.status == "approved"
        assert application.reviewed_by is not None
        assert application.reviewed_at is not None
        assert application.review_note == "资料完整"
        assert application.approved_profile_id == profile.id
        assert application.competition_interest == "yes"
        assert application.cycling_experience == "competition"
        assert application.bicycle_status == "road_bike"
        assert application.additional_note == "private application note"
        approve_log = AuditLog.query.filter_by(
            action="membership_application.approve",
            target_type="membership_application",
            target_id=str(application_id),
        ).first()
        assert approve_log is not None
        approve_detail = json.loads(approve_log.detail)
        assert approve_detail["profile_id"] == profile.id
        assert approve_detail["member_user_id"] == member_id
        assert approve_detail["review_note_present"] is True
        assert approve_detail["auto_matched_member_user"] is False
        create_log = AuditLog.query.filter_by(action="member_profile.create", target_id=str(profile.id)).first()
        bind_log = AuditLog.query.filter_by(action="member_profile.account_bind", target_id=str(profile.id)).first()
        assert create_log is not None
        assert bind_log is not None
        assert profile.phone not in create_log.detail
        assert "private application note" not in approve_log.detail


def test_membership_application_approve_auto_matches_unbound_same_student_account(app_and_client):
    app, client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20261603",
            nickname="Auto Match",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        application = _membership_application(
            student_id="20261603",
            full_name="Auto Match Applicant",
            member_user_id=None,
            status="pending",
        )
        db.session.add_all([member, application])
        _create_membership_writer("application_auto_writer")
        db.session.commit()
        application_id = application.id
        member_id = member.id

    assert login(client, "application_auto_writer", "writer123456789").status_code == 200
    page = client.get(f"/manage/membership-applications/{application_id}")
    assert "将自动匹配并绑定同学号社员账号" in page.get_data(as_text=True)
    token = _extract_csrf(page.get_data(as_text=True))
    resp = client.post(
        f"/manage/membership-applications/{application_id}/approve",
        data={"csrf_token": token},
        follow_redirects=True,
    )
    assert resp.status_code == 200
    with app.app_context():
        profile = MemberProfile.query.filter_by(student_id="20261603").first()
        assert profile.member_user_id == member_id
        application = db.session.get(MembershipApplication, application_id)
        log = AuditLog.query.filter_by(action="membership_application.approve", target_id=str(application_id)).first()
        detail = json.loads(log.detail)
        assert detail["auto_matched_member_user"] is True
        assert application.approved_profile_id == profile.id


def test_membership_application_approve_blocks_profile_and_account_conflicts(app_and_client):
    app, client = app_and_client
    with app.app_context():
        existing = MemberProfile(student_id="20261604", full_name="Existing Profile")
        bound_member = MemberUser(
            student_id="20261605",
            nickname="Already Bound",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        mismatched_member = MemberUser(
            student_id="DIFFERENT61606",
            nickname="Mismatch",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        db.session.add_all([existing, bound_member, mismatched_member])
        db.session.flush()
        db.session.add(MemberProfile(member_user_id=bound_member.id, student_id="OTHER61605", full_name="Other Bound"))
        profile_conflict = _membership_application(student_id="20261604", full_name="Profile Conflict", status="pending")
        bound_conflict = _membership_application(student_id="20261605", full_name="Bound Conflict", status="pending")
        mismatch_conflict = _membership_application(
            member_user_id=mismatched_member.id,
            student_id="20261606",
            full_name="Mismatch Conflict",
            status="pending",
        )
        db.session.add_all([profile_conflict, bound_conflict, mismatch_conflict])
        _create_membership_writer("application_conflict_writer")
        db.session.commit()
        ids = [profile_conflict.id, bound_conflict.id, mismatch_conflict.id]

    assert login(client, "application_conflict_writer", "writer123456789").status_code == 200
    expected_messages = ["已经存在社员档案", "已经关联其他社员档案", "学号与申请学号不一致"]
    for application_id, message in zip(ids, expected_messages):
        token = _extract_csrf(client.get(f"/manage/membership-applications/{application_id}").get_data(as_text=True))
        resp = client.post(
            f"/manage/membership-applications/{application_id}/approve",
            data={"csrf_token": token},
            follow_redirects=True,
        )
        html = resp.get_data(as_text=True)
        assert resp.status_code == 200
        assert message in html
    with app.app_context():
        for application_id in ids:
            assert db.session.get(MembershipApplication, application_id).status == "pending"
        assert MemberProfile.query.filter(MemberProfile.full_name.in_(["Profile Conflict", "Bound Conflict", "Mismatch Conflict"])).count() == 0


def test_membership_application_reject_does_not_create_profile_and_updates_nav_count(app_and_client):
    app, client = app_and_client
    with app.app_context():
        application = _membership_application(student_id="20261607", full_name="Rejected Applicant", status="pending")
        db.session.add(application)
        _create_membership_writer("application_reject_writer")
        db.session.commit()
        application_id = application.id

    assert login(client, "application_reject_writer", "writer123456789").status_code == 200
    page = client.get(f"/manage/membership-applications/{application_id}")
    token = _extract_csrf(page.get_data(as_text=True))
    resp = client.post(
        f"/manage/membership-applications/{application_id}/reject",
        data={"csrf_token": token, "review_note": "暂不通过"},
        follow_redirects=True,
    )
    html = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "入社申请已拒绝" in html
    assert "同意申请" not in html
    assert "入社申请（1）" not in html

    with app.app_context():
        application = db.session.get(MembershipApplication, application_id)
        assert application.status == "rejected"
        assert application.reviewed_by is not None
        assert application.reviewed_at is not None
        assert application.review_note == "暂不通过"
        assert application.approved_profile_id is None
        assert MemberProfile.query.filter_by(student_id="20261607").count() == 0
        reject_log = AuditLog.query.filter_by(
            action="membership_application.reject",
            target_type="membership_application",
            target_id=str(application_id),
        ).first()
        assert reject_log is not None
        detail = json.loads(reject_log.detail)
        assert detail["after_status"] == "rejected"
        assert detail["review_note_present"] is True


def test_membership_application_processed_records_cannot_be_reviewed_again(app_and_client):
    app, client = app_and_client
    with app.app_context():
        approved = _membership_application(student_id="20261608", full_name="Already Approved", status="approved")
        rejected = _membership_application(student_id="20261609", full_name="Already Rejected", status="rejected")
        db.session.add_all([approved, rejected])
        _create_membership_writer("application_repeat_writer")
        db.session.commit()
        approved_id = approved.id
        rejected_id = rejected.id

    assert login(client, "application_repeat_writer", "writer123456789").status_code == 200
    for application_id, route_suffix in ((approved_id, "approve"), (rejected_id, "reject")):
        token = get_manage_csrf(client)
        resp = client.post(
            f"/manage/membership-applications/{application_id}/{route_suffix}",
            data={"csrf_token": token},
            follow_redirects=True,
        )
        assert resp.status_code == 200
        assert "该申请已处理" in resp.get_data(as_text=True)


def test_membership_application_review_requires_csrf(app_and_client):
    app, client = app_and_client
    with app.app_context():
        application = _membership_application(student_id="20261610", full_name="CSRF Applicant", status="pending")
        db.session.add(application)
        _create_membership_writer("application_csrf_writer")
        db.session.commit()
        application_id = application.id

    assert login(client, "application_csrf_writer", "writer123456789").status_code == 200
    resp = client.post(f"/manage/membership-applications/{application_id}/approve", data={})
    assert resp.status_code == 400


def test_membership_application_approval_rolls_back_on_audit_failure(app_and_client, monkeypatch):
    app, _client = app_and_client
    with app.app_context():
        application = _membership_application(student_id="20261611", full_name="Rollback Approval", status="pending")
        writer = User(
            username="rollback_writer",
            password=generate_password_hash("writer123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add_all([application, writer])
        db.session.commit()
        application_id = application.id
        writer_id = writer.id
        audit_count_before = AuditLog.query.count()

    def fail_audit(*_args, **_kwargs):
        raise RuntimeError("audit failed")

    monkeypatch.setattr("app.services.add_audit_log", fail_audit)
    with app.app_context():
        from app.services import approve_membership_application

        with pytest.raises(RuntimeError):
            approve_membership_application(application_id, writer_id)
        application = db.session.get(MembershipApplication, application_id)
        assert application.status == "pending"
        assert application.approved_profile_id is None
        assert MemberProfile.query.filter_by(student_id="20261611").count() == 0
        assert AuditLog.query.count() == audit_count_before


def test_manage_membership_applications_export_and_delete_require_permissions(app_and_client):
    app, client = app_and_client
    with app.app_context():
        _create_membership_reader("application_export_reader")
        _create_membership_writer("application_export_writer")
        _create_membership_admin("application_export_admin")
        db.session.add_all(
            [
                _membership_application(student_id="20261750", full_name="Exported Pending", status="pending"),
                _membership_application(student_id="20261751", full_name="Exported Approved", status="approved"),
            ]
        )
        db.session.commit()

    assert login(client, "application_export_reader", "reader123456789").status_code == 200
    list_page = client.get("/manage/membership-applications")
    assert list_page.status_code == 200
    assert "导出当前条件" not in list_page.get_data(as_text=True)
    export_forbidden = client.get("/manage/membership-applications/export.xlsx", follow_redirects=False)
    assert export_forbidden.status_code in {302, 403}
    _clear_manage_session(client)
    assert login(client, "application_export_writer", "writer123456789").status_code == 200
    write_list_page = client.get("/manage/membership-applications")
    assert "导出当前条件" in write_list_page.get_data(as_text=True)


def test_manage_membership_applications_export_filtering_and_excel_content(app_and_client):
    app, client = app_and_client
    base_local_now = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=8)))
    today = base_local_now.date()
    recent_date = (today - timedelta(days=1)).isoformat()
    with app.app_context():
        db.session.add_all(
            [
                _membership_application(
                    student_id="20261752",
                    full_name="Very Old Pending",
                    status="pending",
                    submitted_at=datetime.combine(today - timedelta(days=40), time(10, 0), tzinfo=timezone(timedelta(hours=8))).astimezone(
                        timezone.utc
                    ),
                ),
                _membership_application(
                    student_id="20261753",
                    full_name="Recent Pending",
                    status="pending",
                    submitted_at=datetime.combine(today - timedelta(days=1), time(10, 0), tzinfo=timezone(timedelta(hours=8))).astimezone(
                        timezone.utc
                    ),
                ),
                _membership_application(
                    student_id="20261754",
                    full_name="Recent Approved",
                    status="approved",
                    submitted_at=datetime.combine(today - timedelta(days=1), time(12, 0), tzinfo=timezone(timedelta(hours=8))).astimezone(
                        timezone.utc
                    ),
                ),
            ]
        )
        _create_membership_writer("application_filter_export_writer")
        db.session.commit()

    assert login(client, "application_filter_export_writer", "writer123456789").status_code == 200
    default_export = client.get("/manage/membership-applications/export.xlsx")
    assert default_export.status_code == 200
    assert (
        default_export.headers["Content-Type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    default_workbook = __import__("openpyxl").load_workbook(BytesIO(default_export.data), data_only=True)
    default_sheet = default_workbook.active
    headers = list(default_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    expected_headers = [
        "申请编号",
        "学号",
        "真实姓名",
        "性别",
        "入学年份",
        "学院",
        "书院",
        "手机号",
        "比赛意愿",
        "骑行经验",
        "车辆状况",
        "其他车辆说明",
        "补充说明",
        "是否绑定账号",
        "申请状态",
        "表单版本",
        "提交时间",
        "审核时间",
        "审核管理员",
        "审核备注",
        "关联社员档案 ID",
    ]
    assert headers[: len(expected_headers)] == tuple(expected_headers)
    default_rows = list(default_sheet.iter_rows(min_row=2, values_only=True))
    ids = [row[0] for row in default_rows]
    with app.app_context():
        old_application = MembershipApplication.query.filter_by(student_id="20261752").first()
        recent_pending = MembershipApplication.query.filter_by(student_id="20261753").first()
        assert old_application is not None and recent_pending is not None
        assert old_application.id not in ids
        assert recent_pending.id in ids

    filtered = client.get(
        f"/manage/membership-applications/export.xlsx?status=approved&start_date={recent_date}&end_date={recent_date}"
    )
    assert filtered.status_code == 200
    filtered_sheet = __import__("openpyxl").load_workbook(BytesIO(filtered.data), data_only=True).active
    filtered_headers = list(filtered_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    assert filtered_headers[0] == "申请编号"
    filtered_rows = list(filtered_sheet.iter_rows(min_row=2, values_only=True))
    with app.app_context():
        approved_application = MembershipApplication.query.filter_by(student_id="20261754").first()
        filtered_start = datetime.combine(datetime.fromisoformat(recent_date), time(0, 0), tzinfo=timezone(timedelta(hours=8))).astimezone(
            timezone.utc
        )
        filtered_end = filtered_start + timedelta(days=1)
        expected_filtered_ids = [
            application.id
            for application in (
                MembershipApplication.query.filter(
                    MembershipApplication.status == "approved",
                    MembershipApplication.submitted_at >= filtered_start,
                    MembershipApplication.submitted_at < filtered_end,
                )
                .order_by(MembershipApplication.submitted_at.desc(), MembershipApplication.id.desc())
                .all()
            )
        ]
        assert approved_application is not None
        filtered_ids = [row[0] for row in filtered_rows]
        assert approved_application.id in filtered_ids
        assert set(filtered_ids) == set(expected_filtered_ids)


def test_manage_membership_applications_export_audit_log_and_confirmation(app_and_client):
    app, client = app_and_client
    with app.app_context():
        _create_membership_writer("application_export_audit_writer")
        admin = db.session.query(User).filter_by(username="application_export_audit_writer").first()
        assert admin is not None
        admin_id = admin.id
        audit_before = AuditLog.query.count()
        db.session.add(
            _membership_application(
                student_id="20261755",
                full_name="Audit Logged Applicant",
                status="pending",
                submitted_at=datetime.now(timezone.utc),
            )
        )
        db.session.commit()
    assert login(client, "application_export_audit_writer", "writer123456789").status_code == 200
    search = "Audit Logged"
    export = client.get(
        f"/manage/membership-applications/export.xlsx?status=pending&q={search}&start_date=2026-01-01&end_date=2030-01-01"
    )
    assert export.status_code == 200
    with app.app_context():
        log = (
            AuditLog.query.filter_by(action="membership_application.export", target_type="membership_application")
            .order_by(AuditLog.id.desc())
            .first()
        )
        assert log is not None
        assert log.actor_id == admin_id
        assert log.target_id is None
        detail = json.loads(log.detail)
        assert detail["status"] == "pending"
        assert detail["search_keyword"] == search
        assert detail["start_date"] == "2026-01-01"
        assert detail["end_date"] == "2030-01-01"
        assert detail["count"] >= 1
        assert AuditLog.query.count() > audit_before


def test_manage_membership_applications_delete_requires_admin_and_confirmation(app_and_client):
    app, client = app_and_client
    with app.app_context():
        application = _membership_application(
            student_id="20261760",
            full_name="Delete Readonly Applicant",
            status="pending",
        )
        _create_membership_reader("application_delete_reader")
        _create_membership_admin("application_delete_admin")
        db.session.add(application)
        db.session.commit()
        application_id = application.id

    assert login(client, "application_delete_reader", "reader123456789").status_code == 200
    detail = client.get(f"/manage/membership-applications/{application_id}")
    assert "删除申请" not in detail.get_data(as_text=True)
    token = get_manage_csrf(client)
    refuse = client.post(
        f"/manage/membership-applications/{application_id}/delete",
        data={"csrf_token": token, "confirm_value": str(application_id)},
        follow_redirects=False,
    )
    assert refuse.status_code in {302, 403}
    with app.app_context():
        assert db.session.get(MembershipApplication, application_id) is not None

    _clear_manage_session(client)
    assert login(client, "application_delete_admin", "admin123456789").status_code == 200
    wrong = client.post(
        f"/manage/membership-applications/{application_id}/delete",
        data={"csrf_token": get_manage_csrf(client), "confirm_value": "wrong-value"},
        follow_redirects=True,
    )
    assert wrong.status_code == 200
    assert "确认内容不匹配" in wrong.get_data(as_text=True)
    with app.app_context():
        assert db.session.get(MembershipApplication, application_id) is not None


def test_manage_membership_application_delete_removes_only_application_and_audit(app_and_client):
    app, client = app_and_client
    with app.app_context():
        member = MemberUser(
            student_id="20261761",
            nickname="Delete Bound Member",
            password_hash=generate_password_hash("memberpass123"),
            account_status=MEMBER_ACCOUNT_ACTIVE,
        )
        profile = MemberProfile(
            student_id="20261761",
            full_name="Bound Profile",
            school="SSE",
            phone="13800000061",
        )
        db.session.add_all([member, profile])
        db.session.flush()
        application = _membership_application(
            student_id="20261761",
            full_name="Delete Approved Applicant",
            status="approved",
            member_user_id=member.id,
            approved_profile_id=profile.id,
            additional_note="hidden note for delete",
            phone="+86 13800000061",
            reviewed_at=datetime(2026, 7, 14, 10, 30, tzinfo=timezone.utc),
            reviewed_by=1,
            review_note="approved earlier",
            form_version=1,
        )
        _create_membership_admin("application_delete_admin2")
        db.session.add(application)
        db.session.commit()
        application_id = application.id
        member_id = member.id
        profile_id = profile.id

    assert login(client, "application_delete_admin2", "admin123456789").status_code == 200
    delete_page = client.get(f"/manage/membership-applications/{application_id}")
    delete_token = _extract_csrf(delete_page.get_data(as_text=True))
    resp = client.post(
        f"/manage/membership-applications/{application_id}/delete",
        data={"csrf_token": delete_token, "confirm_value": str(application_id)},
        follow_redirects=True,
    )
    assert resp.status_code == 200
    with app.app_context():
        assert db.session.get(MembershipApplication, application_id) is None
        assert db.session.get(MemberUser, member_id) is not None
        assert db.session.get(MemberProfile, profile_id) is not None
        log = AuditLog.query.filter_by(
            action="membership_application.delete",
            target_type="membership_application",
            target_id=str(application_id),
        ).first()
        assert log is not None
        detail = json.loads(log.detail)
        assert detail["status"] == "approved"
        assert detail["approved_profile_id"] == profile_id
        assert detail["has_linked_account"] is True
        assert "13800000061" not in log.detail
        assert "hidden note for delete" not in log.detail


def test_manage_membership_application_delete_rolls_back_on_audit_failure(app_and_client, monkeypatch):
    app, client = app_and_client
    with app.app_context():
        _create_membership_admin("application_delete_fail_admin")
        application = _membership_application(student_id="20261762", full_name="Rollback Delete", status="pending")
        db.session.add(application)
        db.session.commit()
        application_id = application.id
        log_count_before = AuditLog.query.filter_by(
            action="membership_application.delete",
            target_type="membership_application",
            target_id=str(application_id),
        ).count()

    assert login(client, "application_delete_fail_admin", "admin123456789").status_code == 200
    token = get_manage_csrf(client)

    def fail_add_audit_log(*_args, **_kwargs):
        raise RuntimeError("delete audit failed")

    monkeypatch.setattr("app.routes_admin.add_audit_log", fail_add_audit_log)
    response = client.post(
        f"/manage/membership-applications/{application_id}/delete",
        data={"csrf_token": token, "confirm_value": str(application_id)},
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert "删除申请失败" in response.get_data(as_text=True)

    with app.app_context():
        assert db.session.get(MembershipApplication, application_id) is not None
        assert (
            AuditLog.query.filter_by(
                action="membership_application.delete",
                target_type="membership_application",
                target_id=str(application_id),
            ).count()
            == log_count_before
        )


def test_membership_application_migration_upgrade_downgrade(tmp_path, monkeypatch):
    db_path = tmp_path / "migration.db"
    monkeypatch.setenv("FLASK_ENV", "development")
    engine = create_engine(f"sqlite:///{db_path.as_posix()}")
    config = Config("alembic.ini")
    db.metadata.create_all(engine)
    MembershipApplication.__table__.drop(engine)

    with engine.begin() as connection:
        config.attributes["connection"] = connection
        command.stamp(config, "20260713_0007")
    with engine.begin() as connection:
        config.attributes["connection"] = connection
        command.upgrade(config, "20260714_0008")
    inspector = inspect(engine)
    assert "membership_applications" in inspector.get_table_names()
    columns = {column["name"]: column for column in inspector.get_columns("membership_applications")}
    expected_columns = {
        "id",
        "member_user_id",
        "student_id",
        "full_name",
        "gender",
        "entry_year",
        "school",
        "college",
        "phone",
        "competition_interest",
        "cycling_experience",
        "bicycle_status",
        "other_bicycle_description",
        "additional_note",
        "status",
        "form_version",
        "submitted_at",
        "reviewed_at",
        "reviewed_by",
        "review_note",
        "approved_profile_id",
        "created_at",
        "updated_at",
    }
    assert set(columns) == expected_columns
    assert columns["student_id"]["nullable"] is False
    assert columns["member_user_id"]["nullable"] is True
    indexes = {index["name"]: tuple(index["column_names"]) for index in inspector.get_indexes("membership_applications")}
    assert indexes["idx_membership_applications_student_id"] == ("student_id",)
    assert indexes["idx_membership_applications_member_user_id"] == ("member_user_id",)
    assert indexes["idx_membership_applications_status_submitted_at"] == ("status", "submitted_at")
    foreign_keys = {
        tuple(foreign_key["constrained_columns"]): (
            foreign_key["referred_table"],
            tuple(foreign_key["referred_columns"]),
            foreign_key.get("options", {}).get("ondelete"),
        )
        for foreign_key in inspector.get_foreign_keys("membership_applications")
    }
    assert foreign_keys[("member_user_id",)] == ("member_users", ("id",), "SET NULL")
    assert foreign_keys[("reviewed_by",)] == ("users", ("id",), "SET NULL")
    assert foreign_keys[("approved_profile_id",)] == ("member_profiles", ("id",), "SET NULL")

    with engine.begin() as connection:
        config.attributes["connection"] = connection
        command.downgrade(config, "20260713_0007")
    assert "membership_applications" not in inspect(engine).get_table_names()

    with engine.begin() as connection:
        config.attributes["connection"] = connection
        command.upgrade(config, "head")
    assert "membership_applications" in inspect(engine).get_table_names()
    engine.dispose()


def test_unauth_manage_redirect_writes_audit_log(app_and_client):
    app, client = app_and_client
    resp = client.get("/manage", follow_redirects=False)
    assert resp.status_code == 302
    assert "/manage/login" in resp.headers.get("Location", "")

    with app.app_context():
        log = AuditLog.query.order_by(AuditLog.id.desc()).first()
        assert log is not None
        assert log.action == "auth.required_redirect"
        assert log.target_type == "admin"
        assert log.target_id == "/manage"
        assert '"path": "/manage"' in (log.detail or "")


def test_login_failed_writes_audit_log_with_ip(app_and_client):
    app, client = app_and_client
    login_page = client.get("/manage/login")
    token = _extract_csrf(login_page.get_data(as_text=True))
    resp = client.post(
        "/manage/login",
        data={"username": "admin", "password": "wrong-password", "csrf_token": token},
        headers={"X-Forwarded-For": "203.0.113.10"},
        follow_redirects=False,
    )
    assert resp.status_code == 302

    with app.app_context():
        log = (
            AuditLog.query.filter_by(action="auth.login_failed")
            .order_by(AuditLog.id.desc())
            .first()
        )
        assert log is not None
        assert "ip=203.0.113.10" in (log.detail or "")


def test_admin_can_change_own_password(app_and_client):
    app, client = app_and_client
    resp = login_admin(client)
    assert resp.status_code == 200

    page = client.get("/manage/account/password")
    assert page.status_code == 200
    html = page.get_data(as_text=True)
    assert "修改密码" in html

    token = _extract_csrf(html)
    bad_resp = client.post(
        "/manage/account/password",
        data={
            "csrf_token": token,
            "current_password": "wrong-password",
            "new_password": "new-admin-password",
            "confirm_password": "new-admin-password",
        },
        follow_redirects=True,
    )
    assert "当前密码不正确" in bad_resp.get_data(as_text=True)

    token = _extract_csrf(client.get("/manage/account/password").get_data(as_text=True))
    good_resp = client.post(
        "/manage/account/password",
        data={
            "csrf_token": token,
            "current_password": "admin123456789",
            "new_password": "new-admin-password",
            "confirm_password": "new-admin-password",
        },
        follow_redirects=True,
    )
    assert good_resp.status_code == 200
    assert "密码已更新" in good_resp.get_data(as_text=True)

    with app.app_context():
        user = User.query.filter_by(username="admin").first()
        assert user is not None
        assert check_password_hash(user.password, "new-admin-password")
        log = AuditLog.query.filter_by(action="user.password_change").order_by(AuditLog.id.desc()).first()
        assert log is not None
        assert log.target_id == str(user.id)

    logout_token = get_manage_csrf(client)
    client.post("/manage/logout", data={"csrf_token": logout_token}, follow_redirects=True)
    new_login = login(client, "admin", "new-admin-password")
    assert new_login.status_code == 200
    assert "管理总览" in new_login.get_data(as_text=True)

    with app.app_context():
        user = User.query.filter_by(username="admin").first()
        assert user is not None
        user.password = generate_password_hash("admin123456789")
        db.session.commit()


def test_access_log_persisted_for_web_request(app_and_client):
    app, client = app_and_client
    resp = client.get("/")
    assert resp.status_code == 200

    with app.app_context():
        log = (
            AccessLog.query.filter_by(path="/", method="GET")
            .order_by(AccessLog.id.desc())
            .first()
        )
        assert log is not None
        assert log.status_code == 200


def test_access_log_prefers_cf_connecting_ip(app_and_client):
    app, client = app_and_client
    resp = client.get("/", headers={"CF-Connecting-IP": "198.51.100.8", "X-Forwarded-For": "203.0.113.77"})
    assert resp.status_code == 200

    with app.app_context():
        log = (
            AccessLog.query.filter_by(path="/", method="GET")
            .order_by(AccessLog.id.desc())
            .first()
        )
        assert log is not None
        assert log.ip_address == "198.51.100.8"


def test_site_feedback_records_forwarded_ip(app_and_client):
    app, client = app_and_client
    resp = client.post(
        "/feedback",
        data={
            "category": "bug",
            "content": "这里有一个页面展示问题，麻烦排查。",
            "contact": "123456789",
            "source": "/",
        },
        headers={"X-Forwarded-For": "203.0.113.66"},
        follow_redirects=False,
    )
    assert resp.status_code == 302

    with app.app_context():
        entry = SiteFeedback.query.order_by(SiteFeedback.id.desc()).first()
        assert entry is not None
        assert entry.ip_address == "203.0.113.66"


def test_robots_txt_served(app_and_client):
    _app, client = app_and_client
    resp = client.get("/robots.txt")
    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    assert "User-agent: *" in body
    assert "Disallow: /manage/" in body
    assert "Sitemap:" in body


def test_sitemap_xml_served(app_and_client):
    _app, client = app_and_client
    resp = client.get("/sitemap.xml")
    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    assert "<urlset" in body
    assert "/about" in body
    assert "/events" in body
    assert "/routes" in body


def test_v41_static_pages_available(app_and_client):
    _app, client = app_and_client
    home = client.get("/")
    assert home.status_code == 200
    home_body = home.get_data(as_text=True)
    assert "骑行服预定" in home_body
    assert "/kit" in home_body
    assert client.get("/about").status_code == 200
    assert client.get("/team").status_code == 404
    assert client.get("/contact").status_code == 200
    assert client.get("/routes").status_code == 200


def test_events_alias_pages_available(app_and_client):
    app, client = app_and_client
    with app.app_context():
        activity = Activity(title="V4.1 Test Event")
        db.session.add(activity)
        db.session.commit()
        activity_id = activity.id

    list_resp = client.get("/events")
    assert list_resp.status_code == 200
    detail_resp = client.get(f"/events/{activity_id}")
    assert detail_resp.status_code == 200


def test_kit_preorder_submit_update_lookup_and_cancel(app_and_client):
    app, client = app_and_client
    deadline = datetime.now(timezone.utc) + timedelta(days=7)
    with app.app_context():
        batch = MerchPreorderBatch(
            title="Kit Batch A",
            status=MERCH_BATCH_ACTIVE,
            deadline_at=deadline,
            description="This overview description should stay hidden",
            price_min=180,
            price_max=220,
            is_visible=True,
        )
        db.session.add(batch)
        db.session.commit()
        batch_id = batch.id

    detail = client.get(f"/kit/{batch_id}")
    assert detail.status_code == 200
    assert "Kit Batch A" in detail.get_data(as_text=True)
    list_resp = client.get("/kit")
    assert list_resp.status_code == 200
    assert "This overview description should stay hidden" not in list_resp.get_data(as_text=True)

    submit_resp = client.post(
        f"/kit/{batch_id}/submit",
        data={
            "name": "Alice",
            "student_id": "123456",
            "phone": "13800000000",
            "gender": "女",
            "size": "L",
            "quantity": "1",
            "notes": "first",
        },
        follow_redirects=False,
    )
    assert submit_resp.status_code == 302
    assert f"/kit/{batch_id}/success" in submit_resp.headers.get("Location", "")

    check_resp = client.get(f"/kit/{batch_id}/check-student?student_id=123456")
    assert check_resp.status_code == 200
    assert check_resp.get_json()["exists"] is True
    registration_id = check_resp.get_json()["registration_id"]

    update_resp = client.post(
        f"/kit/{batch_id}/submit",
        data={
            "name": "Alice",
            "student_id": "123456",
            "phone": "13800000001",
            "gender": "女",
            "size": "XL",
            "quantity": "2",
            "notes": "updated",
            "update_registration_id": str(registration_id),
        },
        follow_redirects=False,
    )
    assert update_resp.status_code == 302

    with app.app_context():
        rows = MerchPreorderRegistration.query.filter_by(batch_id=batch_id).all()
        assert len(rows) == 1
        assert rows[0].size == "XL"
        assert rows[0].quantity == 2

    lookup = client.get(f"/kit/{batch_id}/lookup?name=Alice&student_id=123456")
    assert lookup.status_code == 200
    body = lookup.get_data(as_text=True)
    assert "XL" in body
    assert "取消预报名" in body

    cancel_resp = client.post(
        f"/kit/{batch_id}/cancel",
        data={"name": "Alice", "student_id": "123456"},
        follow_redirects=False,
    )
    assert cancel_resp.status_code == 302
    with app.app_context():
        row = MerchPreorderRegistration.query.filter_by(batch_id=batch_id).first()
        assert row is not None
        assert row.status == MERCH_ORDER_CANCELLED


def test_manage_kit_preorder_create_and_registrations_page(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    manage_page = client.get("/manage")
    assert manage_page.status_code == 200
    assert f"当前版本：{_project_version()}" in manage_page.get_data(as_text=True)
    new_page = client.get("/manage/kit-preorders/new")
    assert new_page.status_code == 200
    new_page_body = new_page.get_data(as_text=True)
    assert 'name="status"' not in new_page_body
    assert 'name="start_date"' in new_page_body
    assert 'name="deadline_date"' in new_page_body

    csrf_token = get_manage_csrf(client)
    resp = client.post(
        "/manage/kit-preorders/create",
        data={
            "csrf_token": csrf_token,
            "title": "Managed Kit Batch",
            "start_date": "2026-12-01",
            "deadline_date": "2026-12-31",
            "price_min": "180",
            "price_max": "220",
            "price_note": "最终价格按人数确认",
            "size_note": "尺码偏小，建议买大一码。",
            "description": "后台创建测试",
            "is_visible": "1",
        },
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert "Managed Kit Batch" in resp.get_data(as_text=True)

    with app.app_context():
        batch = MerchPreorderBatch.query.filter_by(title="Managed Kit Batch").first()
        assert batch is not None
        assert batch.status == MERCH_BATCH_UPCOMING
        assert batch.start_at is not None
        assert batch.deadline_at is not None
        assert batch.deadline_at.minute == 59
        batch_id = batch.id
        db.session.add(
            MerchPreorderRegistration(
                batch_id=batch_id,
                name="Bob",
                student_id="654321",
                phone="13900000000",
                gender="男",
                size="M",
                quantity=1,
                status=MERCH_ORDER_PENDING,
            )
        )
        db.session.commit()

    list_resp = client.get("/manage/kit-preorders")
    assert list_resp.status_code == 200
    assert "Managed Kit Batch" in list_resp.get_data(as_text=True)

    registrations_resp = client.get(f"/manage/kit-preorders/{batch_id}/registrations")
    assert registrations_resp.status_code == 200
    text = registrations_resp.get_data(as_text=True)
    assert "Bob" in text
    assert "654321" in text


def test_ended_kit_preorder_only_allows_lookup(app_and_client):
    app, client = app_and_client
    deadline = datetime.now(timezone.utc) - timedelta(days=1)
    with app.app_context():
        batch = MerchPreorderBatch(
            title="Ended Kit Batch",
            status=MERCH_BATCH_ENDED,
            deadline_at=deadline,
            is_visible=True,
        )
        db.session.add(batch)
        db.session.flush()
        registration = MerchPreorderRegistration(
            batch_id=batch.id,
            name="Ended User",
            student_id="20260001",
            phone="13800000002",
            gender="男",
            size="L",
            quantity=1,
            status=MERCH_ORDER_PENDING,
        )
        db.session.add(registration)
        db.session.commit()
        batch_id = batch.id

    lookup = client.get(f"/kit/{batch_id}/lookup?name=Ended+User&student_id=20260001")
    body = lookup.get_data(as_text=True)
    assert lookup.status_code == 200
    assert "查询预报名" in body
    assert "取消预报名" not in body

    cancel_resp = client.post(
        f"/kit/{batch_id}/cancel",
        data={"name": "Ended User", "student_id": "20260001"},
        follow_redirects=False,
    )
    assert cancel_resp.status_code == 302
    with app.app_context():
        row = MerchPreorderRegistration.query.filter_by(batch_id=batch_id).first()
        assert row is not None
        assert row.status == MERCH_ORDER_PENDING


def test_global_kit_preorder_lookup_lists_all_visible_records(app_and_client):
    app, client = app_and_client
    active_deadline = datetime.now(timezone.utc) + timedelta(days=7)
    ended_deadline = datetime.now(timezone.utc) - timedelta(days=1)
    with app.app_context():
        active_batch = MerchPreorderBatch(
            title="Global Active Batch",
            status=MERCH_BATCH_ACTIVE,
            deadline_at=active_deadline,
            is_visible=True,
        )
        ended_batch = MerchPreorderBatch(
            title="Global Ended Batch",
            status=MERCH_BATCH_ENDED,
            deadline_at=ended_deadline,
            is_visible=True,
        )
        hidden_batch = MerchPreorderBatch(
            title="Global Hidden Batch",
            status=MERCH_BATCH_ACTIVE,
            deadline_at=active_deadline,
            is_visible=False,
        )
        db.session.add_all([active_batch, ended_batch, hidden_batch])
        db.session.flush()
        db.session.add_all(
            [
                MerchPreorderRegistration(
                    batch_id=active_batch.id,
                    name="Global User",
                    student_id="20260002",
                    phone="13800000003",
                    gender="女",
                    size="M",
                    quantity=1,
                    status=MERCH_ORDER_PENDING,
                ),
                MerchPreorderRegistration(
                    batch_id=ended_batch.id,
                    name="Global User",
                    student_id="20260002",
                    phone="13800000003",
                    gender="女",
                    size="L",
                    quantity=2,
                    status=MERCH_ORDER_PENDING,
                ),
                MerchPreorderRegistration(
                    batch_id=hidden_batch.id,
                    name="Global User",
                    student_id="20260002",
                    phone="13800000003",
                    gender="女",
                    size="XL",
                    quantity=1,
                    status=MERCH_ORDER_PENDING,
                ),
            ]
        )
        db.session.commit()
        active_batch_id = active_batch.id

    lookup = client.get("/kit/lookup?name=Global+User&student_id=20260002")
    body = lookup.get_data(as_text=True)
    assert lookup.status_code == 200
    assert "Global Active Batch" in body
    assert "Global Ended Batch" in body
    assert "Global Hidden Batch" not in body
    assert body.count("取消预报名") == 1

    cancel_resp = client.post(
        f"/kit/{active_batch_id}/cancel",
        data={"name": "Global User", "student_id": "20260002", "source": "global"},
        follow_redirects=False,
    )
    assert cancel_resp.status_code == 302
    assert "/kit/lookup" in cancel_resp.headers.get("Location", "")
    with app.app_context():
        active_row = MerchPreorderRegistration.query.filter_by(batch_id=active_batch_id).first()
        assert active_row is not None
        assert active_row.status == MERCH_ORDER_CANCELLED


def test_upcoming_kit_preorder_hides_lookup_entry(app_and_client):
    app, client = app_and_client
    start_at = datetime.now(timezone.utc) + timedelta(days=3)
    deadline = datetime.now(timezone.utc) + timedelta(days=7)
    with app.app_context():
        batch = MerchPreorderBatch(
            title="Upcoming Kit Batch",
            status=MERCH_BATCH_UPCOMING,
            start_at=start_at,
            deadline_at=deadline,
            is_visible=True,
        )
        db.session.add(batch)
        db.session.commit()
        batch_id = batch.id

    list_resp = client.get("/kit")
    assert list_resp.status_code == 200
    body = list_resp.get_data(as_text=True)
    assert "Upcoming Kit Batch" in body
    assert f"/kit/{batch_id}/lookup" not in body

    detail_resp = client.get(f"/kit/{batch_id}")
    assert detail_resp.status_code == 200
    detail_body = detail_resp.get_data(as_text=True)
    assert "查询/取消预报名" not in detail_body
    assert "查询预报名" not in detail_body


def test_manage_announcements_crud(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)
    with app.app_context():
        linked_route = Route(route_name="Announcement Route", gpx_filename="announcement_route.gpx", status="published")
        linked_activity = Activity(title="Announcement Activity")
        db.session.add_all([linked_route, linked_activity])
        db.session.commit()
        linked_route_id = linked_route.id
        linked_activity_id = linked_activity.id

    create_resp = client.post(
        "/manage/announcements/create",
        data={
            "csrf_token": csrf_token,
            "title": "Announcement A",
            "content": "公告内容 A",
            "status": "published",
            "is_pinned": "1",
            "sort_order": "8",
            "published_at": "2026-03-20T12:00",
            "offline_at": "2026-03-25T12:00",
            "activity_ids": [str(linked_activity_id)],
            "route_ids": [str(linked_route_id)],
        },
        follow_redirects=True,
    )
    assert create_resp.status_code == 200

    with app.app_context():
        item = Announcement.query.filter_by(title="Announcement A").first()
        assert item is not None
        assert item.is_pinned is True
        assert item.status == "published"
        assert len(item.activities) == 1
        assert len(item.routes) == 1
        announcement_id = item.id

    edit_page = client.get(f"/manage/announcements/{announcement_id}/edit")
    assert edit_page.status_code == 200
    edit_body = edit_page.get_data(as_text=True)
    assert "添加关联" in edit_body
    assert f"/events/{linked_activity_id}" in edit_body
    assert f"/routes/{linked_route_id}" in edit_body

    update_resp = client.post(
        f"/manage/announcements/{announcement_id}/update",
        data={
            "csrf_token": csrf_token,
            "title": "Announcement B",
            "content": "公告内容 B",
            "status": "offline",
            "is_pinned": "0",
            "sort_order": "2",
            "published_at": "",
            "offline_at": "",
            "activity_ids": [],
            "route_ids": [],
        },
        follow_redirects=True,
    )
    assert update_resp.status_code == 200

    with app.app_context():
        item = db.session.get(Announcement, announcement_id)
        assert item is not None
        assert item.title == "Announcement B"
        assert item.status == "offline"
        assert item.is_pinned is False
        assert len(item.activities) == 0
        assert len(item.routes) == 0

    list_resp = client.get("/manage/announcements")
    assert list_resp.status_code == 200
    assert "Announcement B" in list_resp.get_data(as_text=True)

    delete_resp = client.post(
        f"/manage/announcements/{announcement_id}/delete",
        data={"csrf_token": csrf_token},
        follow_redirects=True,
    )
    assert delete_resp.status_code == 200
    with app.app_context():
        assert db.session.get(Announcement, announcement_id) is None


def test_announcement_detail_visibility_and_associations(app_and_client):
    app, client = app_and_client
    now = datetime.now(timezone.utc)
    with app.app_context():
        route = Route(route_name="Ann Detail Route", gpx_filename="ann_detail_route.gpx", status="published")
        activity = Activity(title="Ann Detail Activity")
        announcement = Announcement(
            title="Ann Detail",
            content="公告详情正文\n[[/events/1|本次活动]]\n[[javascript:alert(1)|危险链接]]",
            status="published",
            published_at=now - timedelta(hours=1),
            offline_at=now + timedelta(days=1),
        )
        announcement.routes.append(route)
        announcement.activities.append(activity)
        db.session.add_all([route, activity, announcement])
        db.session.commit()
        announcement_id = announcement.id

    detail = client.get(f"/announcements/{announcement_id}")
    body = detail.get_data(as_text=True)
    assert detail.status_code == 200
    assert "Ann Detail" in body
    assert "Ann Detail Route" not in body
    assert "Ann Detail Activity" not in body
    assert 'href="/events/1"' in body
    assert "本次活动" in body
    assert "javascript:alert" in body
    assert 'href="javascript:alert(1)"' not in body


def test_announcement_schedule_blocks_future_or_expired(app_and_client):
    app, client = app_and_client
    now = datetime.now(timezone.utc)
    with app.app_context():
        future_item = Announcement(
            title="Future Ann",
            content="not yet",
            status="published",
            published_at=now + timedelta(days=1),
        )
        expired_item = Announcement(
            title="Expired Ann",
            content="expired",
            status="published",
            published_at=now - timedelta(days=2),
            offline_at=now - timedelta(days=1),
        )
        db.session.add_all([future_item, expired_item])
        db.session.commit()
        future_id = future_item.id
        expired_id = expired_item.id

    assert client.get(f"/announcements/{future_id}").status_code == 404
    assert client.get(f"/announcements/{expired_id}").status_code == 404
    index_body = client.get("/").get_data(as_text=True)
    assert "Future Ann" not in index_body
    assert "Expired Ann" not in index_body


def test_route_detail_back_link_from_activity_detail(app_and_client):
    app, client = app_and_client
    with app.app_context():
        route = Route(
            route_name="Back Link Route",
            gpx_filename="back_link_route.gpx",
            status="published",
            is_deleted=False,
        )
        activity = Activity(title="Back Link Event")
        activity.routes.append(route)
        db.session.add_all([route, activity])
        db.session.commit()
        route_id = route.id
        activity_id = activity.id

    resp = client.get(f"/routes/{route_id}?from_activity_id={activity_id}&from_detail=web.events_detail")
    body = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "返回活动详情" in body
    assert f"/events/{activity_id}" in body


def test_activity_route_options_saved_and_rendered(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)
    create_route(client, csrf_token, name="Tier Route A")
    create_route(client, csrf_token, name="Tier Route B")
    create_route(client, csrf_token, name="Tier Route C")
    routes = client.get("/api/v1/routes").get_json()["items"]
    route_ids = [item["id"] for item in routes[:3]]

    resp = client.post(
        "/manage/activities/create",
        data={
            "csrf_token": csrf_token,
            "title": "Tier Activity",
            "activity_date": "2026-03-21",
            "route_option_beginner": str(route_ids[0]),
            "route_option_beginner_time": "09:00",
            "route_option_beginner_participants": "12",
            "route_option_intermediate": str(route_ids[1]),
            "route_option_intermediate_time": "09:30",
            "route_option_intermediate_participants": "8",
            "route_option_advanced": str(route_ids[2]),
            "route_option_advanced_time": "10:00",
            "route_option_advanced_participants": "5",
        },
        follow_redirects=True,
    )
    assert resp.status_code == 200

    with app.app_context():
        activity = Activity.query.filter_by(title="Tier Activity").first()
        assert activity is not None
        options = ActivityRouteOption.query.filter_by(activity_id=activity.id).all()
        assert len(options) == 3
        assert len(activity.routes) >= 3
        assert sum(item.participant_count for item in options) == 25
        sh_tz = timezone(timedelta(hours=8))
        option_times = {
            item.level_key: item.activity_time.replace(tzinfo=timezone.utc).astimezone(sh_tz).strftime("%Y-%m-%d %H:%M")
            for item in options
        }
        assert option_times["beginner"] == "2026-03-21 09:00"
        assert option_times["intermediate"] == "2026-03-21 09:30"
        assert option_times["advanced"] == "2026-03-21 10:00"
        assert activity.activity_time.replace(tzinfo=timezone.utc).astimezone(sh_tz).strftime("%Y-%m-%d %H:%M") == "2026-03-21 09:00"
        activity_id = activity.id

    detail = client.get(f"/events/{activity_id}")
    body = detail.get_data(as_text=True)
    assert detail.status_code == 200
    assert "初级路线" in body
    assert "中级路线" in body
    assert "高级路线" in body
    assert "参与人数" in body


def test_event_signup_does_not_mutate_manual_participant_counts(app_and_client):
    app, client = app_and_client
    starts_at = datetime.now(timezone.utc) + timedelta(days=14)
    deadline = starts_at - timedelta(days=1)
    with app.app_context():
        route = Route(
            route_name="Signup Count Route",
            gpx_filename="signup-count-route.gpx",
            distance_km=20,
            status="published",
        )
        activity = Activity(
            title="Signup Count Activity",
            activity_time=starts_at,
            needs_registration=True,
            registration_deadline=deadline,
            registration_limit=10,
            participant_count=7,
        )
        db.session.add_all([route, activity])
        db.session.flush()
        option = ActivityRouteOption(
            activity_id=activity.id,
            route_id=route.id,
            level_key="beginner",
            level_label="初级",
            activity_time=starts_at,
            participant_count=7,
            sort_order=1,
        )
        activity.routes = [route]
        db.session.add(option)
        db.session.commit()
        activity_id = activity.id
        option_id = option.id

    resp = client.post(
        f"/events/{activity_id}/signup",
        data={
            "name": "Signup User",
            "student_id": "SID001",
            "option_id": str(option_id),
            "source": "events_detail",
            "consent_required": "1",
            "consent_image": "1",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 302

    with app.app_context():
        activity = db.session.get(Activity, activity_id)
        option = db.session.get(ActivityRouteOption, option_id)
        registration_count = EventRegistration.query.filter_by(activity_id=activity_id).count()
        assert activity is not None
        assert option is not None
        assert registration_count == 1
        assert activity.participant_count == 7
        assert option.participant_count == 7

    detail = client.get(f"/events/{activity_id}")
    body = detail.get_data(as_text=True)
    assert detail.status_code == 200
    assert "当前报名" in body
    assert ">1<" in body


def test_activity_detail_legacy_route_relation_compatible(app_and_client):
    app, client = app_and_client
    with app.app_context():
        route = Route(
            route_name="Legacy Activity Route",
            gpx_filename="legacy_activity_route.gpx",
            status="published",
            is_deleted=False,
            distance_km=12.3,
        )
        activity = Activity(title="Legacy Activity Detail")
        activity.routes.append(route)
        db.session.add_all([route, activity])
        db.session.commit()
        activity_id = activity.id

    resp = client.get(f"/events/{activity_id}")
    body = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "初级路线" in body
    assert "Legacy Activity Route" in body


def test_activity_media_upload_and_render(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)
    create_route(client, csrf_token, name="Media Route")
    route_id = client.get("/api/v1/routes").get_json()["items"][0]["id"]

    png_bytes = b"\\x89PNG\\r\\n\\x1a\\n\\x00\\x00\\x00\\rIHDR\\x00\\x00\\x00\\x01\\x00\\x00\\x00\\x01\\x08\\x02\\x00\\x00\\x00\\x90wS\\xde\\x00\\x00\\x00\\x0cIDATx\\x9cc``\\x00\\x00\\x00\\x04\\x00\\x01\\x0b\\xe7\\x02\\x9d\\x00\\x00\\x00\\x00IEND\\xaeB`\\x82"
    mp4_bytes = b"\\x00\\x00\\x00\\x18ftypmp42\\x00\\x00\\x00\\x00mp42isom"

    resp = client.post(
        "/manage/activities/create",
        data={
            "csrf_token": csrf_token,
            "title": "Media Event",
            "route_option_beginner": str(route_id),
            "route_option_beginner_time": "2026-03-20T10:00",
            "route_option_beginner_participants": "15",
            "media_files_beginner": [
                (BytesIO(png_bytes), "photo.png"),
                (BytesIO(mp4_bytes), "clip.mp4"),
            ],
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert resp.status_code == 200

    with app.app_context():
        activity = Activity.query.filter_by(title="Media Event").first()
        assert activity is not None
        assets = MediaAsset.query.filter_by(activity_id=activity.id).all()
        assert len(assets) >= 2
        first_asset_id = assets[0].id
        activity_id = activity.id

    media_resp = client.get(f"/media/{first_asset_id}")
    assert media_resp.status_code == 200

    detail_resp = client.get(f"/events/{activity_id}")
    body = detail_resp.get_data(as_text=True)
    assert detail_resp.status_code == 200
    assert "/media/" in body


def test_probe_wordpress_path_blocked_early(app_and_client):
    app, client = app_and_client
    resp = client.get("/wordpress/wp-admin/setup-config.php")
    assert resp.status_code == 404

    with app.app_context():
        log = (
            AccessLog.query.filter_by(path="/wordpress/wp-admin/setup-config.php")
            .order_by(AccessLog.id.desc())
            .first()
        )
        assert log is not None
        assert log.status_code == 404


def test_watchlist_probe_path_match():
    assert is_watchlist_probe_path("/wordpress/wp-admin/setup-config.php")
    assert is_watchlist_probe_path("/wp-login.php")
    assert not is_watchlist_probe_path("/routes/1")


def test_404_page_is_lightweight(app_and_client):
    _app, client = app_and_client
    resp = client.get("/not-found-anymore")
    text = resp.get_data(as_text=True)
    assert resp.status_code == 404
    assert "返回首页" in text
    assert "查找路线" not in text


def test_manage_analytics_page_available_after_login(app_and_client):
    _app, client = app_and_client
    assert login_admin(client).status_code == 200
    resp = client.get("/manage/analytics")
    assert resp.status_code == 200
    assert "流量统计" in resp.get_data(as_text=True)
    assert "近5分钟活跃(估算)" not in resp.get_data(as_text=True)


def test_manage_analytics_excludes_self_path(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            AccessLog(
                path="/manage/analytics",
                method="GET",
                endpoint="admin.analytics_page",
                status_code=200,
                ip_address="203.0.113.1",
                user_agent="pytest",
                referer="",
            )
        )
        db.session.add(
            AccessLog(
                path="/manage/routes",
                method="GET",
                endpoint="admin.routes_page",
                status_code=200,
                ip_address="203.0.113.9",
                user_agent="pytest",
                referer="",
            )
        )
        db.session.add(
            AccessLog(
                path="/",
                method="GET",
                endpoint="web.index",
                status_code=200,
                ip_address="203.0.113.2",
                user_agent="pytest",
                referer="",
            )
        )
        db.session.commit()

    assert login_admin(client).status_code == 200
    resp = client.get("/manage/analytics?days=1")
    text = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "<td>/manage/analytics</td>" not in text
    assert "<td>/manage/routes</td>" not in text
    assert "<td>/</td>" in text


def test_manage_analytics_supports_post_deploy_scope(app_and_client):
    app, client = app_and_client
    with app.app_context():
        baseline_utc = datetime.fromisoformat("2026-03-17T18:30:00+08:00").astimezone(timezone.utc)
        db.session.add(
            AccessLog(
                path="/",
                method="GET",
                endpoint="web.index",
                status_code=200,
                ip_address="203.0.113.3",
                user_agent="pytest",
                referer="",
                created_at=baseline_utc - timedelta(hours=1),
            )
        )
        db.session.add(
            AccessLog(
                path="/",
                method="GET",
                endpoint="web.index",
                status_code=200,
                ip_address="203.0.113.4",
                user_agent="pytest",
                referer="",
                created_at=baseline_utc + timedelta(hours=1),
            )
        )
        db.session.commit()

    assert login_admin(client).status_code == 200
    resp = client.get("/manage/analytics?scope=post_deploy")
    text = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "上线后（" in text
    assert "业务PV" in text


def test_manage_analytics_shows_probe_excluded_metrics(app_and_client):
    app, client = app_and_client
    with app.app_context():
        db.session.add(
            AccessLog(
                path="/",
                method="GET",
                endpoint="web.index",
                status_code=200,
                ip_address="203.0.113.20",
                user_agent="pytest",
                referer="",
            )
        )
        db.session.add(
            AccessLog(
                path="/wp-login.php",
                method="GET",
                endpoint="",
                status_code=404,
                ip_address="203.0.113.21",
                user_agent="pytest",
                referer="",
            )
        )
        db.session.commit()

    assert login_admin(client).status_code == 200
    resp = client.get("/manage/analytics?days=1")
    text = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "排除探测" in text
    assert "探测占比" in text


def test_manage_dashboard_shows_active_5m_metric(app_and_client):
    _app, client = app_and_client
    assert login_admin(client).status_code == 200
    resp = client.get("/manage")
    assert resp.status_code == 200
    assert "近5分钟活跃(估算)" in resp.get_data(as_text=True)


def test_manage_dashboard_shows_security_entry(app_and_client):
    _app, client = app_and_client
    assert login_admin(client).status_code == 200
    resp = client.get("/manage")
    assert resp.status_code == 200
    text = resp.get_data(as_text=True)
    assert "安全监控" in text
    assert f"当前版本：{_project_version()}" in text


def test_manage_security_page_available_after_login(app_and_client):
    _app, client = app_and_client
    assert login_admin(client).status_code == 200
    resp = client.get("/manage/security")
    assert resp.status_code == 200
    text = resp.get_data(as_text=True)
    assert "核心指标" in text
    assert "最近安全事件" in text


def test_manage_security_supports_post_deploy_scope(app_and_client):
    _app, client = app_and_client
    assert login_admin(client).status_code == 200
    resp = client.get("/manage/security?scope=post_deploy")
    text = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "核心指标" in text
    assert "上线后" in text


def test_manage_security_events_filter_and_pagination(app_and_client):
    _app, client = app_and_client
    assert login_admin(client).status_code == 200
    resp = client.get("/manage/security?event_type=watchlist&event_status=5xx&event_q=wp-admin&event_page=1")
    text = resp.get_data(as_text=True)
    assert resp.status_code == 200
    assert "事件类型" in text
    assert "状态码" in text
    assert "第 " in text


def test_admin_login_and_create_route(app_and_client):
    _app, client = app_and_client
    login_resp = login_admin(client)
    assert login_resp.status_code == 200

    csrf_token = get_manage_csrf(client)
    resp = create_route(client, csrf_token, name="Test Route")
    assert resp.status_code == 200

    list_resp = client.get("/api/v1/routes")
    items = list_resp.get_json()["items"]
    assert any(item["route_name"] == "Test Route" for item in items)


def test_route_distance_and_elevation_stats_auto_computed_from_gpx(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)

    # ~111m straight line with elevation gain.
    gpx_content = b"""<?xml version='1.0' encoding='UTF-8'?>
<gpx version='1.1' creator='pytest'>
  <trk><name>Auto Stats</name><trkseg>
    <trkpt lat='22.500000' lon='114.100000'><ele>10</ele></trkpt>
    <trkpt lat='22.501000' lon='114.100000'><ele>24</ele></trkpt>
  </trkseg></trk>
</gpx>"""
    resp = create_route(
        client,
        csrf_token,
        name="Auto Stats Route",
        distance_km="999.9",
        gpx_content=gpx_content,
    )
    assert resp.status_code == 200

    with app.app_context():
        route = Route.query.filter_by(route_name="Auto Stats Route").first()
        assert route is not None
        assert route.distance_km != 999.9
        assert route.distance_km > 0
        assert route.ascent_m is not None
        assert route.max_ele_m is not None


def test_recalculate_route_stats_endpoint_refreshes_saved_values(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)

    gpx_content = b"""<?xml version='1.0' encoding='UTF-8'?>
<gpx version='1.1' creator='pytest'>
  <trk><name>Recalc Stats</name><trkseg>
    <trkpt lat='22.500000' lon='114.100000'><ele>18</ele></trkpt>
    <trkpt lat='22.501000' lon='114.100000'><ele>31</ele></trkpt>
  </trkseg></trk>
</gpx>"""
    create_route(client, csrf_token, name="Recalc Stats Route", gpx_content=gpx_content)

    with app.app_context():
        route = Route.query.filter_by(route_name="Recalc Stats Route").first()
        assert route is not None
        route_id = route.id
        upload_dir = Path(app.config["UPLOAD_FOLDER"])
        original_path = upload_dir / route.gpx_filename
        fallback_name = "Recalc_Stats_Route.gpx"
        (upload_dir / fallback_name).write_bytes(original_path.read_bytes())
        route.gpx_filename = f"20260101000000_{fallback_name}"
        route.distance_km = 0.0
        route.ascent_m = 0.0
        route.descent_m = 0.0
        route.suggested_duration_hours = 99.0
        db.session.commit()

    csrf_token = get_manage_csrf(client)
    resp = client.post(
        f"/manage/routes/{route_id}/recalculate-stats",
        data={
            "csrf_token": csrf_token,
            "difficulty": "5",
            "manual_distance_km": "12.5",
            "manual_ascent_m": "100",
        },
        headers={"Accept": "application/json"},
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["ok"] is True
    assert payload["stats"]["distance_km"] > 0

    with app.app_context():
        route = db.session.get(Route, route_id)
        assert route is not None
        assert route.distance_km == 12.5
        assert route.ascent_m == 100
        assert '"distance_km": 12.5' in route.manual_stat_overrides
        expected_duration = round((12.5 / 25 + 100 / 600) * 1.15, 1)
        assert route.suggested_duration_hours == expected_duration


def test_route_preview_returns_waypoints_from_gpx(app_and_client):
    _app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)
    gpx_content = b"""<?xml version='1.0' encoding='UTF-8'?>
<gpx version='1.1' creator='pytest'>
  <wpt lat='22.7423' lon='114.2882'>
    <name>\xe7\xae\xa1\xe5\x88\xb6\xe5\x8c\xba\xe5\x9f\x9f</name>
    <cmt>\xe7\xbb\x95\xe8\xa1\x8c</cmt>
    <desc>\xe7\xbb\x95\xe8\xa1\x8c</desc>
    <type>risk:high</type>
    <sym>Restricted Area</sym>
  </wpt>
  <wpt lat='22.7424' lon='114.2884'>
    <name>\xe6\x8f\x90\xe9\x86\x92\xe7\x82\xb9</name>
    <desc>\xe6\xb3\xa8\xe6\x84\x8f\xe6\xa8\xaa\xe9\xa3\x8e</desc>
    <type>low</type>
  </wpt>
  <trk><trkseg>
    <trkpt lat='22.500000' lon='114.100000'><ele>18</ele></trkpt>
    <trkpt lat='22.501000' lon='114.100000'><ele>31</ele></trkpt>
  </trkseg></trk>
</gpx>"""
    create_route(client, csrf_token, name="Waypoint Route", gpx_content=gpx_content)
    route_id = client.get("/api/v1/routes").get_json()["items"][0]["id"]
    resp = client.get(f"/api/v1/routes/{route_id}/preview")
    assert resp.status_code == 200
    payload = resp.get_json()
    assert "waypoints" in payload
    assert len(payload["waypoints"]) == 2
    assert payload["waypoints"][0]["kind"] == "risk"
    assert payload["waypoints"][0]["risk_level"] == "high"
    assert payload["waypoints"][1]["risk_level"] == "low"


def test_download_updates_stats(app_and_client):
    _app, client = app_and_client
    assert login_admin(client).status_code == 200

    csrf_token = get_manage_csrf(client)
    create_route(client, csrf_token, name="Stats Route")

    list_resp = client.get("/api/v1/routes")
    route_id = list_resp.get_json()["items"][0]["id"]

    scanner_resp = client.get(f"/download/{route_id}")
    assert scanner_resp.status_code == 200
    detail_before = client.get(f"/api/v1/routes/{route_id}").get_json()
    assert detail_before["download_count"] in (0, None)

    tracked_resp = client.post(f"/download/{route_id}/track")
    assert tracked_resp.status_code == 200

    detail = client.get(f"/api/v1/routes/{route_id}").get_json()
    assert detail["download_count"] >= 1
    assert detail["last_downloaded_at"] is not None


def test_viewer_cannot_create_route(app_and_client):
    app, client = app_and_client
    with app.app_context():
        if not User.query.filter_by(username="viewer").first():
            viewer = User(
                username="viewer",
                password=generate_password_hash("viewer123456789"),
                role=ROLE_VIEWER,
                is_active=True,
            )
            db.session.add(viewer)
            db.session.commit()

    login(client, "viewer", "viewer123456789")
    csrf_token = get_manage_csrf(client)

    resp = create_route(client, csrf_token, name="Should Fail", follow_redirects=False)
    assert resp.status_code in (302, 403)

    check = client.get("/api/v1/routes").get_json()["items"]
    assert all(item["route_name"] != "Should Fail" for item in check)


def test_user_with_manage_users_permission_can_open_users_page(app_and_client):
    app, client = app_and_client
    with app.app_context():
        manager = User(
            username="manager",
            password=generate_password_hash("manager123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(manager)
        grant_page_permission(manager, PAGE_ACCOUNTS, PERMISSION_READ)
        grant_page_permission(manager, PAGE_ANALYTICS, PERMISSION_READ)
        db.session.commit()

    login(client, "manager", "manager123456789")
    resp = client.get("/manage/users")
    assert resp.status_code == 200
    assert "账号列表" in resp.get_data(as_text=True)


def test_user_create_without_custom_permissions_uses_role_preset(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)

    resp = client.post(
        "/manage/users/create",
        data={
            "csrf_token": csrf_token,
            "username": "preset_viewer",
            "password": "viewer123456789",
            "role": ROLE_VIEWER,
            "page_perm_routes": PERMISSION_ADMIN,
        },
        follow_redirects=True,
    )
    assert resp.status_code == 200

    with app.app_context():
        user = User.query.filter_by(username="preset_viewer").first()
        assert user is not None
        route_permission = UserPagePermission.query.filter_by(user_id=user.id, page_key=PAGE_ROUTES).first()
        assert route_permission is not None
        assert route_permission.permission_level == PERMISSION_READ


def test_user_without_analytics_permission_forbidden(app_and_client):
    app, client = app_and_client
    with app.app_context():
        blocked = User(
            username="blocked_user",
            password=generate_password_hash("blocked123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(blocked)
        grant_page_permission(blocked, PAGE_ANALYTICS, PERMISSION_NONE)
        db.session.commit()

    login(client, "blocked_user", "blocked123456789")
    resp = client.get("/manage/analytics")
    assert resp.status_code == 302
    assert "/manage/login" in (resp.headers.get("Location") or "")


def test_dashboard_hides_sections_without_permissions(app_and_client):
    app, client = app_and_client
    with app.app_context():
        viewer = User(
            username="limited_viewer",
            password=generate_password_hash("viewer123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(viewer)
        grant_page_permission(viewer, PAGE_ANALYTICS, PERMISSION_NONE)
        grant_page_permission(viewer, PAGE_SECURITY, PERMISSION_NONE)
        grant_page_permission(viewer, PAGE_FEEDBACK, PERMISSION_NONE)
        grant_page_permission(viewer, PAGE_ACCOUNTS, PERMISSION_NONE)
        grant_page_permission(viewer, PAGE_AUDIT_LOGS, PERMISSION_NONE)
        db.session.commit()

    login(client, "limited_viewer", "viewer123456789")
    resp = client.get("/manage")
    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    assert "流量统计" not in body
    assert "安全监控" not in body
    assert "最新审计日志" not in body
    assert "管理员账号" not in body


def test_user_without_edit_or_review_cannot_download_import_report(app_and_client):
    app, client = app_and_client
    with app.app_context():
        viewer = User(
            username="report_blocked",
            password=generate_password_hash("viewer123456789"),
            role=ROLE_VIEWER,
            is_active=True,
        )
        db.session.add(viewer)
        grant_page_permission(viewer, PAGE_ROUTES, PERMISSION_NONE)
        grant_page_permission(viewer, PAGE_FEEDBACK, PERMISSION_NONE)
        db.session.commit()

    login(client, "report_blocked", "viewer123456789")
    resp = client.get("/manage/import-report/any-token")
    assert resp.status_code == 302
    assert "/manage/login" in (resp.headers.get("Location") or "")


def test_bulk_import_invalid_csv_does_not_leave_uploaded_files(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200

    csrf_token = get_manage_csrf(client)
    upload_dir = app.config["UPLOAD_FOLDER"]
    before = {p.name for p in Path(upload_dir).glob("*.gpx")}

    bad_csv = "bad,header\n1,2\n".encode("utf-8")
    gpx_content = b"<?xml version='1.0'?><gpx version='1.1'></gpx>"
    resp = client.post(
        "/manage/bulk-import",
        data={
            "csrf_token": csrf_token,
            "csv_file": (BytesIO(bad_csv), "bad.csv"),
            "gpx_files": (BytesIO(gpx_content), "bulk_orphan.gpx"),
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert resp.status_code == 200

    after = {p.name for p in Path(upload_dir).glob("*.gpx")}
    assert after == before


def test_feedback_review_flow(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)
    create_route(client, csrf_token, name="Feedback Route")

    route_id = client.get("/api/v1/routes").get_json()["items"][0]["id"]

    feedback_resp = client.post(
        f"/api/v1/routes/{route_id}/feedback",
        json={
            "rating": 4,
            "comment": "路况良好",
            "road_condition_update": "中段有施工",
            "report_type": "changed",
        },
    )
    assert feedback_resp.status_code == 201
    feedback_id = feedback_resp.get_json()["id"]

    with app.app_context():
        reviewer = User(
            username="reviewer",
            password=generate_password_hash("reviewer123456789"),
            role=ROLE_OPS_ADMIN,
            is_active=True,
        )
        db.session.add(reviewer)
        grant_page_permission(reviewer, PAGE_FEEDBACK, PERMISSION_WRITE)
        db.session.commit()
    admin_csrf = get_manage_csrf(client)
    forged_resp = client.post(
        f"/api/v1/admin/feedback/{feedback_id}/review",
        headers={"X-Admin-User": "99999", "X-CSRF-Token": admin_csrf},
        json={"status": "approved", "reviewer_note": "forged"},
    )
    assert forged_resp.status_code == 403
    assert forged_resp.get_json()["error"] == "admin_user_mismatch"

    client.post("/manage/logout", data={"csrf_token": admin_csrf}, follow_redirects=True)
    assert login(client, "reviewer", "reviewer123456789").status_code == 200
    reviewer_csrf = get_manage_csrf(client)

    review_resp = client.post(
        f"/api/v1/admin/feedback/{feedback_id}/review",
        headers={"X-CSRF-Token": reviewer_csrf},
        json={"status": "approved", "reviewer_note": "已核实"},
    )
    assert review_resp.status_code == 200
    assert review_resp.get_json()["status"] == "approved"

    with app.app_context():
        feedback = db.session.get(RouteFeedback, feedback_id)
        assert feedback is not None
        assert feedback.status == "approved"


def test_route_version_rollback(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)
    create_route(client, csrf_token, name="Rollback Route")

    with app.app_context():
        route = Route.query.filter_by(route_name="Rollback Route").first()
        assert route is not None
        route_id = route.id

    csrf_token = get_manage_csrf(client)
    update_resp = client.post(
        f"/manage/routes/{route_id}/update",
        data={
            "csrf_token": csrf_token,
            "route_name": "Rollback Route v2",
            "distance_km": "8",
            "difficulty": "hard",
            "category": "hiking",
            "description": "updated",
            "status": "published",
            "suggested_duration_hours": "4",
            "supply_points": "none",
            "risk_warning": "heat",
        },
        follow_redirects=True,
    )
    assert update_resp.status_code == 200

    with app.app_context():
        versions = (
            RouteVersion.query.filter_by(route_id=route_id)
            .order_by(RouteVersion.version_no.asc())
            .all()
        )
        assert len(versions) >= 2
        first_version = versions[0].version_no

    csrf_token = get_manage_csrf(client)
    rollback_resp = client.post(
        f"/manage/routes/{route_id}/rollback",
        data={"csrf_token": csrf_token, "version_no": str(first_version)},
        follow_redirects=True,
    )
    assert rollback_resp.status_code == 200

    with app.app_context():
        route = db.session.get(Route, route_id)
        assert route is not None
        assert route.route_name == "Rollback Route"


def test_route_rollback_rejects_missing_gpx_snapshot(app_and_client):
    app, client = app_and_client
    assert login_admin(client).status_code == 200
    csrf_token = get_manage_csrf(client)
    create_route(client, csrf_token, name="Rollback Missing GPX")

    with app.app_context():
        route = Route.query.filter_by(route_name="Rollback Missing GPX").first()
        assert route is not None
        route_id = route.id
        first_version = (
            RouteVersion.query.filter_by(route_id=route_id)
            .order_by(RouteVersion.version_no.asc())
            .first()
        )
        assert first_version is not None
        payload = json.loads(first_version.snapshot_json)
        payload["gpx_filename"] = "not_exists_rollback.gpx"
        first_version.snapshot_json = json.dumps(payload, ensure_ascii=False)
        db.session.commit()
        old_name = route.route_name

    csrf_token = get_manage_csrf(client)
    rollback_resp = client.post(
        f"/manage/routes/{route_id}/rollback",
        data={"csrf_token": csrf_token, "version_no": "1"},
        follow_redirects=True,
    )
    assert rollback_resp.status_code == 200

    with app.app_context():
        route = db.session.get(Route, route_id)
        assert route is not None
        assert route.route_name == old_name

